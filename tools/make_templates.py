"""Blank starter templates for the two input files, with the exact headers,
dropdowns for controlled vocabularies, correct column formats, one example
row block (marked for deletion), and a bilingual README sheet each."""
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

HDR_FONT = Font(name="Arial", bold=True, size=10)
BODY_FONT = Font(name="Arial", size=10)
NOTE_FONT = Font(name="Arial", size=10, color="6B6B6B")
HDR_FILL = PatternFill("solid", fgColor="F0EEE9")
EX_FILL = PatternFill("solid", fgColor="FFF6DF")   # example rows — delete me
THIN = Side(style="thin", color="C9C6C0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
OUTPUT_DIR = (
    Path(sys.argv[1]).expanduser()
    if len(sys.argv) > 1
    else Path(__file__).resolve().parents[1] / "docs"
)
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)


def style_header(ws, row, headers, widths):
    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = HDR_FONT
        c.fill = HDR_FILL
        c.border = BORDER
        c.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = w


def readme_sheet(wb, lines):
    ws = wb.create_sheet("README 使用说明")
    ws.column_dimensions["A"].width = 118
    for i, (text, bold) in enumerate(lines, start=1):
        c = ws.cell(row=i, column=1, value=text)
        c.font = Font(name="Arial", size=10, bold=bold)
        c.alignment = Alignment(wrap_text=True, vertical="top")
    return ws


# ================================================================== PLOG ==
PLOG_HEADERS = ["NO", "MCN", "CAMPAIGN", "TYPE", "LEVEL", "NAME",
                "FAN BASE（K)", "POST DATE", "MICRO MACRO", "POST LINK",
                "IMPRESSION", "LIKE", "COLLECTION", "COMMENT",
                "TTL ENGAGEMENT", "PRICE", "CPM", "CPE"]
PLOG_WIDTHS = [5, 10, 16, 11, 8, 20, 11, 12, 12, 34, 11, 9, 11, 9, 13, 10, 8, 8]

wb = Workbook()
ws = wb.active
ws.title = "MASTER KOL LIST"
style_header(ws, 1, PLOG_HEADERS, PLOG_WIDTHS)
ws.freeze_panes = "A2"

examples = [
    [1, "ORANGE", "CAMPAIGN #001", "报备图文", "腰部", "示例博主A（删除此行）", 350,
     datetime(2026, 6, 25), "MICRO", "https://xhslink.com/o/EXAMPLE1",
     242757, 838, 143, 74, "=L2+M2+N2", 30300, None, None],
    [2, "ORANGE", None, "软植图文", "KOC", "示例博主B（删除此行）", 88,
     datetime(2026, 6, 28), "MICRO", "https://xhslink.com/o/EXAMPLE2",
     34211, 1037, 677, 18, "=L3+M3+N3", 8000, None, None],
]
for r, row in enumerate(examples, start=2):
    for cidx, v in enumerate(row, start=1):
        c = ws.cell(row=r, column=cidx, value=v)
        c.font = BODY_FONT
        c.fill = EX_FILL
        c.border = BORDER

for r in range(2, 502):
    ws.cell(row=r, column=8).number_format = "yyyy-mm-dd"        # POST DATE
    ws.cell(row=r, column=10).number_format = "@"                # POST LINK text
    for cidx in (11, 12, 13, 14, 15, 16):
        ws.cell(row=r, column=cidx).number_format = "#,##0"
    ws.cell(row=r, column=7).number_format = "0"                 # FAN BASE (K)

dv_type = DataValidation(type="list",
                         formula1='"报备图文,软植图文,报备视频,软植视频"',
                         allow_blank=True, showErrorMessage=True,
                         error="TYPE 必须以 报备 或 软植 开头")
dv_level = DataValidation(type="list", formula1='"头部,腰部,尾部,底部,KOC"',
                          allow_blank=True, showErrorMessage=True,
                          error="LEVEL 只能填 头部/腰部/尾部/底部/KOC")
dv_mm = DataValidation(type="list", formula1='"MICRO,MACRO"', allow_blank=True)
ws.add_data_validation(dv_type)
ws.add_data_validation(dv_level)
ws.add_data_validation(dv_mm)
dv_type.add("D2:D501")
dv_level.add("E2:E501")
dv_mm.add("I2:I501")

readme_sheet(wb, [
    ("PLOG tracker template · PLOG 投放追踪表模板", True),
    ("", False),
    ("• One row = one post. The two YELLOW rows are format examples — DELETE them before adding real data.", False),
    ("  一行 = 一篇帖子。黄色两行是格式示例——填写真实数据前请删除。", False),
    ("• Do not rename, reorder, or delete the header row. Leave column S and everything after it EMPTY (the tool writes there).", False),
    ("  不要改名、调换或删除表头行。S 列及之后请留空（工具会写入判定结果）。", False),
    ("• POST LINK: one unique xhslink.com / xiaohongshu.com link per row — never the same link on two rows.", False),
    ("  POST LINK：每行一条 xhslink.com / xiaohongshu.com 链接，绝不能两行共用。", False),
    ("• This template defaults to THOUSANDS (K): 130,000 followers → write 130; 1.74M → write 1741. If raw mode is selected in the report form, every row must use raw followers. Never mix units.", False),
    ("  本模板默认单位为「千」：13 万粉写 130；174 万粉写 1741。如在报告表单选择原始粉丝数，整份文件每行都必须使用原始值，绝不能混用单位。", False),
    ("• TTL ENGAGEMENT must equal LIKE + COLLECTION + COMMENT. The example rows carry the formula =L2+M2+N2 — copy it down.", False),
    ("  TTL ENGAGEMENT 必须等于 点赞+收藏+评论。示例行已带公式 =L2+M2+N2——请向下复制。", False),
    ("• TYPE and LEVEL cells have dropdowns — use them. POST DATE must be a real date (the column is date-formatted).", False),
    ("  TYPE 和 LEVEL 已设下拉选项——请直接选择。POST DATE 必须是真实日期（该列已设为日期格式）。", False),
    ("• PRICE in CNY, numbers only. CPM/CPE are optional — the tool recomputes them and never reuses yours.", False),
    ("  PRICE 为人民币，只填数字。CPM/CPE 选填——工具会重新计算，绝不复用。", False),
    ("", False),
    ("Full rules: see “DMR Reconciler — Input File Formatting Rubric”. 完整规则见《输入文件格式规范》。", False),
])
wb.save(OUTPUT_DIR / "PLOG_Tracker_Template.xlsx")

# =================================================================== DMR ==
DMR_HEADERS = ["Country", "Category", "Blogger", "Username", "Platform",
               "PostID", "Likes_Retweet", "Share_Favorites", "PostDate",
               "Followers", "Sector", "Brand", "Line", "HashTag", "Link",
               "Engagement", "WEIGHTED ENG.", "Tag", "HiddenEngagement",
               "SponsoredBy", "Comments"]
DMR_WIDTHS = [16, 12, 24, 26, 10, 26, 13, 14, 19, 11, 10, 12, 13, 24, 12,
              12, 13, 8, 16, 12, 10]

wb = Workbook()
ws = wb.active
ws.title = "Streaming"
meta = ws.cell(row=1, column=1, value=(
    "User: (exporter name)\n"
    "Generation date: 20/07/2026 05:22:55\n"
    "Top Bloggers - From 01/01/2026 To 20/07/2026"))
meta.font = NOTE_FONT
meta.alignment = Alignment(wrap_text=True, vertical="top")
ws.row_dimensions[1].height = 42
style_header(ws, 3, DMR_HEADERS, DMR_WIDTHS)
ws.freeze_panes = "A4"

example = ["MAINLAND CHINA", "Influencer", "Shi Li Bozhu 示例博主（删除此行）",
           "5f00000000000000000000aa", "RedBook", "69674d920000000000000001",
           12, 4, datetime(2026, 1, 14, 8, 2, 26), 121132, "Fashion",
           "DIOR Fashion", "Non Product", "迪奥,示例标签", "Show Post",
           23, 23, "", 0, "", 7]
for cidx, v in enumerate(example, start=4 and 1):
    c = ws.cell(row=4, column=cidx, value=v)
    c.font = BODY_FONT
    c.fill = EX_FILL
    c.border = BORDER

for r in range(4, 3004):
    ws.cell(row=r, column=4).number_format = "@"    # Username as TEXT
    ws.cell(row=r, column=6).number_format = "@"    # PostID as TEXT
    ws.cell(row=r, column=9).number_format = "yyyy-mm-dd hh:mm:ss"

readme_sheet(wb, [
    ("DMR export template · DMR 导出文件模板", True),
    ("", False),
    ("• The YELLOW row is a format example — DELETE it before adding real data.", False),
    ("  黄色一行是格式示例——填入真实数据前请删除。", False),
    ("• Cell A1 must keep all three lines, and the “From … To …” dates must cover every campaign post date — they define the export window.", False),
    ("  A1 单元格三行信息必须保留，其中「From … To …」日期必须覆盖所有投放发帖日期——它定义导出窗口。", False),
    ("• PostID: the 24-character hexadecimal Xiaohongshu note id. The column is pre-formatted as Text — never let Excel turn it into scientific notation.", False),
    ("  PostID：24 位十六进制的小红书笔记 ID。该列已预设为文本格式——绝不能被 Excel 转成科学计数法。", False),
    ("• Username: the author's 24-hex platform user id. REQUIRED on every row — without it 无博主 and 无帖子 cannot be told apart.", False),
    ("  Username：作者的 24 位十六进制用户 ID。每行必填——没有它无法区分「无博主」和「无帖子」。", False),
    ("• Engagement columns are the first-crawl snapshot, whole numbers. Link should be a hyperlink embedding the note id where possible.", False),
    ("  互动各列为首次抓取快照，填整数。Link 最好是内嵌笔记 ID 的超链接。", False),
    ("", False),
    ("Full rules: see “DMR Reconciler — Input File Formatting Rubric”. 完整规则见《输入文件格式规范》。", False),
])
wb.save(OUTPUT_DIR / "DMR_Export_Template.xlsx")
print(f"templates written to {OUTPUT_DIR}")
