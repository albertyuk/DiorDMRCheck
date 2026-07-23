#!/usr/bin/env python3
"""Evaluation harness — how you know you're done.

Runs the pipeline on the real PLOG + DMR source files and diffs the produced
column-S annotations against the human-made reference
(``PLOG_DMR_CHECK_1.xlsx``). Prints a confusion matrix and a per-row
disagreement report.

Known label noise in the reference (verified by direct inspection — do not
chase 100% agreement): 兔子糖糖公主Rinrin (2026-06-26) and 宅鱼日常
(2026-06-29) are marked 无博主 by the human but exist in the DMR file with
exact-date posts. The correct behavior is to match them; this harness lists
them as *expected* disagreements and excludes them from the acceptance gate
(all rows must agree after excusing only those two explicit noisy labels).

Usage:
    python tools/evaluate.py PLOG_DMR_CHECK.xlsx YTD_DMR_MICRO_0720.xlsx \
        PLOG_DMR_CHECK_1.xlsx [--no-llm]
"""
from __future__ import annotations

import argparse
import io
import sys
import time
from collections import Counter

from openpyxl import load_workbook

from app.core.xlsx import cell_str, find_header_row
from app.reconciler import perimeter as perimeter_mod
from app.reconciler.adjudicator import adjudicate
from app.reconciler.domain import (LINK_ERROR, NAME_MISLABEL, NO_BLOGGER,
                                   NO_BLOGGER_NOT_IN_PERIMETER, NO_POST,
                                   NO_POST_IN_PERIMETER, REVIEW, S_TEXT)
from app.reconciler.export import S_COL
from app.reconciler.parsers import PLOG_REQUIRED, parse_dmr, parse_plog
from app.reconciler.pipeline import run_pipeline

# (stable row identity: blogger name + PLOG date + matched note-id prefix,
# reviewed reference class, reviewed pipeline class) → why that exact
# transition is excused. A different row or failure is not label noise.
KNOWN_LABEL_NOISE = {
    ("兔子糖糖公主Rinrin", "2026-06-26", "6a3e4f7a",
     S_TEXT[NO_BLOGGER], "MATCH"):
        "human wrote 无博主 but DMR has an exact-date post (PostID 6a3e4f7a…)",
    ("宅鱼日常", "2026-06-29", "6a421ff9",
     S_TEXT[NO_BLOGGER], "MATCH"):
        "human wrote 无博主 but DMR has an exact-date post (PostID 6a421ff9…)",
}


def classify(text: str) -> str:
    """Collapse an annotation string to a comparable class (S_TEXT values)."""
    t = (text or "").strip()
    if not t:
        return "MATCH"
    if t.startswith(S_TEXT[NO_BLOGGER]):
        return S_TEXT[NO_BLOGGER]
    if t.startswith(S_TEXT[NO_POST]):
        return S_TEXT[NO_POST]
    # 人工复核 must be classified before the 链接 check — REVIEW reasons can
    # mention 链接 (e.g. 人工复核（链接已解析但…）).
    if t.startswith("人工") or S_TEXT[REVIEW] in t:
        return S_TEXT[REVIEW]
    if t.lower().startswith("check") or "链接" in t:
        return S_TEXT[LINK_ERROR]
    if "标注错误" in t or "名字" in t:
        return NAME_MISLABEL
    return t


def _index_occurrences(rows, value_for) -> dict[tuple[str, str, int], object]:
    """Index duplicate business keys without silently overwriting rows."""
    seen: Counter[tuple[str, str]] = Counter()
    out = {}
    for row in rows:
        base = (row.campaign, row.no)
        seen[base] += 1
        out[(row.campaign, row.no, seen[base])] = value_for(row)
    return out


def _comparison_keys(reference: dict, ours: dict) -> list:
    """Union is intentional: missing *and* unexpected rows must fail."""
    return sorted(set(reference) | set(ours))


def _known_noise_reason(verdict, want: str, got: str) -> str:
    if verdict is None:
        return ""
    note_id = str(getattr(verdict, "matched_post_id", "") or "").casefold()
    for identity, reason in KNOWN_LABEL_NOISE.items():
        name, post_date, note_prefix, expected_want, expected_got = identity
        if ((verdict.name, verdict.post_date or "", want, got)
                == (name, post_date, expected_want, expected_got)
                and note_id.startswith(note_prefix)):
            return reason
    return ""


def load_reference(path: str) -> dict[tuple[str, str, int], str]:
    """Reference annotations keyed by (CAMPAIGN, NO, occurrence)."""
    ref = parse_plog(path)  # same row identity logic as the pipeline
    wb = load_workbook(path, data_only=True)
    ws = None
    for candidate in wb.worksheets:
        if find_header_row(candidate, PLOG_REQUIRED):
            ws = candidate
            break
    if ws is None:
        raise ValueError("reference workbook has no PLOG header row")
    return _index_occurrences(
        ref.rows,
        lambda row: cell_str(ws.cell(row=row.excel_row, column=S_COL).value),
    )


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("plog")
    ap.add_argument("dmr")
    ap.add_argument("reference")
    ap.add_argument("--no-llm", action="store_true",
                    help="skip Tier-4 adjudication (deterministic tiers only)")
    ap.add_argument("--perimeter", default="",
                    help="optional Micro perimeter workbook — splits 无博主 by "
                         "membership (both split statuses map back to 无博主 "
                         "for the agreement math, since the reference predates "
                         "this feature)")
    args = ap.parse_args()

    t0 = time.time()
    plog = parse_plog(args.plog)
    dmr = parse_dmr(args.dmr)
    for w in plog.warnings + dmr.warnings:
        print(f"  [parse warning] {w}")

    perim = None
    if args.perimeter:
        data = open(args.perimeter, "rb").read()
        h = perimeter_mod.file_hash(data)
        perim = perimeter_mod.load_cached(h)
        if perim is None:
            print("  [perimeter] parsing (first time for this file hash)…",
                  file=sys.stderr)
            parsed = perimeter_mod.parse_perimeter(
                io.BytesIO(data), filename=args.perimeter, content_hash=h)
            perimeter_mod.store_parsed(parsed)
            perim = perimeter_mod.load_cached(h)
            for w in parsed.warnings:
                print(f"  [perimeter warning] {w}")
        print(f"  [perimeter] {len(perim.rows):,} rows, "
              f"{len(perim.by_redbook):,} with REDBOOK_ID, "
              f"extracted {perim.extraction_date or '?'}")

    tikhub_calls = [0]

    def counter():
        tikhub_calls[0] += 1

    def progress(phase, done, total, msg):
        if done in (1, total) or done % 10 == 0:
            print(f"  [{phase}] {msg}", file=sys.stderr)

    verdicts = run_pipeline(plog, dmr, progress=progress, tikhub_counter=counter,
                            perimeter=perim)
    if not args.no_llm:
        adjudicate(verdicts)
    elapsed = time.time() - t0

    if perim is not None:
        inside = [v for v in verdicts if v.status == NO_POST_IN_PERIMETER]
        outside = [v for v in verdicts if v.status == NO_BLOGGER_NOT_IN_PERIMETER]
        print("\n=== Perimeter split of 无博主 rows ===")
        print(f"  in perimeter (→ DMR gap): {len(inside)}")
        for v in inside:
            print(f"    {v.name} — REDBOOK {v.perimeter_redbook_id} "
                  f"({v.perimeter_name or v.perimeter_namebis})")
        print(f"  not in perimeter: {len(outside)}")
        for v in outside:
            extra = f" [{v.perimeter_note}]" if v.perimeter_note else ""
            print(f"    {v.name}{extra}")

    reference = load_reference(args.reference)
    ours = _index_occurrences(verdicts, lambda verdict: verdict)

    confusion: Counter[tuple[str, str]] = Counter()
    disagreements, excused = [], []
    all_keys = _comparison_keys(reference, ours)
    for key in all_keys:
        ref_text = reference.get(key)
        v = ours.get(key)
        got = classify(v.column_s()) if v else "(row missing)"
        want = classify(ref_text) if ref_text is not None else "(unexpected row)"
        confusion[(want, got)] += 1
        if got != want:
            name = v.name if v else "?"
            pdate = v.post_date if v else "?"
            reason = _known_noise_reason(v, want, got)
            entry = (key, name, pdate, want, got, reason)
            if reason:
                excused.append(entry)
            else:
                disagreements.append(entry)

    total = len(all_keys)
    agree = sum(n for (w, g), n in confusion.items() if w == g)

    print("\n=== Confusion matrix (reference → pipeline) ===")
    labels = sorted({k for pair in confusion for k in pair})
    width = max(len(x) for x in labels) + 2
    print(" " * width + "".join(x.ljust(width) for x in labels))
    for want in labels:
        row = [str(confusion.get((want, got), 0)).ljust(width) for got in labels]
        print(want.ljust(width) + "".join(row))

    print(f"\nAgreement: {agree}/{total}"
          f"  (excused known-noise disagreements: {len(excused)})")
    print(f"Elapsed: {elapsed:.1f}s · TikHub calls: {tikhub_calls[0]}")

    if excused:
        print("\n=== Expected disagreements (known reference label noise) ===")
        for key, name, pdate, want, got, why in excused:
            print(f"  {key} {name} {pdate}: reference={want!r} pipeline={got!r} — {why}")
    if disagreements:
        print("\n=== Unexplained disagreements ===")
        for key, name, pdate, want, got, _ in disagreements:
            v = ours.get(key)
            print(f"  {key} {name} {pdate}: reference={want!r} pipeline={got!r}"
                  f" [tier={v.tier if v else '?'}"
                  f" note={v.notes[0][:100] if v and v.notes else ''}]")

    # Only individually reviewed, explicit allow-list entries may be excused.
    # The pipeline's own tier/status can never validate itself.
    effective = agree + len(excused)
    ok = effective == total
    print(f"\nAcceptance (all rows after explicit known-noise exceptions): "
          f"{'PASS' if ok else 'FAIL'} ({effective}/{total}; "
          f"{len(disagreements)} unexplained)")
    return 0 if ok else 1


if __name__ == "__main__":
    sys.exit(main())
