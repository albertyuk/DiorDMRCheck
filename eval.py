#!/usr/bin/env python3
"""Evaluation harness — how you know you're done.

Runs the pipeline on the real PLOG + DMR source files and diffs the produced
column-S annotations against the human-made reference
(``PLOG_DMR_CHECK_1.xlsx``). Prints a confusion matrix and a per-row
disagreement report.

Known label noise in the reference (verified by direct inspection — do not
chase 100% agreement): 兔子糖糖公主Rinrin (2026-06-26) and 宅鱼日常
(2026-06-29) are marked 无博主 by the human but exist in the DMR file with
exact-date posts. The correct behavior is to match them; this harness lists
them as *expected* disagreements and excludes them from the acceptance gate
(≥ 99/101 agreement after excusing the two noisy labels).

Usage:
    python eval.py PLOG_DMR_CHECK.xlsx YTD_DMR_MICRO_0720.xlsx \
        PLOG_DMR_CHECK_1.xlsx [--no-llm]
"""
from __future__ import annotations

import argparse
import sys
import time
from collections import Counter

from openpyxl import load_workbook

from app.adjudicator import adjudicate
from app.matcher import run_pipeline
from app.parsers import PLOG_REQUIRED, _cell_str, _find_header_row, parse_dmr, parse_plog

# (blogger name, PLOG post date ISO) → why it is excused
KNOWN_LABEL_NOISE = {
    ("兔子糖糖公主Rinrin", "2026-06-26"):
        "human wrote 无博主 but DMR has an exact-date post (PostID 6a3e4f7a…)",
    ("宅鱼日常", "2026-06-29"):
        "human wrote 无博主 but DMR has an exact-date post (PostID 6a421ff9…)",
}

S_COL = 19


def classify(text: str) -> str:
    """Collapse an annotation string to a comparable class."""
    t = (text or "").strip()
    if not t:
        return "MATCH"
    if t.startswith("无博主"):
        return "无博主"
    if t.startswith("无帖子"):
        return "无帖子"
    # 人工复核 must be classified before the 链接 check — REVIEW reasons can
    # mention 链接 (e.g. 人工复核（链接已解析但…）).
    if t.startswith("人工") or "人工复核" in t:
        return "人工复核"
    if t.lower().startswith("check") or "链接" in t:
        return "Check链接错误"
    if "标注错误" in t or "名字" in t:
        return "有 但是DMR博主名字标注错误"
    return t


def load_reference(path: str) -> dict[tuple[str, str], str]:
    """Reference column-S annotations keyed by (CAMPAIGN, NO)."""
    ref = parse_plog(path)  # same row identity logic as the pipeline
    wb = load_workbook(path, data_only=True)
    ws = None
    for candidate in wb.worksheets:
        if _find_header_row(candidate, PLOG_REQUIRED):
            ws = candidate
            break
    assert ws is not None
    out = {}
    for row in ref.rows:
        out[row.key] = _cell_str(ws.cell(row=row.excel_row, column=S_COL).value)
    return out


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("plog")
    ap.add_argument("dmr")
    ap.add_argument("reference")
    ap.add_argument("--no-llm", action="store_true",
                    help="skip Tier-4 adjudication (deterministic tiers only)")
    args = ap.parse_args()

    t0 = time.time()
    plog = parse_plog(args.plog)
    dmr = parse_dmr(args.dmr)
    for w in plog.warnings + dmr.warnings:
        print(f"  [parse warning] {w}")

    tikhub_calls = [0]

    def counter():
        tikhub_calls[0] += 1

    def progress(phase, done, total, msg):
        if done in (1, total) or done % 10 == 0:
            print(f"  [{phase}] {msg}", file=sys.stderr)

    verdicts = run_pipeline(plog, dmr, progress=progress, tikhub_counter=counter)
    if not args.no_llm:
        adjudicate(verdicts)
    elapsed = time.time() - t0

    reference = load_reference(args.reference)
    ours = {(v.campaign, v.no): v for v in verdicts}

    confusion: Counter[tuple[str, str]] = Counter()
    disagreements, excused = [], []
    for key, ref_text in reference.items():
        v = ours.get(key)
        got = classify(v.column_s()) if v else "(row missing)"
        want = classify(ref_text)
        confusion[(want, got)] += 1
        if got != want:
            name = v.name if v else "?"
            pdate = v.post_date if v else "?"
            noise_key = (name, pdate or "")
            entry = (key, name, pdate, want, got,
                     KNOWN_LABEL_NOISE.get(noise_key, ""))
            (excused if noise_key in KNOWN_LABEL_NOISE else disagreements).append(entry)

    total = len(reference)
    agree = sum(n for (w, g), n in confusion.items() if w == g)

    print("\n=== Confusion matrix (reference → pipeline) ===")
    labels = sorted({k for pair in confusion for k in pair})
    width = max(len(x) for x in labels) + 2
    print(" " * width + "".join(x.ljust(width) for x in labels))
    for want in labels:
        row = [str(confusion.get((want, got), 0)).ljust(width) for got in labels]
        print(want.ljust(width) + "".join(row))

    print(f"\nAgreement: {agree}/{total}"
          f"  (excused known-noise disagreements: {len(excused)})")
    print(f"Elapsed: {elapsed:.1f}s · TikHub calls: {tikhub_calls[0]}")

    if excused:
        print("\n=== Expected disagreements (known reference label noise) ===")
        for key, name, pdate, want, got, why in excused:
            print(f"  {key} {name} {pdate}: reference={want!r} pipeline={got!r} — {why}")
    if disagreements:
        print("\n=== Disagreements ===")
        for key, name, pdate, want, got, _ in disagreements:
            v = ours.get(key)
            print(f"  {key} {name} {pdate}: reference={want!r} pipeline={got!r}"
                  f" [tier={v.tier if v else '?'}"
                  f" note={v.notes[0][:100] if v and v.notes else ''}]")

    effective = agree + len(excused)
    ok = effective >= total - 2
    print(f"\nAcceptance (≥ {total - 2}/{total} after excusing noise): "
          f"{'PASS' if ok else 'FAIL'} ({effective}/{total})")
    return 0 if ok else 1


if __name__ == "__main__":
    sys.exit(main())
