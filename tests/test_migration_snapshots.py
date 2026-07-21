"""Migration safety harnesses.

Two snapshot tests that must stay green — with UNCHANGED committed snapshots —
after every step of the reorganization:

1. Route table: every (path, methods, kind) the ASGI app exposes.
2. Golden export: the full deterministic output of the reconciliation pipeline
   on the synthetic fixtures (annotated-xlsx cell values, verdict documents,
   counts/buckets, reverse audit). Network resolution is faked; the LLM tier
   is off. Any value drift means a move changed behavior.

Regenerate intentionally with:  UPDATE_SNAPSHOTS=1 pytest tests/test_migration_snapshots.py
"""
from __future__ import annotations

import json
import os
from pathlib import Path

from openpyxl import load_workbook

SNAPSHOT_DIR = Path(__file__).resolve().parent / "snapshots"


def _check(name: str, produced: dict) -> None:
    """Compare `produced` against the committed snapshot (JSON-normalized)."""
    path = SNAPSHOT_DIR / name
    normalized = json.loads(json.dumps(produced, ensure_ascii=False, default=str))
    if os.environ.get("UPDATE_SNAPSHOTS") == "1" or not path.exists():
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(json.dumps(normalized, ensure_ascii=False, indent=1,
                                   sort_keys=True) + "\n", encoding="utf-8")
        return
    committed = json.loads(path.read_text(encoding="utf-8"))
    assert normalized == json.loads(json.dumps(committed)), (
        f"{name} drifted from the committed snapshot. If this change is "
        f"intentional, regenerate with UPDATE_SNAPSHOTS=1 and review the diff.")


def test_route_table_snapshot():
    from app.main import app

    routes = sorted(
        [getattr(r, "path", getattr(r, "path_format", "?")),
         sorted(getattr(r, "methods", None) or []),
         type(r).__name__]
        for r in app.routes
    )
    _check("route_table.json", {"routes": routes})


def test_golden_export_snapshot(plog_path, dmr_path, fake_resolver, tmp_path):
    from app.matcher import run_pipeline, status_counts, summary_buckets
    from app.parsers import parse_dmr, parse_plog
    from app.report import write_annotated_xlsx
    from app.reverse_audit import reverse_audit

    plog = parse_plog(plog_path)
    dmr = parse_dmr(dmr_path)
    verdicts = run_pipeline(plog, dmr)

    out = tmp_path / "annotated.xlsx"
    write_annotated_xlsx(plog_path, str(out), verdicts,
                         header_row=plog.header_row, sheet_name=plog.sheet)
    ws = load_workbook(str(out))[plog.sheet]
    cells = [[c.row, c.column, str(c.value)]
             for row in ws.iter_rows() for c in row if c.value is not None]

    counts = status_counts(verdicts)
    _check("golden_export.json", {
        "plog_meta": {"sheet": plog.sheet, "header_row": plog.header_row,
                      "rows": len(plog.rows), "campaigns": plog.campaigns,
                      "warnings": plog.warnings},
        "dmr_meta": {"sheet": dmr.sheet, "header_row": dmr.header_row,
                     "rows": len(dmr.rows), "warnings": dmr.warnings},
        "cells": cells,
        "verdicts": [v.to_dict() for v in verdicts],
        "counts": counts,
        "buckets": summary_buckets(counts),
        "reverse_audit": reverse_audit(plog, dmr, verdicts),
    })
