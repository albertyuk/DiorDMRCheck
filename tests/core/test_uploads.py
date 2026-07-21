from __future__ import annotations

import asyncio
import io
import os
import threading
import time
import zipfile
from xml.etree import ElementTree

import pytest
from fastapi import FastAPI, UploadFile
from openpyxl import Workbook, load_workbook
from starlette import formparsers

from app.core.uploads import (RequestBodyLimitMiddleware, UploadLimitError,
                              active_upload_names, cleanup_expired,
                              cleanup_over_budget, read_limited,
                              register_active_upload, save_limited,
                              run_upload_task, run_upload_task_sync,
                              unregister_active_upload,
                              validate_xlsx_archive)


def _http_scope(*, headers=()):
    return {
        "type": "http",
        "asgi": {"version": "3.0"},
        "http_version": "1.1",
        "method": "POST",
        "scheme": "http",
        "path": "/upload",
        "raw_path": b"/upload",
        "query_string": b"",
        "root_path": "",
        "headers": list(headers),
        "client": ("testclient", 50000),
        "server": ("testserver", 80),
    }


def test_request_body_limit_rejects_content_length_before_reading():
    app_called = False
    receive_called = False
    sent = []

    async def inner(_scope, _receive, _send):
        nonlocal app_called
        app_called = True

    async def receive():
        nonlocal receive_called
        receive_called = True
        return {"type": "http.request", "body": b""}

    async def send(message):
        sent.append(message)

    middleware = RequestBodyLimitMiddleware(
        inner, lambda _method, _path: 5
    )
    asyncio.run(
        middleware(
            _http_scope(headers=((b"content-length", b"6"),)),
            receive,
            send,
        )
    )

    assert not app_called
    assert not receive_called
    assert sent[0]["status"] == 413
    assert sent[1]["body"] == b"Request body too large."


def test_request_body_limit_rejects_streamed_overflow():
    completed = False
    sent = []
    requests = iter(
        [
            {"type": "http.request", "body": b"abc", "more_body": True},
            {"type": "http.request", "body": b"def", "more_body": False},
        ]
    )

    async def inner(_scope, receive, _send):
        nonlocal completed
        while (await receive()).get("more_body"):
            pass
        completed = True

    async def receive():
        return next(requests)

    async def send(message):
        sent.append(message)

    middleware = RequestBodyLimitMiddleware(
        inner, lambda _method, _path: 5
    )
    asyncio.run(middleware(_http_scope(), receive, send))

    assert not completed
    assert sent[0]["status"] == 413
    assert sent[1]["body"] == b"Request body too large."


def test_streamed_multipart_overflow_closes_partial_spool(monkeypatch):
    """The real multipart parser must close files created before overflow."""
    boundary = b"upload-boundary"
    first_chunk = (
        b"--" + boundary + b"\r\n"
        b'Content-Disposition: form-data; name="report"; '
        b'filename="report.xlsx"\r\n'
        b"Content-Type: application/octet-stream\r\n\r\n"
        b"partial"
    )
    final_chunk = b"overflow\r\n--" + boundary + b"--\r\n"
    requests = iter(
        [
            {
                "type": "http.request",
                "body": first_chunk,
                "more_body": True,
            },
            {
                "type": "http.request",
                "body": final_chunk,
                "more_body": False,
            },
        ]
    )
    sent = []
    spools = []
    real_spooled_file = formparsers.SpooledTemporaryFile

    def recording_spooled_file(*args, **kwargs):
        spool = real_spooled_file(*args, **kwargs)
        spools.append(spool)
        return spool

    monkeypatch.setattr(
        formparsers, "SpooledTemporaryFile", recording_spooled_file
    )

    app = FastAPI()

    @app.post("/upload")
    async def receive_report(report: UploadFile):
        return {"filename": report.filename}

    middleware = RequestBodyLimitMiddleware(
        app, lambda _method, _path: len(first_chunk)
    )

    async def receive():
        return next(requests)

    async def send(message):
        sent.append(message)

    asyncio.run(
        middleware(
            _http_scope(
                headers=(
                    (
                        b"content-type",
                        b"multipart/form-data; boundary=" + boundary,
                    ),
                )
            ),
            receive,
            send,
        )
    )

    starts = [
        message for message in sent
        if message["type"] == "http.response.start"
    ]
    assert [message["status"] for message in starts] == [413]
    assert spools
    assert all(spool.closed for spool in spools)


def test_upload_admission_precedes_downstream_body_handling():
    entered = 0
    first_entered = asyncio.Event()
    release = asyncio.Event()

    async def inner(_scope, _receive, send):
        nonlocal entered
        entered += 1
        first_entered.set()
        await release.wait()
        await send({"type": "http.response.start", "status": 204,
                    "headers": []})
        await send({"type": "http.response.body", "body": b""})

    async def receive():
        return {"type": "http.request", "body": b"", "more_body": False}

    async def exercise():
        middleware = RequestBodyLimitMiddleware(
            inner,
            lambda _method, _path: 100,
            admission_for_path=lambda _method, _path: True,
            max_concurrent_uploads=1,
        )
        sent = []

        async def send(message):
            sent.append(message)

        first = asyncio.create_task(
            middleware(_http_scope(), receive, send)
        )
        await first_entered.wait()
        second = asyncio.create_task(
            middleware(_http_scope(), receive, send)
        )
        await asyncio.sleep(0.02)
        assert entered == 1
        release.set()
        await asyncio.gather(first, second)
        assert entered == 2

    asyncio.run(exercise())


def test_read_limited_rejects_oversize():
    upload = UploadFile(io.BytesIO(b"x" * 11), filename="large.xlsx")
    with pytest.raises(UploadLimitError, match="limit"):
        asyncio.run(read_limited(upload, 10))


def test_save_limited_removes_partial_file(tmp_path):
    upload = UploadFile(io.BytesIO(b"x" * 11), filename="large.xlsx")
    target = tmp_path / "large.xlsx"
    with pytest.raises(UploadLimitError, match="limit"):
        asyncio.run(save_limited(upload, target, 10))
    assert not target.exists()


def test_save_limited_removes_partial_file_when_cancelled(tmp_path):
    class CancellingUpload:
        def __init__(self):
            self.reads = 0

        async def read(self, _size):
            self.reads += 1
            if self.reads == 1:
                return b"partial"
            raise asyncio.CancelledError

    target = tmp_path / "cancelled.xlsx"
    with pytest.raises(asyncio.CancelledError):
        asyncio.run(save_limited(CancellingUpload(), target, 100))
    assert not target.exists()


def test_validate_xlsx_archive_caps_expansion():
    data = io.BytesIO()
    with zipfile.ZipFile(data, "w", zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("xl/worksheets/sheet1.xml", b"x" * 10_000)
    with pytest.raises(UploadLimitError, match="expands"):
        validate_xlsx_archive(data.getvalue(), max_uncompressed_bytes=100,
                              max_entries=10)


def test_validate_xlsx_archive_caps_populated_cells():
    data = io.BytesIO()
    workbook = Workbook()
    workbook.active.append([1, 2])
    workbook.save(data)

    with pytest.raises(UploadLimitError, match="populated cells"):
        validate_xlsx_archive(
            data.getvalue(),
            max_uncompressed_bytes=1_000_000,
            max_entries=100,
            max_cells=1,
        )


def test_cell_cap_covers_non_xml_worksheet_relationship_targets():
    source = io.BytesIO()
    workbook = Workbook()
    workbook.active.append(["first", "second"])
    workbook.save(source)

    original_part = "xl/worksheets/sheet1.xml"
    renamed_part = "xl/custom/sheet1.data"
    data = io.BytesIO()
    with (
        zipfile.ZipFile(io.BytesIO(source.getvalue())) as source_archive,
        zipfile.ZipFile(data, "w", zipfile.ZIP_DEFLATED) as output_archive,
    ):
        for info in source_archive.infolist():
            member = source_archive.read(info)
            output_name = info.filename
            if output_name == "xl/_rels/workbook.xml.rels":
                member = member.replace(
                    b"/xl/worksheets/sheet1.xml",
                    b"/xl/custom/sheet1.data",
                )
            elif output_name == "[Content_Types].xml":
                member = member.replace(
                    b"/xl/worksheets/sheet1.xml",
                    b"/xl/custom/sheet1.data",
                )
            elif output_name == original_part:
                output_name = renamed_part
            output_archive.writestr(output_name, member)

    # The renamed package is still a usable workbook, despite the worksheet
    # no longer having an XML extension or the conventional directory name.
    loaded = load_workbook(io.BytesIO(data.getvalue()), read_only=True)
    try:
        assert list(loaded.active.values)[0] == ("first", "second")
    finally:
        loaded.close()

    with pytest.raises(UploadLimitError, match="populated cells"):
        validate_xlsx_archive(
            data.getvalue(),
            max_uncompressed_bytes=1_000_000,
            max_entries=100,
            max_cells=1,
        )


def test_cell_cap_covers_percent_encoded_worksheet_part_names():
    source = io.BytesIO()
    workbook = Workbook()
    workbook.active.append(["first", "second", "third"])
    workbook.save(source)

    original_part = "xl/worksheets/sheet1.xml"
    renamed_part = "xl/custom/sheet%201.data"
    data = io.BytesIO()
    with (
        zipfile.ZipFile(io.BytesIO(source.getvalue())) as source_archive,
        zipfile.ZipFile(data, "w", zipfile.ZIP_DEFLATED) as output_archive,
    ):
        for info in source_archive.infolist():
            member = source_archive.read(info)
            output_name = info.filename
            if output_name == "xl/_rels/workbook.xml.rels":
                member = member.replace(
                    b"/xl/worksheets/sheet1.xml",
                    b"/xl/custom/sheet%201.data",
                )
            elif output_name == "[Content_Types].xml":
                member = member.replace(
                    b"/xl/worksheets/sheet1.xml",
                    b"/xl/custom/sheet%201.data",
                )
            elif output_name == original_part:
                output_name = renamed_part
            output_archive.writestr(output_name, member)

    loaded = load_workbook(io.BytesIO(data.getvalue()), read_only=True)
    try:
        assert list(loaded.active.values)[0] == (
            "first", "second", "third"
        )
    finally:
        loaded.close()

    with pytest.raises(UploadLimitError, match="populated cells"):
        validate_xlsx_archive(
            data.getvalue(),
            max_uncompressed_bytes=1_000_000,
            max_entries=100,
            max_cells=1,
        )


def test_cell_cap_canonicalizes_relationship_targets_exactly_once():
    source = io.BytesIO()
    workbook = Workbook()
    workbook.active.append(["first", "second", "third"])
    workbook.save(source)

    original_part = "xl/worksheets/sheet1.xml"
    renamed_part = "xl/custom/sheet%25201.data"
    data = io.BytesIO()
    with (
        zipfile.ZipFile(io.BytesIO(source.getvalue())) as source_archive,
        zipfile.ZipFile(data, "w", zipfile.ZIP_DEFLATED) as output_archive,
    ):
        for info in source_archive.infolist():
            member = source_archive.read(info)
            output_name = info.filename
            if output_name == "xl/_rels/workbook.xml.rels":
                member = member.replace(
                    b"/xl/worksheets/sheet1.xml",
                    b"/xl/custom/sheet%25201.data",
                )
            elif output_name == "[Content_Types].xml":
                root = ElementTree.fromstring(member)
                namespace = root.tag.removesuffix("Types")
                for element in list(root):
                    if (element.tag.endswith("Override")
                            and element.attrib.get("PartName")
                            == "/xl/worksheets/sheet1.xml"):
                        root.remove(element)
                ElementTree.SubElement(root, namespace + "Default", {
                    "Extension": "data",
                    "ContentType": (
                        "application/vnd.openxmlformats-officedocument."
                        "spreadsheetml.worksheet+xml"
                    ),
                })
                member = ElementTree.tostring(root)
            elif output_name == original_part:
                output_name = renamed_part
            output_archive.writestr(output_name, member)

    # A Default content type is valid OPC and forces discovery to rely on the
    # relationship target. openpyxl follows that double-encoded name verbatim.
    loaded = load_workbook(io.BytesIO(data.getvalue()), read_only=True)
    try:
        assert list(loaded.active.values)[0] == (
            "first", "second", "third"
        )
    finally:
        loaded.close()

    with pytest.raises(UploadLimitError, match="populated cells"):
        validate_xlsx_archive(
            data.getvalue(),
            max_uncompressed_bytes=1_000_000,
            max_entries=100,
            max_cells=1,
        )


def test_cell_cap_ignores_unrelated_custom_xml_elements():
    source = io.BytesIO()
    workbook = Workbook()
    workbook.active["A1"] = "only worksheet cell"
    workbook.save(source)

    data = io.BytesIO()
    with (
        zipfile.ZipFile(io.BytesIO(source.getvalue())) as source_archive,
        zipfile.ZipFile(data, "w", zipfile.ZIP_DEFLATED) as output_archive,
    ):
        for info in source_archive.infolist():
            output_archive.writestr(info, source_archive.read(info))
        output_archive.writestr(
            "customXml/item1.xml",
            b"<metadata><c/><c/><c/><c/><c/></metadata>",
        )

    loaded = load_workbook(io.BytesIO(data.getvalue()), read_only=True)
    try:
        assert loaded.active["A1"].value == "only worksheet cell"
    finally:
        loaded.close()

    validate_xlsx_archive(
        data.getvalue(),
        max_uncompressed_bytes=1_000_000,
        max_entries=100,
        max_cells=1,
    )


def test_cleanup_expired_only_removes_old_directories(tmp_path):
    old = tmp_path / "old"
    fresh = tmp_path / "fresh"
    old.mkdir()
    fresh.mkdir()
    (old / "client.xlsx").write_bytes(b"private")
    now = time.time()
    old_time = now - 1000
    os.utime(old, (old_time, old_time))

    assert cleanup_expired(tmp_path, 500, now=now) == 1
    assert not old.exists()
    assert fresh.exists()


def test_cleanup_expired_honors_predicate_and_calls_callback(tmp_path):
    removable = tmp_path / "removable"
    protected = tmp_path / "protected"
    removable.mkdir()
    protected.mkdir()
    now = time.time()
    old_time = now - 1000
    os.utime(removable, (old_time, old_time))
    os.utime(protected, (old_time, old_time))
    removed = []

    count = cleanup_expired(
        tmp_path,
        500,
        now=now,
        should_remove=lambda path: path.name != "protected",
        on_remove=lambda path: removed.append(path.name),
    )

    assert count == 1
    assert removed == ["removable"]
    assert not removable.exists()
    assert protected.exists()


def test_cleanup_over_budget_removes_oldest_eligible_directory(tmp_path):
    protected = tmp_path / "protected"
    oldest_eligible = tmp_path / "oldest-eligible"
    newest = tmp_path / "newest"
    for path, size in (
        (protected, 6),
        (oldest_eligible, 5),
        (newest, 4),
    ):
        path.mkdir()
        (path / "workbook.xlsx").write_bytes(b"x" * size)
    now = time.time()
    os.utime(protected, (now - 300, now - 300))
    os.utime(oldest_eligible, (now - 200, now - 200))
    os.utime(newest, (now - 100, now - 100))
    removed = []

    count = cleanup_over_budget(
        tmp_path,
        10,
        should_remove=lambda path: path.name != "protected",
        on_remove=lambda path: removed.append(path.name),
    )

    assert count == 1
    assert removed == ["oldest-eligible"]
    assert protected.exists()
    assert not oldest_eligible.exists()
    assert newest.exists()


def test_cleanup_delete_does_not_block_active_registration(
        tmp_path, monkeypatch):
    from app.core import uploads

    victim = tmp_path / "victim"
    victim.mkdir()
    (victim / "workbook.xlsx").write_bytes(b"private")
    original_remove_tree = uploads.remove_tree
    deletion_started = threading.Event()
    allow_deletion = threading.Event()
    registration_done = threading.Event()
    cleanup_result = []
    registration_result = []

    def stalled_remove_tree(path):
        deletion_started.set()
        assert allow_deletion.wait(2)
        original_remove_tree(path)

    def cleanup():
        cleanup_result.append(cleanup_over_budget(tmp_path, 0))

    candidate = tmp_path / "new-upload"

    def register_candidate():
        registration_result.append(register_active_upload(candidate))
        registration_done.set()

    monkeypatch.setattr(uploads, "remove_tree", stalled_remove_tree)
    cleanup_thread = threading.Thread(target=cleanup)
    registration_thread = threading.Thread(target=register_candidate)
    cleanup_thread.start()
    try:
        assert deletion_started.wait(2)
        registration_thread.start()
        assert registration_done.wait(0.5)
    finally:
        allow_deletion.set()
        cleanup_thread.join(2)
        if registration_thread.ident is not None:
            registration_thread.join(2)
        if registration_result and registration_result[0]:
            unregister_active_upload(candidate)

    assert not cleanup_thread.is_alive()
    assert not registration_thread.is_alive()
    assert cleanup_result == [1]
    assert registration_result == [True]


def test_active_upload_registry_protects_inflight_directory(tmp_path):
    inflight = tmp_path / "inflight"
    inflight.mkdir()
    (inflight / "partial.xlsx").write_bytes(b"client data")

    assert register_active_upload(inflight)
    try:
        assert inflight.name in active_upload_names()
        assert cleanup_over_budget(
            tmp_path,
            0,
            should_remove=lambda path: path.name not in active_upload_names(),
        ) == 0
        assert inflight.exists()
    finally:
        unregister_active_upload(inflight)

    assert cleanup_over_budget(tmp_path, 0) == 1
    assert not inflight.exists()


def test_active_upload_registry_reference_counts_overlapping_leases(tmp_path):
    run_dir = tmp_path / "same-run"
    assert register_active_upload(run_dir)
    assert register_active_upload(run_dir)
    unregister_active_upload(run_dir)
    assert run_dir.name in active_upload_names()
    unregister_active_upload(run_dir)
    assert run_dir.name not in active_upload_names()


def test_async_and_background_work_share_memory_gate(monkeypatch):
    from app.core import uploads

    monkeypatch.setattr(uploads, "_workbook_gate",
                        threading.BoundedSemaphore(1))
    lock = threading.Lock()
    active = 0
    peak = 0

    def heavy_work():
        nonlocal active, peak
        with lock:
            active += 1
            peak = max(peak, active)
        time.sleep(0.03)
        with lock:
            active -= 1

    async def exercise():
        await asyncio.gather(
            run_upload_task(None, heavy_work),
            asyncio.to_thread(run_upload_task_sync, heavy_work),
        )

    asyncio.run(exercise())
    assert peak == 1


def test_cancelled_gate_waiter_cannot_leak_or_release_permit_early(
        monkeypatch):
    from app.core import uploads

    gate = threading.BoundedSemaphore(1)
    monkeypatch.setattr(uploads, "_workbook_gate", gate)
    gate.acquire()
    worker_finished = threading.Event()

    async def exercise():
        task = asyncio.create_task(run_upload_task(
            None, worker_finished.set
        ))
        await asyncio.sleep(0.02)
        task.cancel()
        try:
            await asyncio.sleep(0.02)
            # Cancellation retains ownership until the worker safely leaves
            # the gate queue, without running abandoned workbook work.
            assert not task.done()
        finally:
            gate.release()
        with pytest.raises(asyncio.CancelledError):
            await asyncio.wait_for(task, 2)
        assert not worker_finished.is_set()

    asyncio.run(exercise())
    assert gate.acquire(blocking=False)
    gate.release()
