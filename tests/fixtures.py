"""Synthetic .xlsx fixtures reproducing every observed real-data quirk (§8):
emoji in NAME, full-width paren + double-space headers, NO resetting per
campaign, DMR metadata rows above the header, romanized-only Blogger variants,
early-crawl engagement snapshots, dead links, duplicate bloggers across
campaigns, and an export window that doesn't cover every PLOG date.
"""
from __future__ import annotations

import io
from datetime import datetime

from openpyxl import Workbook

PLOG_HEADERS = [
    "NO", "MCN", "CAMPAIGN", "TYPE", "LEVEL", "NAME", "FAN BASE（K)",
    "POST DATE", "MICRO MACRO", "POST LINK", "IMPRESSION", "LIKE",
    "COLLECTION", "COMMENT", "TTL  ENGAGEMENT", "PRICE", "CPM", "CPE",
]

DMR_HEADERS = [
    "Country", "Category", "Blogger", "Username", "Platform", "PostID",
    "Likes_Retweet", "Share_Favorites", "PostDate", "Followers", "Sector",
    "Brand", "Line", "HashTag", "Link", "Engagement", "WEIGHTED ENG.", "Tag",
    "HiddenEngagement", "SponsoredBy", "Comments",
]

# 24-hex note ids
N_MOCHI_MAY = "6a1a0000000000000000a001"   # 墨池墨吟 May post (PLOG 05-13)
N_MOCHI_MAY_DMR = "6a1a0000000000000000a002"  # different 墨池墨吟 post DMR crawled 05-11
N_MOCHI_JUN = "6a1a0000000000000000a003"   # 墨池墨吟 verified June match
N_BING = "6a1a0000000000000000a004"        # 饼饼 Δ4d drift match
N_GUNGUN = "6a1a0000000000000000a005"      # gungun_ match, DMR name mislabeled
N_JITUI = "6a1a0000000000000000a006"       # 鸡腿子's real DMR post (PLOG link dead)
N_EGG = "6a1a0000000000000000a007"         # 一颗鸡蛋🥚 — blogger absent from DMR
N_DUP_C3 = "6a1a0000000000000000a008"      # duplicate blogger, campaign #003 post
N_OLD = "6a1a0000000000000000a009"         # out-of-window December post
N_EXTRA = "6a1a0000000000000000a00a"       # DMR extra post → reverse audit
N_CONFLICT = "6a1a0000000000000000a00b"    # tier-2 name-conflict row
N_DEAD_DETAIL = "6a1a0000000000000000a00c"  # note resolves, detail dead → sibling tier

U_MOCHI = "5f00000000000000000000b1"
U_BING = "5f00000000000000000000b2"
U_GUNGUN = "5f00000000000000000000b3"
U_JITUI = "5f00000000000000000000b4"
U_EGG = "5f00000000000000000000b5"       # never appears in DMR
U_CONFLICT_PLOG = "5f00000000000000000000b6"  # resolved author, absent from DMR
U_CONFLICT_DMR = "5f00000000000000000000b7"   # same-name blogger's DMR Username

LINK = "http://xhslink.com/o/{}"


def plog_rows() -> list[dict]:
    """PLOG data rows in original order: campaign #002 first, then #001,
    then #003 (NO resets at each new CAMPAIGN value)."""
    return [
        # -------- PLOG #002
        dict(no=1, campaign="PLOG #002", name="墨池墨吟", date=datetime(2026, 5, 13),
             link=LINK.format("mochi-may"), like=1200, collection=80, comment=30),
        dict(no=2, campaign="PLOG #002", name="一颗鸡蛋🥚", date=datetime(2026, 5, 20),
             link=LINK.format("egg"), like=333, collection=21, comment=8),
        # -------- PLOG #001
        dict(no=1, campaign="PLOG #001", name="墨池墨吟", date=datetime(2026, 6, 24),
             link=LINK.format("mochi-jun"), like=607, collection=44, comment=19),
        dict(no=2, campaign="PLOG #001", name="饼饼", date=datetime(2026, 7, 1),
             link=LINK.format("bing"), like=5000, collection=400, comment=120),
        dict(no=3, campaign="PLOG #001", name="gungun_", date=datetime(2026, 6, 10),
             link=LINK.format("gungun"), like=900, collection=70, comment=25),
        dict(no=4, campaign="PLOG #001", name="鸡腿子", date=datetime(2026, 6, 15),
             link=LINK.format("jitui-dead"), like=777, collection=55, comment=12),
        dict(no=5, campaign="PLOG #001", name="早春的树", date=datetime(2026, 6, 18),
             link=LINK.format("conflict"), like=450, collection=33, comment=9),
        dict(no=6, campaign="PLOG #001", name="冬日限定", date=datetime(2025, 12, 1),
             link=LINK.format("old"), like=100, collection=5, comment=2),
        # -------- PLOG #003 (duplicate blogger across campaigns)
        dict(no=1, campaign="PLOG #003", name="墨池墨吟", date=datetime(2026, 7, 10),
             link=LINK.format("mochi-jul"), like=300, collection=20, comment=6),
        # note resolves but detail is dead/blocked (no author) — the author id
        # must come from the sibling 墨池墨吟 rows above → deterministic 无帖子
        dict(no=2, campaign="PLOG #003", name="墨池墨吟", date=datetime(2026, 7, 12),
             link=LINK.format("mochi-dead-detail"), like=50, collection=4, comment=1),
    ]


def build_plog(path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "MASTER KOL LIST"
    ws.append(PLOG_HEADERS)
    for r in plog_rows():
        ws.append([
            r["no"], "SomeMCN", r["campaign"], "图文", "KOC", r["name"], 12.5,
            r["date"], "MICRO", r["link"], 10000, r["like"], r["collection"],
            r["comment"], r["like"] + r["collection"] + r["comment"], 2000,
            3.1, 1.2,
        ])
    wb.save(path)


def dmr_rows() -> list[dict]:
    return [
        # 墨池墨吟's nearby-but-different May post (crawled 05-11; the PLOG
        # 05-13 note is NOT in DMR → 无帖子, provable only via note-ID join)
        dict(blogger="墨池墨吟", user=U_MOCHI, pid=N_MOCHI_MAY_DMR,
             likes=88, favs=6, date=datetime(2026, 5, 11, 9, 30), comments=4),
        # verified June match — early-crawl snapshot reads 14 likes vs PLOG 607
        dict(blogger="墨池墨吟", user=U_MOCHI, pid=N_MOCHI_JUN,
             likes=14, favs=1, date=datetime(2026, 6, 24, 8, 0), comments=0),
        # 饼饼 with Δ4d PostDate drift
        dict(blogger="Bing Bing 饼饼", user=U_BING, pid=N_BING,
             likes=4100, favs=350, date=datetime(2026, 7, 5, 22, 0), comments=100),
        # romanized-only variant of the handle gungun_
        dict(blogger="gungunnnnn", user=U_GUNGUN, pid=N_GUNGUN,
             likes=850, favs=60, date=datetime(2026, 6, 10, 12, 0), comments=20),
        # 鸡腿子 — real DMR post; the PLOG link for it is dead
        dict(blogger="Ji Tui Zi 鸡腿子", user=U_JITUI, pid=N_JITUI,
             likes=700, favs=48, date=datetime(2026, 6, 16, 10, 0), comments=11),
        # same-name blogger under a different Username (tier-2 conflict row)
        dict(blogger="早春的树", user=U_CONFLICT_DMR, pid=N_CONFLICT,
             likes=430, favs=30, date=datetime(2026, 6, 19, 7, 0), comments=8),
        # duplicate blogger campaign #003 post
        dict(blogger="墨池墨吟", user=U_MOCHI, pid=N_DUP_C3,
             likes=250, favs=18, date=datetime(2026, 7, 10, 20, 0), comments=5),
        # extra 墨池墨吟 post PLOG doesn't track → reverse audit
        dict(blogger="墨池墨吟", user=U_MOCHI, pid=N_EXTRA,
             likes=61, favs=3, date=datetime(2026, 6, 1, 13, 0), comments=2),
    ]


def build_dmr(path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Streaming"
    ws.append(["User: tester, Generation date: 2026-07-20, "
               "Top Bloggers - From 2026/01/01 To 2026/07/20"])
    ws.append(["(metadata row 2)"])
    ws.append(DMR_HEADERS)
    for r in dmr_rows():
        row = [
            "CN", "Beauty", r["blogger"], r["user"], "RED", r["pid"],
            r["likes"], r["favs"], r["date"], 150000, "Fashion", "Dior",
            "Makeup", "#tag", "Show Post", r["likes"] + r["favs"] + r["comments"],
            r["likes"] + 0.5, "", 0, "", r["comments"],
        ]
        ws.append(row)
        link_cell = ws.cell(row=ws.max_row, column=DMR_HEADERS.index("Link") + 1)
        link_cell.hyperlink = (
            "https://www.dmr.st/redi.html?url=https%3A%2F%2Fwww.xiaohongshu.com"
            f"%2Fdiscovery%2Fitem%2F{r['pid']}"
        )
    wb.save(path)


def fake_resolutions() -> dict[str, dict]:
    """URL → resolver outcome used by the tests' fake resolve_link."""
    return {
        LINK.format("mochi-may"): dict(note_id=N_MOCHI_MAY, author=U_MOCHI, nick="墨池墨吟"),
        LINK.format("egg"): dict(note_id=N_EGG, author=U_EGG, nick="一颗鸡蛋🥚"),
        LINK.format("mochi-jun"): dict(note_id=N_MOCHI_JUN, author=U_MOCHI, nick="墨池墨吟"),
        LINK.format("bing"): dict(note_id=N_BING, author=U_BING, nick="饼饼"),
        LINK.format("gungun"): dict(note_id=N_GUNGUN, author=U_GUNGUN, nick="gungun_"),
        LINK.format("jitui-dead"): dict(fail="link expired (404)"),
        LINK.format("conflict"): dict(note_id=N_CONFLICT[:-1] + "f",
                                      author=U_CONFLICT_PLOG, nick="早春的树"),
        LINK.format("old"): dict(note_id=N_OLD, author=U_MOCHI, nick="墨池墨吟"),
        LINK.format("mochi-jul"): dict(note_id=N_DUP_C3, author=U_MOCHI, nick="墨池墨吟"),
        LINK.format("mochi-dead-detail"): dict(note_id=N_DEAD_DETAIL, author="", nick=""),
    }


# --------------------------------------------------------------- efficiency
# Synthetic KOL efficiency workbook (shared by test_effreport and
# test_efficiency_web). 43 active rows; every metric hand-computable and each
# validation rule V2-V10 has a row that trips it — see build_eff_bytes.

EFF_HEADERS = [
    "NO", "MCN", "CAMPAIGN", "TYPE", "LEVEL", "NAME", "FAN BASE（K)",
    "POST DATE", "MICRO MACRO", "POST LINK", "IMPRESSION", "LIKE",
    "COLLECTION", "COMMENT", "TTL  ENGAGEMENT", "PRICE", "CPM", "CPE",
]

EFF_PAID, EFF_SOFT = "报备图文", "软植图文"


def _eff_row(no, type_, level, name, fan, link, impr, like, coll, comm, ttl,
             price, cpm=None):
    return [no, "", "EFF #001", type_, level, name, fan,
            datetime(2026, 6, 1 + (no - 1) % 28), "", link, impr, like, coll,
            comm, ttl,
            price, cpm, None]


def build_eff_bytes() -> bytes:
    """43 active rows. Hand-computed groups:

    MID PAID  n=2: prices 20k+10k, impr 200k+100k, eng 2000+1000
              → pooled CPM 100, per-post CPM 100, pooled CPE 10, avg 15k
    MID SOFT  n=3: 6k×3, impr 100k/50k/50k, eng 500/250/250
              → pooled CPM 90, per-post CPM 100, pooled CPE 18, avg 6k
    BOT SOFT  n=2: one normal + one IMPRESSION=0 (V3)
              → pooled CPM 50 (zero-impr row out of CPM only), CPE 10
    KOC PAID  n=4: impr 90k/90k/10k/10k → top-2 hold 90% (V10)
    TOP SOFT  n=1 (V9); BOT PAID n=30 (filler, makes TOP SOFT a <3% sliver)
    plus: V7 unknown TYPE row (in totals, no group), V2 missing-PRICE row
    (excluded), V6 duplicate link, V4 identity break, V5 CPM-column drift,
    V8 尾部+底部 both present.
    """
    PAID, SOFT = EFF_PAID, EFF_SOFT
    wb = Workbook()
    ws = wb.active
    ws.title = "MASTER KOL LIST"
    ws.append(EFF_HEADERS)
    rows = [
        _eff_row(1, PAID, "腰部达人", "mp1", 800, "http://x.co/mp1",
                 200_000, 1_500, 300, 200, 2_000, 20_000,
                 cpm=0.5),                                    # V5: true 0.1
        _eff_row(2, PAID, "腰部达人", "mp2", 700, "http://x.co/mp2",
                 100_000, 800, 100, 100, 1_000, 10_000),
        _eff_row(3, SOFT, "腰部达人", "ms1", 600, "http://x.co/dup",
                 100_000, 400, 50, 50, 500, 6_000),
        _eff_row(4, SOFT, "腰部达人", "ms2", 600, "http://x.co/dup",  # V6 pair
                 50_000, 200, 25, 25, 250, 6_000),
        _eff_row(5, SOFT, "腰部达人", "ms3", 600, "http://x.co/ms3",
                 50_000, 200, 25, 25, 250, 6_000),
        _eff_row(6, SOFT, "尾部达人", "bs1", 250, "http://x.co/bs1",
                 100_000, 400, 50, 50, 500, 5_000),
        _eff_row(7, SOFT, "底部达人", "bs2", 150, "http://x.co/bs2",
                 0, 400, 50, 50, 500, 5_000),                 # V3 zero impr
        _eff_row(8, PAID, "KOC", "kp1", 50, "http://x.co/kp1",
                 90_000, 500, 50, 50, 600, 1_000),
        _eff_row(9, PAID, "KOC", "kp2", 50, "http://x.co/kp2",
                 90_000, 500, 50, 50, 999, 1_000),            # V4: 600≠999
        _eff_row(10, PAID, "KOC", "kp3", 50, "http://x.co/kp3",
                 10_000, 100, 10, 10, 120, 1_000),
        _eff_row(11, PAID, "KOC", "kp4", 50, "http://x.co/kp4",
                 10_000, 100, 10, 10, 120, 1_000),
        _eff_row(12, SOFT, "头部达人", "ts1", 1_500, "http://x.co/ts1",
                 1_000_000, 5_000, 500, 500, 6_000, 90_000),  # V9 n=1
        _eff_row(13, "其他合作", "腰部达人", "unk", 500, "http://x.co/unk",
                 10_000, 50, 5, 5, 60, 2_000),                # V7 TYPE
        _eff_row(14, PAID, "腰部达人", "gap", 500, "http://x.co/gap",
                 10_000, 50, 5, 5, 60, None),                 # V2 no PRICE
    ]
    for i in range(30):                                       # BOT PAID filler
        rows.append(_eff_row(15 + i, PAID, "尾部达人", f"bp{i}", 250,
                             f"http://x.co/bp{i}", 10_000, 80, 10, 10, 100,
                             1_000))
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
