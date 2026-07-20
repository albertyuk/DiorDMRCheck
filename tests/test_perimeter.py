"""Perimeter cross-check: parsing, caching, and the NO_BLOGGER split.

Uses its own small fixture workbooks (separate from the shared fixtures so
existing campaign-shape assertions stay untouched)."""
from __future__ import annotations

import io
from datetime import datetime

import pytest
from openpyxl import Workbook

from app.matcher import (LINK_ERROR, NO_BLOGGER, NO_BLOGGER_NOT_IN_PERIMETER,
                         NO_POST_IN_PERIMETER, run_pipeline)
from app.parsers import parse_dmr, parse_plog
from app.perimeter import (PerimeterIndex, file_hash, load_cached,
                           parse_perimeter, store_parsed)
from app.resolver import Resolution
from tests import fixtures

PERIM_HEADERS = [
    "NAME", "NAMEBIS", "DMRID", "COUNTRY", "TYPE", "INSTAGRAM_ID",
    "INSTAGRAM_FOLLOWERS", "YOUTUBE_ID", "YOUTUBE_FOLLOWERS", "TIKTOK_ID",
    "TIKTOK_FOLLOWERS", "WEIBO_ID", "WEIBO_FOLLOWERS", "REDBOOK_ID",
    "REDBOOK_FOLLOWERS",
]

U_IN = "5f000000000000000000c001"      # in-perimeter author
U_NEAR = "5f000000000000000000c002"    # resolved author for the near-miss row
U_NEAR_PERIM = "5f000000000000000000c003"  # perimeter's different account
U_MULTI = "5f000000000000000000c004"   # esther-style multi-hit author
U_NOREG = "5f000000000000000000c005"   # name in perimeter, no REDBOOK_ID
U_NOWHERE = "5f000000000000000000c006" # not in perimeter at all


def _perimeter_rows():
    return [
        ("Yi Ke Ji Dan", "一颗鸡蛋", "100001", U_IN, 52000),
        ("Xiao Ce Shi", "小小测试", "100002", U_NEAR_PERIM, 9000),
        ("esther", "", "100003", "5f000000000000000000c0aa", 1000),
        ("esther lee", "", "100004", "", None),
        ("Esther W", "艾斯特", "100005", "5f000000000000000000c0bb", 2000),
        ("Wu Dang An", "无档案", "100006", "", None),
        ("Unrelated Person", "路人甲", "100007", "5f000000000000000000c0cc", 500),
    ]


def build_perimeter_bytes(extraction="19/05/2026 10:30:00") -> bytes:
    wb = Workbook()
    macro = wb.active
    macro.title = "List Macro"           # must be ignored
    macro.append(["NAME", "REDBOOK_ID"])
    macro.append(["Macro Person", "5f000000000000000000ffff"])
    ws = wb.create_sheet("List Micro")
    ws.append([])
    ws.append(["2026 MICRO social perimeter"])
    ws.append(["LVMH internal use only"])
    ws.append(["The perimeter is regularly updated"])
    ws.append([f"Date of extraction : {extraction}"])
    ws.append([])
    ws.append(PERIM_HEADERS)
    for name, namebis, dmrid, rid, followers in _perimeter_rows():
        ws.append([name, namebis, dmrid, "China", "Influencer",
                   "", None, "", None, "", None, "", None, rid, followers])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.fixture
def perim_index(tmp_path, monkeypatch) -> PerimeterIndex:
    from app import config, db
    monkeypatch.setattr(config, "DB_PATH", tmp_path / "perim.sqlite3")
    monkeypatch.setattr(db, "_initialized", False)
    data = build_perimeter_bytes()
    h = file_hash(data)
    parsed = parse_perimeter(io.BytesIO(data), filename="perim.xlsx",
                             content_hash=h)
    store_parsed(parsed)
    idx = load_cached(h)     # exercise the cache round-trip
    assert idx is not None
    return idx


def test_parse_perimeter_contract(perim_index):
    assert perim_index.extraction_date == "19/05/2026 10:30:00"
    assert len(perim_index.rows) == len(_perimeter_rows())
    # Macro sheet ignored
    assert "5f000000000000000000ffff" not in perim_index.by_redbook
    assert U_IN in perim_index.by_redbook
    row = perim_index.by_redbook[U_IN]
    assert row["namebis"] == "一颗鸡蛋" and row["redbook_followers"] == 52000


def test_scan_name_multi_hit(perim_index):
    hits = perim_index.scan_name("esther")
    assert len(hits) >= 2  # esther / esther lee / Esther W collide


# --------------------------------------------------------- pipeline split

def _mini_plog(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "MASTER KOL LIST"
    ws.append(fixtures.PLOG_HEADERS)
    rows = [
        (1, "一颗鸡蛋🥚", "in-perim"),
        (2, "小小测试", "near-miss"),
        (3, "esther", "multi"),
        (4, "无档案", "no-reg"),
        (5, "查无此人", "nowhere"),
        (6, "断链博主", "dead-link"),
    ]
    for no, name, tag in rows:
        ws.append([no, "", "C#1", "", "", name, 1, datetime(2026, 6, 1), "",
                   f"http://xhslink.com/o/{tag}", 1, 1, 1, 1, 3, 1, 1, 1])
    wb.save(path)


def _mini_dmr(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Streaming"
    ws.append(["Top Bloggers - From 2026/01/01 To 2026/07/20"])
    ws.append([])
    ws.append(fixtures.DMR_HEADERS)
    # one unrelated row so indexes are non-empty (none of the test authors)
    ws.append(["CN", "B", "别人", "5f00000000000000000000d1",
               "RED", "6a1a0000000000000000d001", 1, 1,
               datetime(2026, 6, 2), 1, "", "", "", "", "", 1, 1, "", 0, "", 1])
    wb.save(path)


RESOLUTIONS = {
    "http://xhslink.com/o/in-perim": (U_IN, "6a1a0000000000000000e001"),
    "http://xhslink.com/o/near-miss": (U_NEAR, "6a1a0000000000000000e002"),
    "http://xhslink.com/o/multi": (U_MULTI, "6a1a0000000000000000e003"),
    "http://xhslink.com/o/no-reg": (U_NOREG, "6a1a0000000000000000e004"),
    "http://xhslink.com/o/nowhere": (U_NOWHERE, "6a1a0000000000000000e005"),
}


@pytest.fixture
def split_verdicts(tmp_path, monkeypatch, perim_index):
    _mini_plog(tmp_path / "p.xlsx")
    _mini_dmr(tmp_path / "d.xlsx")

    def fake_resolve(url, run_counter=None, retry_failed=False):
        if url in RESOLUTIONS:
            author, note = RESOLUTIONS[url]
            return Resolution(status="ok", note_id=note, author_id=author,
                              source="fixture")
        return Resolution(status="failed", error="dead link")

    import app.matcher as m
    monkeypatch.setattr(m, "resolve_link", fake_resolve)
    monkeypatch.setattr(m, "ensure_author",
                        lambda url, res, run_counter=None, retry_failed=False: res)
    plog = parse_plog(str(tmp_path / "p.xlsx"))
    dmr = parse_dmr(str(tmp_path / "d.xlsx"))
    vs = run_pipeline(plog, dmr, perimeter=perim_index)
    return {v.name: v for v in vs}


def test_primary_id_join_flips_to_in_perimeter(split_verdicts):
    v = split_verdicts["一颗鸡蛋🥚"]
    assert v.status == NO_POST_IN_PERIMETER
    assert v.perimeter_method == "redbook-id"
    assert v.perimeter_redbook_id == U_IN
    assert v.column_s() == "无博主但在Perimeter内→无帖子"


def test_near_miss_same_name_different_id(split_verdicts):
    v = split_verdicts["小小测试"]
    assert v.status == NO_BLOGGER_NOT_IN_PERIMETER
    assert "REDBOOK_ID不同" in v.perimeter_note
    assert v.column_s() == "无博主（不在Perimeter内）"


def test_multi_hit_never_classifies_by_name(split_verdicts):
    v = split_verdicts["esther"]
    assert v.status == NO_BLOGGER_NOT_IN_PERIMETER
    assert len(v.perimeter_candidates) >= 2
    assert v.perimeter_method == ""  # nothing auto-picked


def test_single_hit_without_redbook_id(split_verdicts):
    v = split_verdicts["无档案"]
    assert v.status == NO_BLOGGER_NOT_IN_PERIMETER
    assert "未登记REDBOOK_ID" in v.perimeter_note


def test_not_in_perimeter_plain(split_verdicts):
    v = split_verdicts["查无此人"]
    assert v.status == NO_BLOGGER_NOT_IN_PERIMETER
    assert not v.perimeter_method and not v.perimeter_candidates


def test_dead_link_verdict_never_changes(split_verdicts):
    v = split_verdicts["断链博主"]
    assert v.status == LINK_ERROR
    assert v.column_s().startswith("Check链接错误")


def test_without_perimeter_behavior_unchanged(tmp_path, monkeypatch):
    _mini_plog(tmp_path / "p.xlsx")
    _mini_dmr(tmp_path / "d.xlsx")

    def fake_resolve(url, run_counter=None, retry_failed=False):
        if url in RESOLUTIONS:
            author, note = RESOLUTIONS[url]
            return Resolution(status="ok", note_id=note, author_id=author,
                              source="fixture")
        return Resolution(status="failed", error="dead link")

    import app.matcher as m
    monkeypatch.setattr(m, "resolve_link", fake_resolve)
    monkeypatch.setattr(m, "ensure_author",
                        lambda url, res, run_counter=None, retry_failed=False: res)
    plog = parse_plog(str(tmp_path / "p.xlsx"))
    dmr = parse_dmr(str(tmp_path / "d.xlsx"))
    vs = run_pipeline(plog, dmr, perimeter=None)
    by = {v.name: v for v in vs}
    assert by["一颗鸡蛋🥚"].status == NO_BLOGGER
    assert by["一颗鸡蛋🥚"].column_s() == "无博主"


def test_eval_classify_maps_new_statuses_to_no_blogger():
    import eval as ev
    assert ev.classify("无博主但在Perimeter内→无帖子") == "无博主"
    assert ev.classify("无博主（不在Perimeter内）") == "无博主"
