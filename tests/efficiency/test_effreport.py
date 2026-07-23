"""KOL efficiency report: engine (effreport) and deck generation.

The synthetic workbook is built so every metric is hand-computable, and the
core validation rules have rows that trip them. The golden test against the
real client workbook runs only when the (gitignored) file is present.
"""
from __future__ import annotations

import io
import zipfile
from pathlib import Path

import pytest
from openpyxl import Workbook

from app.efficiency.deck import assert_chart_cache, build_deck
from app.efficiency.analysis import (ReportConfig, VerificationError, analyze,
                           compute_metrics,
                           compute_metrics_pandas, verify_dual_path)
from tests.fixtures import build_eff_bytes


@pytest.fixture(scope="module")
def analysis() -> dict:
    return analyze(io.BytesIO(build_eff_bytes()), ReportConfig())


def _codes(analysis):
    return {f["code"] for f in analysis["findings"]}


# ------------------------------------------------------------------ metrics

def test_group_metrics_hand_computed(analysis):
    g = analysis["metrics"]["groups"]
    mp = g["MID PAID"]
    assert mp["n"] == 2
    assert mp["avg_price"] == pytest.approx(15_000)
    assert mp["cpm_pooled"] == pytest.approx(100.0)
    assert mp["cpm_perpost"] == pytest.approx(100.0)
    assert mp["cpe_pooled"] == pytest.approx(10.0)
    ms = g["MID SOFT"]
    assert ms["cpm_pooled"] == pytest.approx(90.0)
    assert ms["cpm_perpost"] == pytest.approx(100.0)
    assert ms["cpe_pooled"] == pytest.approx(18.0)


def test_zero_impression_row_excluded_from_cpm_only(analysis):
    bs = analysis["metrics"]["groups"]["BOT SOFT"]
    assert bs["n"] == 2                       # still counted in share/n
    # Both posts incurred spend.  The zero-impression post is excluded only
    # from per-post ratios, not from the pooled campaign numerator.
    assert bs["cpm_pooled"] == pytest.approx(100.0)  # 10000/100000*1000
    assert bs["cpe_pooled"] == pytest.approx(10.0)   # both rows have eng
    assert "V3" in _codes(analysis)


def test_missing_group_is_absent_not_zero(analysis):
    assert "TOP PAID" not in analysis["metrics"]["groups"]


def test_unclassified_counted_in_totals_not_groups(analysis):
    t = analysis["metrics"]["totals"]
    assert t["unclassified"] == 1             # the V7 TYPE row
    assert t["rows"] == 43                    # 44 parsed - 1 excluded (V2)
    n_sum = sum(g["n"] for g in analysis["metrics"]["groups"].values())
    assert n_sum + t["unclassified"] == t["rows"]


def test_validation_findings_fire(analysis):
    assert {"V2", "V3", "V4", "V5", "V6", "V7", "V8", "V9",
            "V10"} <= _codes(analysis)
    v10 = [f for f in analysis["findings"] if f["code"] == "V10"]
    assert any(f["message"].startswith("KOC PAID") for f in v10)
    # trivially-concentrated small groups are V9's job, not V10's
    assert not any(f["message"].startswith("TOP SOFT") for f in v10)


def test_dual_path_and_reconciliation_pass(analysis):
    assert not analysis["blocked"]            # WARNs never block


def test_dual_path_catches_divergence():
    rows_bytes = build_eff_bytes()
    from app.efficiency.analysis import classify, parse_report, validate
    cfg = ReportConfig()
    rows, findings, _ = parse_report(io.BytesIO(rows_bytes))
    classify(rows, cfg, findings)
    validate(rows, cfg, findings)
    primary = compute_metrics(rows, cfg)
    secondary = compute_metrics_pandas(rows, cfg)
    secondary["MID PAID"]["cpm_pooled"] += 1.0
    with pytest.raises(VerificationError):
        verify_dual_path(primary, secondary)


def test_block_policy_marks_blocked():
    cfg = ReportConfig(missing_row_policy="block")
    a = analyze(io.BytesIO(build_eff_bytes()), cfg)
    assert a["blocked"]
    assert any(f["severity"] == "ERROR" for f in a["findings"])


def test_fanbase_tier_mode():
    a = analyze(io.BytesIO(build_eff_bytes()), ReportConfig(tier_mode="fanbase"))
    g = a["metrics"]["groups"]
    # ts1 has 1500K fans → TOP either way; bs1 (250K) lands in BOT,
    # bs2 (150K) drops to KOC under thresholds
    assert g["TOP SOFT"]["n"] == 1
    assert g["KOC SOFT"]["n"] == 1


def test_v1_missing_columns_raises():
    wb = Workbook()
    ws = wb.active
    ws.append(["NAME", "POST LINK"])          # header found, columns missing
    ws.append(["x", "http://x.co/1"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    with pytest.raises(ValueError, match="V1"):
        analyze(buf)


def test_parse_report_closes_workbook_when_validation_raises(monkeypatch):
    from app.efficiency import analysis as analysis_module

    wb = Workbook()
    wb.active.append(["NAME", "POST LINK"])
    closed = False
    real_close = wb.close

    def record_close():
        nonlocal closed
        closed = True
        real_close()

    monkeypatch.setattr(wb, "close", record_close)
    monkeypatch.setattr(analysis_module, "load_workbook",
                        lambda *_args, **_kwargs: wb)

    with pytest.raises(ValueError, match="V1"):
        analysis_module.parse_report(io.BytesIO())
    assert closed


def test_header_probe_skips_early_partial_header():
    from tests.fixtures import EFF_HEADERS, EFF_PAID, _eff_row
    wb = Workbook()
    ws = wb.active
    ws.title = "MASTER KOL LIST"
    ws.append(["NAME", "POST LINK"])
    ws.append(EFF_HEADERS)
    ws.append(_eff_row(1, EFF_PAID, "KOC", "x", 20, "https://x/1",
                       100, 5, 2, 1, 8, 10))
    buf = io.BytesIO()
    wb.save(buf)
    result = analyze(io.BytesIO(buf.getvalue()))
    assert result["meta"]["header_row"] == 2


def test_report_finds_complete_schema_on_later_worksheet():
    from tests.fixtures import EFF_HEADERS, EFF_PAID, _eff_row
    wb = Workbook()
    wb.active.title = "Instructions"
    wb.active.append(["Read me first"])
    ws = wb.create_sheet("Campaign data")
    ws.append(EFF_HEADERS)
    ws.append(_eff_row(1, EFF_PAID, "KOC", "x", 20, "https://x/1",
                       100, 5, 2, 1, 8, 10))
    buf = io.BytesIO()
    wb.save(buf)

    result = analyze(io.BytesIO(buf.getvalue()))

    assert result["meta"]["sheet"] == "Campaign data"
    assert result["rows_total"] == 1


@pytest.mark.parametrize("field", ["CPM", "CPE"])
def test_invalid_diagnostics_only_values_warn_without_blocking(field):
    from tests.fixtures import EFF_HEADERS, EFF_PAID, _eff_row
    wb = Workbook()
    ws = wb.active
    ws.title = "MASTER KOL LIST"
    ws.append(EFF_HEADERS)
    row = _eff_row(1, EFF_PAID, "KOC", "x", 20, "https://x/1",
                   100, 5, 2, 1, 8, 10)
    row[EFF_HEADERS.index(field)] = "Infinity"
    ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)

    result = analyze(io.BytesIO(buf.getvalue()))

    assert not result["blocked"]
    assert result["metrics"]["totals"]["rows"] == 1
    assert any(f["code"] == "V11" and f["severity"] == "WARN"
               for f in result["findings"])


def test_invalid_fanbase_does_not_block_label_mode():
    from tests.fixtures import EFF_HEADERS, EFF_PAID, _eff_row
    wb = Workbook()
    ws = wb.active
    ws.title = "MASTER KOL LIST"
    ws.append(EFF_HEADERS)
    row = _eff_row(1, EFF_PAID, "KOC", "x", 20, "https://x/1",
                   100, 5, 2, 1, 8, 10)
    row[EFF_HEADERS.index("FAN BASE（K)")] = -1
    ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)

    result = analyze(io.BytesIO(buf.getvalue()), ReportConfig(tier_mode="label"))

    assert not result["blocked"]
    assert result["metrics"]["groups"]["KOC PAID"]["n"] == 1


@pytest.mark.parametrize("field,value", [
    ("IMPRESSION", -1),
    ("LIKE", 1.5),
    ("PRICE", -500),
    ("PRICE", "NaN"),
    ("PRICE", "Infinity"),
])
def test_invalid_numeric_domains_block_with_row_and_field(field, value):
    from tests.fixtures import EFF_HEADERS, EFF_PAID, _eff_row
    wb = Workbook()
    ws = wb.active
    ws.title = "MASTER KOL LIST"
    ws.append(EFF_HEADERS)
    row = _eff_row(1, EFF_PAID, "KOC", "x", 20, "https://x/1",
                   100, 5, 2, 1, 8, 10)
    row[EFF_HEADERS.index(field)] = value
    ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    with pytest.raises(ValueError, match=rf"V11:.*{field}.*Excel rows 2"):
        analyze(io.BytesIO(buf.getvalue()))


def test_empty_and_all_missing_workbooks_are_user_validation_errors():
    from tests.fixtures import EFF_HEADERS, EFF_PAID, _eff_row
    for add_row in (False, True):
        wb = Workbook()
        ws = wb.active
        ws.title = "MASTER KOL LIST"
        ws.append(EFF_HEADERS)
        if add_row:
            ws.append(_eff_row(1, EFF_PAID, "KOC", "x", 20,
                               "https://x/1", None, None, None, None,
                               None, None))
        buf = io.BytesIO()
        wb.save(buf)
        with pytest.raises(ValueError, match="V2:.*no analyzable"):
            analyze(io.BytesIO(buf.getvalue()))


def test_efficiency_row_limit_is_enforced(monkeypatch):
    from app import config
    monkeypatch.setattr(config, "MAX_EFFICIENCY_ROWS", 1)
    with pytest.raises(ValueError, match="more than 1 efficiency rows"):
        analyze(io.BytesIO(build_eff_bytes()))


def test_data_after_more_than_200_blank_rows_is_not_silently_dropped():
    from tests.fixtures import EFF_HEADERS, EFF_PAID, _eff_row
    wb = Workbook()
    ws = wb.active
    ws.title = "MASTER KOL LIST"
    ws.append(EFF_HEADERS)
    ws.append(_eff_row(1, EFF_PAID, "KOC", "first", 20, "https://x/1",
                       100, 5, 2, 1, 8, 10))
    for _ in range(205):
        ws.append([])
    ws.append(_eff_row(2, EFF_PAID, "KOC", "later", 20, "https://x/2",
                       100, 5, 2, 1, 8, 10))
    # A far-away styled cell must not turn this into a million-row loop.
    ws.cell(row=1_000_000, column=30).number_format = "0.00"
    buf = io.BytesIO()
    wb.save(buf)
    result = analyze(io.BytesIO(buf.getvalue()))
    assert result["rows_total"] == 2


# ----------------------------------------------------------------- insights

def test_insights_are_data_driven_and_never_cross_wave(analysis):
    ins = analysis["insights"]
    joined = " ".join(ins["price"] + ins["efficiency"] + ins["caveats"]
                      + [ins["footnote"]])
    # the hard rule: no fabricated wave-over-wave comparisons
    for banned in ("WAVE", "PREVIOUS", "LAST ROUND", "VS. PRIOR", "环比", "上一波"):
        assert banned not in joined.upper()
    assert any("MID" in b for b in ins["price"])       # premium exists (24k>6k… wait 15k avg)
    assert ins["footnote"].startswith("Basis: pooled")


def test_caveats_cover_v9_and_v10(analysis):
    caveats = " ".join(analysis["insights"]["caveats"])
    assert "TOP SOFT = 1 POST(S) ONLY" in caveats
    assert "KOC PAID" in caveats and "VIRAL" in caveats


def test_per_post_basis_footnote():
    a = analyze(io.BytesIO(build_eff_bytes()), ReportConfig(basis="per_post"))
    assert a["insights"]["footnote"].startswith("Basis: per-post")


# --------------------------------------------------------------------- deck

def test_donut_order_covers_every_group():
    """DONUT_ORDER filters which groups the donut plots — a TIER × COOP combo
    missing from it silently vanishes from the chart (the reference file has
    no TOP PAID, which hid exactly that)."""
    from app.efficiency.deck import DONUT_COLORS, DONUT_ORDER
    from app.efficiency.analysis import COOPS, TIERS
    every = {f"{t} {c}" for t in TIERS for c in COOPS}
    assert set(DONUT_ORDER) == every
    assert set(DONUT_COLORS) == every


def test_build_deck_and_chart_cache(analysis):
    pptx = build_deck(analysis)
    assert_chart_cache(pptx, analysis)        # must not raise
    with zipfile.ZipFile(io.BytesIO(pptx)) as z:
        charts = [n for n in z.namelist()
                  if n.startswith("ppt/charts/chart") and n.endswith(".xml")]
        assert len(charts) == 4               # donut, price, CPM, CPE
        donut = min(charts)
        xml = z.read(donut).decode()
    assert 'val="58"' in xml                  # hole size
    assert "showLeaderLines" in xml
    assert 'formatCode=\'0.0"%"\'' in xml or 'formatCode="0.0&quot;%&quot;"' in xml \
        or '0.0"%"' in xml                    # share number format present


def test_slide_geometry_is_integer_emu(analysis):
    """EMU coordinates are xsd integers. A float (cx="3511296.0" — produced
    by true-dividing an Emu) renders fine in LibreOffice but makes PowerPoint
    'repair' the deck by DELETING the chart frame."""
    import re
    pptx = build_deck(analysis)
    with zipfile.ZipFile(io.BytesIO(pptx)) as z:
        slide = z.read("ppt/slides/slide1.xml").decode()
    floats = re.findall(r'(?:\b[xy]|\bc[xy])="-?\d+\.\d*"', slide)
    assert not floats, f"non-integer EMU coordinates: {floats}"


def test_donut_perpoint_labels_carry_numfmt(analysis):
    """PowerPoint does not reliably inherit the ser-level numFmt into a
    per-point c:dLbl override — without its own, the sliver/dark labels
    render as bare numbers ("7.9") instead of "7.9%"."""
    pptx = build_deck(analysis)
    with zipfile.ZipFile(io.BytesIO(pptx)) as z:
        donut = min(n for n in z.namelist()
                    if n.startswith("ppt/charts/chart") and n.endswith(".xml"))
        xml = z.read(donut).decode()
    import html
    import re
    labels = re.findall(r"<c:dLbl>.*?</c:dLbl>", xml, re.S)
    assert labels, "expected per-point label overrides in the donut"
    for dlbl in labels:
        assert "numFmt" in dlbl and '0.0"%"' in html.unescape(dlbl), dlbl[:200]


def test_chart_cache_catches_tampering(analysis):
    pptx = build_deck(analysis)
    import copy
    tampered = copy.deepcopy(analysis)
    tampered["metrics"]["groups"]["MID PAID"]["avg_price"] += 1
    with pytest.raises(VerificationError):
        assert_chart_cache(pptx, tampered)


def _rewrite_chart(pptx: bytes, chart_number: int, mutate) -> bytes:
    from lxml import etree
    source = io.BytesIO(pptx)
    output = io.BytesIO()
    target_name = f"ppt/charts/chart{chart_number}.xml"
    with zipfile.ZipFile(source) as zin, zipfile.ZipFile(output, "w") as zout:
        for info in zin.infolist():
            payload = zin.read(info.filename)
            if info.filename == target_name:
                root = etree.fromstring(payload)
                mutate(root)
                payload = etree.tostring(root, xml_declaration=True,
                                         encoding="UTF-8")
            zout.writestr(info, payload)
    return output.getvalue()


def test_chart_cache_rejects_missing_trailing_point(analysis):
    from app.efficiency.deck import C_NS
    ns = f"{{{C_NS}}}"

    def remove_last(root):
        cache = root.find(f".//{ns}val/{ns}numRef/{ns}numCache")
        points = cache.findall(f"{ns}pt")
        cache.remove(points[-1])
        count = cache.find(f"{ns}ptCount")
        count.set("val", str(int(count.get("val")) - 1))

    tampered = _rewrite_chart(build_deck(analysis), 2, remove_last)
    with pytest.raises(VerificationError, match="expected 4 points, found 3"):
        assert_chart_cache(tampered, analysis)


def test_chart_cache_rejects_category_reordering(analysis):
    from app.efficiency.deck import C_NS
    ns = f"{{{C_NS}}}"

    def rename_category(root):
        value = root.find(
            f".//{ns}cat/{ns}strRef/{ns}strCache/{ns}pt/{ns}v")
        value.text = "WRONG"

    tampered = _rewrite_chart(build_deck(analysis), 2, rename_category)
    with pytest.raises(VerificationError, match="categories"):
        assert_chart_cache(tampered, analysis)


def test_chart_cache_checks_categories_on_every_series(analysis):
    from app.efficiency.deck import C_NS
    ns = f"{{{C_NS}}}"

    def rename_second_series_category(root):
        series = root.findall(f".//{ns}ser")
        value = series[1].find(
            f"{ns}cat/{ns}strRef/{ns}strCache/{ns}pt/{ns}v")
        value.text = "WRONG"

    tampered = _rewrite_chart(
        build_deck(analysis), 2, rename_second_series_category)
    with pytest.raises(VerificationError, match="categories"):
        assert_chart_cache(tampered, analysis)


def test_deck_zh_language():
    a = analyze(io.BytesIO(build_eff_bytes()), ReportConfig(language="zh"))
    pptx = build_deck(a)
    assert_chart_cache(pptx, a)
    joined = " ".join(a["insights"]["price"] + a["insights"]["efficiency"]
                      + a["insights"]["caveats"] + [a["insights"]["footnote"]])
    for english in ("WINNER", "COMPARISON", "CAUTION", "POST(S)",
                    "BASIS", "RESULTS CONCENTRATED"):
        assert english not in joined.upper()


def test_insight_uses_tie_when_displayed_values_are_equal():
    from app.efficiency.analysis import build_insights
    cfg = ReportConfig()
    groups = {
        "KOC PAID": {"tier": "KOC", "n": 3, "avg_price": 1000,
                     "cpm_pooled": 10.44, "cpe_pooled": 2.04,
                     "cpm_perpost": 10.44, "cpe_perpost": 2.04},
        "KOC SOFT": {"tier": "KOC", "n": 3, "avg_price": 1000,
                     "cpm_pooled": 10.45, "cpe_pooled": 2.049,
                     "cpm_perpost": 10.45, "cpe_perpost": 2.049},
    }
    insights = build_insights(
        {"groups": groups, "totals": {"rows": 6}}, cfg, [])
    assert "TIE ¥10" in insights["efficiency"][0]
    assert "TIE ¥2.0" in insights["efficiency"][1]


# ----------------------------------------------------- golden (real file)

REAL = Path(__file__).resolve().parents[2] / "data" / "real" / "PLOG_DMR_CHECK.xlsx"


@pytest.mark.skipif(not REAL.exists(), reason="real client workbook not present")
def test_golden_real_workbook():
    a = analyze(str(REAL), ReportConfig())
    t = a["metrics"]["totals"]
    assert t["spend"] == pytest.approx(1_049_345)
    assert t["impressions"] == 17_820_424
    assert t["engagements"] == 216_300
    g = a["metrics"]["groups"]
    assert set(g) == {"TOP SOFT", "MID PAID", "MID SOFT", "BOT PAID",
                      "BOT SOFT", "KOC PAID", "KOC SOFT"}
    rounded_cpm = {k: round(v["cpm_pooled"]) for k, v in g.items()}
    assert rounded_cpm == {"TOP SOFT": 82, "MID PAID": 63, "MID SOFT": 103,
                           "BOT PAID": 48, "BOT SOFT": 117, "KOC PAID": 14,
                           "KOC SOFT": 88}
    rounded_cpe = {k: round(v["cpe_pooled"], 1) for k, v in g.items()}
    assert rounded_cpe == {"TOP SOFT": 2.8, "MID PAID": 6.7, "MID SOFT": 5.1,
                           "BOT PAID": 3.9, "BOT SOFT": 8.6, "KOC PAID": 1.5,
                           "KOC SOFT": 5.8}
    pptx = build_deck(a)
    assert_chart_cache(pptx, a)


def test_v8_with_missing_fanbase_does_not_crash():
    """尾部+底部 coexisting while either set lacks FAN BASE values used to
    raise min([]) — the range now degrades to '?'."""
    from datetime import datetime
    from tests.fixtures import EFF_HEADERS, EFF_PAID
    wb = Workbook()
    ws = wb.active
    ws.title = "MASTER KOL LIST"
    ws.append(EFF_HEADERS)
    for no, level, fan in ((1, "尾部", None), (2, "尾部", None),
                           (3, "底部", 150), (4, "底部", 180)):
        ws.append([no, "", "W1", EFF_PAID, level, f"kol{no}", fan,
                   datetime(2026, 6, no), "", f"http://xhslink.com/v8{no}",
                   50000, 1000, 100, 50, 1150, 5000, None, None])
    buf = io.BytesIO()
    wb.save(buf)
    a = analyze(io.BytesIO(buf.getvalue()), ReportConfig())
    v8 = [f for f in a["findings"] if f["code"] == "V8"]
    assert len(v8) == 1
    assert "fans ?K" in v8[0]["message"]
    assert "150–180" in v8[0]["message"]     # the populated side still shows


def test_share_tolerance_scales_with_slice_count():
    """Nine slices rounded to 1 decimal can legitimately drift up to 0.45
    from 100; the old fixed 0.3 tolerance rejected valid data (e.g. counts
    22/62/72/25/25/45/25/15/12 → 100.4)."""
    from app.efficiency.analysis import verify_reconciliation
    counts = [22, 62, 72, 25, 25, 45, 25, 15, 12]
    total = sum(counts)
    groups = {f"G{i}": {"n": c, "spend": 0.0,
                        "share": round(c / total * 100, 1)}
              for i, c in enumerate(counts)}
    metrics = {"groups": groups,
               "totals": {"rows": total, "spend": 0.0,
                          "unclassified": 0, "unclassified_share": 0.0}}
    assert abs(sum(g["share"] for g in groups.values()) - 100.0) > 0.3
    verify_reconciliation([], metrics)       # must NOT raise
    # a genuinely broken share table still trips it
    groups["G0"]["share"] += 2.0
    with pytest.raises(VerificationError):
        verify_reconciliation([], metrics)


# ------------------------------------------------- LEVEL → FAN BASE fallback

def _level_fallback_wb(rows):
    from datetime import datetime
    from tests.fixtures import EFF_HEADERS, EFF_PAID
    wb = Workbook()
    ws = wb.active
    ws.title = "MASTER KOL LIST"
    ws.append(EFF_HEADERS)
    for no, (level, fan) in enumerate(rows, start=1):
        ws.append([no, "", "W1", EFF_PAID, level, f"kol{no}", fan,
                   datetime(2026, 6, no), "", f"http://xhslink.com/fb{no}",
                   50000, 1000, 100, 50, 1150, 5000, None, None])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_missing_level_falls_back_to_fanbase_bands():
    """Fallback keeps the established inclusive 200/400/1000 lower bounds."""
    a = analyze(io.BytesIO(_level_fallback_wb([
        ("", 150),                     # KOC
        ("", 200), ("待定", 399),       # BOT
        ("?", 400), ("", 999),         # MID
        ("", 1000),                    # TOP
    ])), ReportConfig())
    g = a["metrics"]["groups"]
    assert g["KOC PAID"]["n"] == 1
    assert g["BOT PAID"]["n"] == 2
    assert g["MID PAID"]["n"] == 2
    assert g["TOP PAID"]["n"] == 1
    assert a["metrics"]["totals"]["unclassified"] == 0
    v12 = [f for f in a["findings"] if f["code"] == "V12"]
    assert len(v12) == 1 and len(v12[0]["rows"]) == 6
    assert not [f for f in a["findings"] if f["code"] == "V11"]
    assert any("V12: FAN BASE FALLBACK" in c
               for c in a["insights"]["caveats"])
    # …and the finding translates
    from app.i18n import make_td
    zh = make_td("zh")(v12[0]["message"])
    assert "FAN BASE 阈值分层" in zh


def test_unknown_nonplaceholder_level_does_not_fall_back():
    a = analyze(io.BytesIO(_level_fallback_wb([("头布", 500)])),
                ReportConfig())
    assert a["metrics"]["totals"]["unclassified"] == 1
    assert [f for f in a["findings"] if f["code"] == "V7"]
    assert not [f for f in a["findings"] if f["code"] == "V12"]


def test_missing_level_and_fanbase_stays_unclassified():
    a = analyze(io.BytesIO(_level_fallback_wb([("", None), ("头部", 1200)])),
                ReportConfig())
    assert a["metrics"]["totals"]["unclassified"] == 1     # no fallback signal
    assert not [f for f in a["findings"] if f["code"] == "V12"]
    assert [f for f in a["findings"] if f["code"] == "V7"]


def test_explicit_level_wins_over_fanbase():
    """The fallback NEVER overrides a recognizable label — a 头部 row with
    tiny fan base stays TOP."""
    a = analyze(io.BytesIO(_level_fallback_wb([("头部", 50)])), ReportConfig())
    assert a["metrics"]["groups"]["TOP PAID"]["n"] == 1


# ------------------------------------------------ FAN BASE workbook-level unit

def test_fanbase_unit_k_never_switches_by_magnitude():
    """The default follows the column's K contract at every magnitude."""
    a = analyze(io.BytesIO(_level_fallback_wb([
        ("", 9999), ("", 10000), ("", 450000),
    ])), ReportConfig())
    g = a["metrics"]["groups"]
    assert g["TOP PAID"]["n"] == 3
    assert a["config"]["fanbase_unit"] == "k"
    assert not [f for f in a["findings"] if f["code"] == "V13"]


def test_fanbase_unit_raw_converts_every_valid_row_and_reaches_deck():
    a = analyze(io.BytesIO(_level_fallback_wb([
        ("", 130), ("", 1741), ("", 9999),
        ("", 10000), ("", 450000), ("", 1300000),
    ])), ReportConfig(fanbase_unit="raw"))
    g = a["metrics"]["groups"]
    assert g["KOC PAID"]["n"] == 4
    assert g["MID PAID"]["n"] == 1
    assert g["TOP PAID"]["n"] == 1
    assert a["config"]["fanbase_unit"] == "raw"
    v13 = [f for f in a["findings"] if f["code"] == "V13"]
    assert len(v13) == 1 and len(v13[0]["rows"]) == 6
    assert any("V13: FAN BASE UNIT RAW FOLLOWERS" in c
               for c in a["insights"]["caveats"])

    from app.i18n import make_td
    zh = make_td("zh")(v13[0]["message"])
    assert "原始粉丝数" in zh and "除以 1,000" in zh

    pptx = build_deck(a)
    assert_chart_cache(pptx, a)
    with zipfile.ZipFile(io.BytesIO(pptx)) as archive:
        slide = archive.read("ppt/slides/slide1.xml").decode()
    assert "V12: FAN BASE FALLBACK USED" in slide
    assert "V13: FAN BASE UNIT RAW FOLLOWERS" in slide


def test_fanbase_fallback_and_unit_caveats_translate_for_chinese_deck():
    a = analyze(io.BytesIO(_level_fallback_wb([("", 450000)])),
                ReportConfig(fanbase_unit="raw", language="zh"))
    caveats = " ".join(a["insights"]["caveats"])
    assert "回退分层" in caveats
    assert "原始粉丝数" in caveats
    assert "FAN BASE FALLBACK USED" not in caveats


@pytest.mark.parametrize("fanbase", ["NaN", "Infinity", -1, 10000.5])
def test_raw_unit_does_not_rescue_invalid_fanbase(fanbase):
    a = analyze(io.BytesIO(_level_fallback_wb([("", fanbase)])),
                ReportConfig(fanbase_unit="raw"))
    assert not a["blocked"]
    assert a["metrics"]["totals"]["unclassified"] == 1
    v11 = [f for f in a["findings"] if f["code"] == "V11"]
    assert len(v11) == 1 and "invalid diagnostics-only" in v11[0]["message"]
    assert not [f for f in a["findings"]
                if f["code"] in {"V12", "V13"}]


def test_fanbase_raw_unit_feeds_fanbase_mode_too():
    a = analyze(io.BytesIO(_level_fallback_wb([("头部", 450000)])),
                ReportConfig(tier_mode="fanbase", fanbase_unit="raw"))
    assert a["metrics"]["groups"]["MID PAID"]["n"] == 1   # 450K → MID


def test_unknown_fanbase_unit_is_rejected():
    with pytest.raises(ValueError, match="fanbase_unit"):
        analyze(io.BytesIO(_level_fallback_wb([("", 500)])),
                ReportConfig(fanbase_unit="guess"))
