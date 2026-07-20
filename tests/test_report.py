"""The annotated export must reproduce the reference layout: columns A–R
untouched, S = human vocabulary with no header, evidence in T+."""
from __future__ import annotations

from openpyxl import load_workbook

from app.matcher import run_pipeline
from app.parsers import parse_dmr, parse_plog
from app.report import write_annotated_xlsx
from tests import fixtures


def test_annotated_export(plog_path, dmr_path, fake_resolver, tmp_path):
    plog = parse_plog(plog_path)
    dmr = parse_dmr(dmr_path)
    verdicts = run_pipeline(plog, dmr)
    out = tmp_path / "annotated.xlsx"
    write_annotated_xlsx(plog_path, str(out), verdicts,
                         header_row=plog.header_row)

    orig = load_workbook(plog_path)["MASTER KOL LIST"]
    ann = load_workbook(str(out))["MASTER KOL LIST"]

    # Columns A–R byte-identical values
    for row in orig.iter_rows(min_col=1, max_col=18):
        for cell in row:
            assert ann.cell(row=cell.row, column=cell.column).value == cell.value

    # S has no header (the reference leaves S1 blank)
    assert ann.cell(row=plog.header_row, column=19).value in (None, "")
    # evidence headers start at T
    assert ann.cell(row=plog.header_row, column=20).value == "STATUS"

    by = {(v.campaign, v.no): v for v in verdicts}
    # matched row → blank S
    v = by[("PLOG #001", "1")]
    assert ann.cell(row=v.excel_row, column=19).value in (None, "")
    # mislabel row → the human's exact vocabulary
    v = by[("PLOG #001", "3")]
    assert ann.cell(row=v.excel_row, column=19).value == "有 但是DMR博主名字标注错误"
    # no-post row → 无帖子
    v = by[("PLOG #002", "1")]
    assert ann.cell(row=v.excel_row, column=19).value == "无帖子"
    # dead link row → Check链接错误 + candidate note
    v = by[("PLOG #001", "4")]
    s = ann.cell(row=v.excel_row, column=19).value
    assert s.startswith("Check链接错误") and fixtures.N_JITUI in s


def test_override_wins_in_export(plog_path, dmr_path, fake_resolver, tmp_path):
    plog = parse_plog(plog_path)
    dmr = parse_dmr(dmr_path)
    verdicts = run_pipeline(plog, dmr)
    out = tmp_path / "annotated.xlsx"
    v = next(x for x in verdicts if (x.campaign, x.no) == ("PLOG #002", "1"))
    overrides = {v.excel_row: {"status": "无博主", "note": "human says so"}}
    write_annotated_xlsx(plog_path, str(out), verdicts,
                         header_row=plog.header_row, sheet_name=plog.sheet,
                         overrides=overrides)
    ann = load_workbook(str(out))["MASTER KOL LIST"]
    assert ann.cell(row=v.excel_row, column=19).value == "无博主"


def test_match_blank_override_forces_empty_s(plog_path, dmr_path, fake_resolver,
                                             tmp_path):
    from app.report import OVERRIDE_MATCH_BLANK
    plog = parse_plog(plog_path)
    dmr = parse_dmr(dmr_path)
    verdicts = run_pipeline(plog, dmr)
    out = tmp_path / "annotated.xlsx"
    # the mislabel row would normally carry 有 但是DMR博主名字标注错误
    v = next(x for x in verdicts if (x.campaign, x.no) == ("PLOG #001", "3"))
    overrides = {v.excel_row: {"status": OVERRIDE_MATCH_BLANK, "note": ""}}
    write_annotated_xlsx(plog_path, str(out), verdicts,
                         header_row=plog.header_row, sheet_name=plog.sheet,
                         overrides=overrides)
    ann = load_workbook(str(out))["MASTER KOL LIST"]
    assert ann.cell(row=v.excel_row, column=19).value in (None, "")
