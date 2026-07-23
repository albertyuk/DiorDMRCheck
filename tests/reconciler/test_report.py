"""The annotated export must reproduce the reference layout: columns A–R
untouched, S = human vocabulary with no header, evidence in T+."""
from __future__ import annotations

from openpyxl import load_workbook

from app.reconciler.pipeline import run_pipeline
from app.reconciler.parsers import parse_dmr, parse_plog
from app.reconciler.export import EVIDENCE_HEADERS, write_annotated_xlsx
from tests import fixtures


def test_annotated_export(plog_path, dmr_path, fake_resolver, tmp_path):
    plog = parse_plog(plog_path)
    dmr = parse_dmr(dmr_path)
    verdicts = run_pipeline(plog, dmr)
    out = tmp_path / "annotated.xlsx"
    write_annotated_xlsx(plog_path, str(out), verdicts,
                         header_row=plog.header_row)

    orig = load_workbook(plog_path)["MASTER KOL LIST"]
    ann = load_workbook(str(out))["MASTER KOL LIST"]

    # Columns A–R byte-identical values
    for row in orig.iter_rows(min_col=1, max_col=18):
        for cell in row:
            assert ann.cell(row=cell.row, column=cell.column).value == cell.value

    # S has no header (the reference leaves S1 blank)
    assert ann.cell(row=plog.header_row, column=19).value in (None, "")
    # evidence headers start at T
    assert ann.cell(row=plog.header_row, column=20).value == "STATUS"

    by = {(v.campaign, v.no): v for v in verdicts}
    # matched row → blank S
    v = by[("PLOG #001", "1")]
    assert ann.cell(row=v.excel_row, column=19).value in (None, "")
    # mislabel row → the human's exact vocabulary
    v = by[("PLOG #001", "3")]
    assert ann.cell(row=v.excel_row, column=19).value == "有 但是DMR博主名字标注错误"
    # no-post row → 无帖子
    v = by[("PLOG #002", "1")]
    assert ann.cell(row=v.excel_row, column=19).value == "无帖子"
    # dead link row → Check链接错误 + candidate note
    v = by[("PLOG #001", "4")]
    s = ann.cell(row=v.excel_row, column=19).value
    assert s.startswith("Check链接错误") and fixtures.N_JITUI in s


def test_override_wins_in_export(plog_path, dmr_path, fake_resolver, tmp_path):
    plog = parse_plog(plog_path)
    dmr = parse_dmr(dmr_path)
    verdicts = run_pipeline(plog, dmr)
    out = tmp_path / "annotated.xlsx"
    v = next(x for x in verdicts if (x.campaign, x.no) == ("PLOG #002", "1"))
    overrides = {v.excel_row: {"status": "无博主", "note": "human says so"}}
    write_annotated_xlsx(plog_path, str(out), verdicts,
                         header_row=plog.header_row, sheet_name=plog.sheet,
                         overrides=overrides)
    ann = load_workbook(str(out))["MASTER KOL LIST"]
    assert ann.cell(row=v.excel_row, column=19).value == "无博主"


def test_match_blank_override_forces_empty_s(plog_path, dmr_path, fake_resolver,
                                             tmp_path):
    from app.reconciler.export import OVERRIDE_MATCH_BLANK
    plog = parse_plog(plog_path)
    dmr = parse_dmr(dmr_path)
    verdicts = run_pipeline(plog, dmr)
    out = tmp_path / "annotated.xlsx"
    # the mislabel row would normally carry 有 但是DMR博主名字标注错误
    v = next(x for x in verdicts if (x.campaign, x.no) == ("PLOG #001", "3"))
    overrides = {v.excel_row: {"status": OVERRIDE_MATCH_BLANK, "note": ""}}
    write_annotated_xlsx(plog_path, str(out), verdicts,
                         header_row=plog.header_row, sheet_name=plog.sheet,
                         overrides=overrides)
    ann = load_workbook(str(out))["MASTER KOL LIST"]
    assert ann.cell(row=v.excel_row, column=19).value in (None, "")


def test_load_verdicts_tolerates_legacy_and_derived_keys():
    """Stored result documents carry derived fields (column_s) and may
    predate the schema (per-row engagement_caveat, unknown future keys) —
    rehydration must drop them, never TypeError on historical runs."""
    import json
    from app.reconciler.export import load_verdicts
    legacy_verdict = {
        "campaign": "C", "no": "1", "name": "n", "post_date": None,
        "post_link": "http://x", "excel_row": 2, "status": "MATCH",
        "column_s": "",                                # derived
        "engagement_caveat": "old per-row caveat",     # pre-schema
        "some_future_field": 123,                      # forward compat
        "candidates": [{
            "dmr_row": 5, "post_id": "p", "blogger": "b", "username": "u",
            "post_date": None, "date_delta_days": None,
            "likes_retweet": None, "name_method": "cjk-substring",
            "novel_candidate_key": True,               # forward compat
        }],
    }
    run = {"result_json": json.dumps({"verdicts": [legacy_verdict]})}
    (v,) = load_verdicts(run)
    assert v.status == "MATCH" and v.candidates[0].post_id == "p"


def test_audit_json_has_document_level_caveat(plog_path, dmr_path,
                                              fake_resolver):
    import json
    from app.reconciler.domain import ENGAGEMENT_CAVEAT
    from app.reconciler.export import build_audit_json
    plog = parse_plog(plog_path)
    dmr = parse_dmr(dmr_path)
    verdicts = run_pipeline(plog, dmr)
    doc = json.loads(build_audit_json(
        {"id": "t", "summary_json": None}, verdicts, {}, {}, {}, []))
    assert doc["engagement_caveat"] == ENGAGEMENT_CAVEAT
    assert all("engagement_caveat" not in v for v in doc["verdicts"])


# ------------------------------------------- never overwrite populated cells

def _prefilled_copy(plog_path, tmp_path, s_edits=None, col_edits=None):
    """Copy the fixture PLOG and pre-populate cells like earlier human work."""
    import shutil
    p = tmp_path / "prefilled.xlsx"
    shutil.copy(plog_path, p)
    wb = load_workbook(str(p))
    ws = wb["MASTER KOL LIST"]
    for row, value in (s_edits or {}).items():
        ws.cell(row=row, column=19, value=value)
    for (row, col), value in (col_edits or {}).items():
        ws.cell(row=row, column=col, value=value)
    wb.save(str(p))
    return str(p)


def test_prefilled_s_cells_are_never_overwritten(plog_path, dmr_path,
                                                 fake_resolver, tmp_path):
    plog = parse_plog(plog_path)
    dmr = parse_dmr(dmr_path)
    verdicts = run_pipeline(plog, dmr)
    by = {(v.campaign, v.no): v for v in verdicts}
    matched_row = by[("PLOG #001", "1")].excel_row      # pipeline wants blank
    nopost_row = by[("PLOG #002", "1")].excel_row       # pipeline wants 无帖子

    src = _prefilled_copy(plog_path, tmp_path, s_edits={
        matched_row: "人工已确认OK",
        nopost_row: "无帖子",                            # human agrees
    })
    out = tmp_path / "ann.xlsx"
    write_annotated_xlsx(src, str(out), verdicts, header_row=plog.header_row)
    ann = load_workbook(str(out))["MASTER KOL LIST"]

    assert ann.cell(row=matched_row, column=19).value == "人工已确认OK"
    assert ann.cell(row=nopost_row, column=19).value == "无帖子"
    # disagreement is noted in the evidence Notes column, agreement is not
    n_cols = len(EVIDENCE_HEADERS)
    notes_col = 20 + n_cols - 1
    notes = ann.cell(row=matched_row, column=notes_col).value or ""
    assert "人工已确认OK" in notes and "kept" in notes
    status = ann.cell(row=matched_row, column=20).value or ""
    assert "(S kept from source)" in status
    agree_notes = ann.cell(row=nopost_row, column=notes_col).value or ""
    assert "kept" not in agree_notes
    # untouched rows still get the pipeline verdict as usual
    mislabel_row = by[("PLOG #001", "3")].excel_row
    assert ann.cell(row=mislabel_row, column=19).value == "有 但是DMR博主名字标注错误"


def test_prefilled_s_cells_preserve_value_type_and_format(
        plog_path, dmr_path, fake_resolver, tmp_path):
    from datetime import date

    plog = parse_plog(plog_path)
    verdicts = run_pipeline(plog, parse_dmr(dmr_path))
    rows = [verdict.excel_row for verdict in verdicts[:3]]
    source = _prefilled_copy(plog_path, tmp_path, s_edits={
        rows[0]: 123.5,
        rows[1]: date(2025, 4, 3),
        rows[2]: True,
    })
    source_wb = load_workbook(source, data_only=False)
    source_ws = source_wb[plog.sheet]
    source_ws.cell(rows[0], 19).number_format = "0.00"
    source_ws.cell(rows[1], 19).number_format = "yyyy-mm-dd"
    source_wb.save(source)
    source_wb = load_workbook(source, data_only=False)
    before = [
        (
            source_wb[plog.sheet].cell(row, 19).value,
            source_wb[plog.sheet].cell(row, 19).data_type,
            source_wb[plog.sheet].cell(row, 19).number_format,
        )
        for row in rows
    ]

    out = tmp_path / "typed-s.xlsx"
    write_annotated_xlsx(
        source, str(out), verdicts, plog.header_row, plog.sheet
    )
    exported = load_workbook(out, data_only=False)[plog.sheet]
    after = [
        (
            exported.cell(row, 19).value,
            exported.cell(row, 19).data_type,
            exported.cell(row, 19).number_format,
        )
        for row in rows
    ]

    assert after == before


def test_ui_override_still_wins_over_prefilled_s(plog_path, dmr_path,
                                                 fake_resolver, tmp_path):
    plog = parse_plog(plog_path)
    verdicts = run_pipeline(plog, parse_dmr(dmr_path))
    by = {(v.campaign, v.no): v for v in verdicts}
    row = by[("PLOG #002", "1")].excel_row
    src = _prefilled_copy(plog_path, tmp_path, s_edits={row: "旧的人工备注"})
    out = tmp_path / "ann.xlsx"
    write_annotated_xlsx(src, str(out), verdicts, header_row=plog.header_row,
                         overrides={row: {"status": "无博主", "note": ""}})
    ann = load_workbook(str(out))["MASTER KOL LIST"]
    assert ann.cell(row=row, column=19).value == "无博主"   # explicit action wins


def test_evidence_block_shifts_past_populated_columns(plog_path, dmr_path,
                                                      fake_resolver, tmp_path):
    plog = parse_plog(plog_path)
    verdicts = run_pipeline(plog, parse_dmr(dmr_path))
    some_row = verdicts[0].excel_row
    # column T (20) already used by the human for their own notes
    src = _prefilled_copy(plog_path, tmp_path,
                          col_edits={(some_row, 20): "我的备注，别动"})
    out = tmp_path / "ann.xlsx"
    write_annotated_xlsx(src, str(out), verdicts, header_row=plog.header_row)
    ann = load_workbook(str(out))["MASTER KOL LIST"]
    assert ann.cell(row=some_row, column=20).value == "我的备注，别动"
    assert ann.cell(row=plog.header_row, column=20).value in (None, "")
    assert ann.cell(row=plog.header_row, column=21).value == "STATUS"


def test_export_neutralizes_formula_capable_external_text(
        plog_path, dmr_path, fake_resolver, tmp_path):
    plog = parse_plog(plog_path)
    dmr = parse_dmr(dmr_path)
    verdicts = run_pipeline(plog, dmr)
    verdict = verdicts[0]
    verdict.matched_blogger = '=HYPERLINK("https://evil.invalid","click")'
    verdict.llm_rationale_en = "+SUM(1,1)"
    verdict.notes = ["@malicious"]
    overrides = {
        verdict.excel_row: {"status": "无帖子", "note": "-1+1"}
    }
    out = tmp_path / "safe.xlsx"
    write_annotated_xlsx(
        plog_path, str(out), verdicts, plog.header_row, plog.sheet, overrides
    )
    ws = load_workbook(out, data_only=False)[plog.sheet]
    for column in (23, 32, 34):  # W blogger, AF rationale, AH notes
        cell = ws.cell(verdict.excel_row, column)
        assert cell.data_type == "s"
        assert cell.value.startswith("'")


def test_export_strips_ooxml_illegal_control_characters(
        plog_path, dmr_path, fake_resolver, tmp_path):
    plog = parse_plog(plog_path)
    verdicts = run_pipeline(plog, parse_dmr(dmr_path))
    verdict = verdicts[0]
    verdict.notes = ["left\x00middle\x0bright"]
    overrides = {
        verdict.excel_row: {"status": "无帖子", "note": "human\x01note"}
    }
    out = tmp_path / "control-safe.xlsx"

    write_annotated_xlsx(
        plog_path, str(out), verdicts, plog.header_row, plog.sheet, overrides
    )

    notes = load_workbook(out)[plog.sheet].cell(verdict.excel_row, 34).value
    assert notes == "leftmiddleright | humannote"


def test_audit_json_uses_effective_override_counts_and_retains_pipeline(
        plog_path, dmr_path, fake_resolver):
    import json
    from app.reconciler.export import build_audit_json
    from app.reconciler.pipeline import status_counts

    plog = parse_plog(plog_path)
    dmr = parse_dmr(dmr_path)
    verdicts = run_pipeline(plog, dmr)
    verdict = next(v for v in verdicts if v.status == "NO_POST")
    overrides = {
        verdict.excel_row: {"status": "无博主", "note": "reviewed"}
    }
    doc = json.loads(build_audit_json(
        {"id": "t", "summary_json": None}, verdicts,
        status_counts(verdicts), {}, {}, [], overrides=overrides,
    ))
    effective = next(v for v in doc["verdicts"]
                     if v["excel_row"] == verdict.excel_row)
    assert effective["status"] == "NO_BLOGGER"
    assert effective["pipeline_status"] == "NO_POST"
    assert doc["counts"]["NO_BLOGGER"] >= 1
    assert doc["counts"].get("NO_POST", 0) == doc["pipeline_counts"]["NO_POST"] - 1
    assert doc["summary_basis"] == "pipeline_before_human_overrides"


def test_invalid_legacy_override_is_ignored_not_executed(
        plog_path, dmr_path, fake_resolver, tmp_path):
    plog = parse_plog(plog_path)
    verdicts = run_pipeline(plog, parse_dmr(dmr_path))
    verdict = verdicts[0]
    overrides = {
        verdict.excel_row: {"status": "=1+1", "note": "legacy"}
    }
    out = tmp_path / "legacy.xlsx"
    write_annotated_xlsx(
        plog_path, str(out), verdicts, plog.header_row, plog.sheet, overrides
    )
    ws = load_workbook(out, data_only=False)[plog.sheet]
    assert ws.cell(verdict.excel_row, 19).data_type != "f"
    assert ws.cell(verdict.excel_row, 19).value == (verdict.column_s() or None)


def test_xlsx_contains_hidden_perimeter_provenance(
        plog_path, dmr_path, fake_resolver, tmp_path):
    plog = parse_plog(plog_path)
    verdicts = run_pipeline(plog, parse_dmr(dmr_path))
    out = tmp_path / "provenance.xlsx"
    write_annotated_xlsx(
        plog_path, str(out), verdicts, plog.header_row, plog.sheet,
        perimeter_meta={"hash": "abc", "warnings": ["stale"]},
        perimeter_warning="cache unavailable",
    )
    wb = load_workbook(out)
    meta = wb["_DMR_AUDIT_META"]
    assert meta.sheet_state == "hidden"
    values = {row[0].value: row[1].value for row in meta.iter_rows(min_row=2)}
    assert values["hash"] == "abc"
    assert values["warning"] == "cache unavailable"


def test_perimeter_provenance_never_deletes_same_named_user_sheet(
        plog_path, dmr_path, fake_resolver, tmp_path):
    source = tmp_path / "collision.xlsx"
    wb = load_workbook(plog_path)
    user_sheet = wb.create_sheet("_DMR_AUDIT_META")
    user_sheet["A1"] = "USER CONTENT"
    wb.save(source)
    plog = parse_plog(str(source))
    out = tmp_path / "collision-out.xlsx"
    write_annotated_xlsx(
        str(source), str(out), run_pipeline(plog, parse_dmr(dmr_path)),
        plog.header_row, plog.sheet, perimeter_meta={"hash": "abc"},
    )
    exported = load_workbook(out)
    assert exported["_DMR_AUDIT_META"]["A1"].value == "USER CONTENT"
    assert exported["_DMR_AUDIT_META_2"].sheet_state == "hidden"


def test_source_s_formula_is_preserved_and_audited(
        plog_path, dmr_path, fake_resolver, tmp_path):
    import json

    plog = parse_plog(plog_path)
    verdicts = run_pipeline(plog, parse_dmr(dmr_path))
    row = verdicts[0].excel_row
    source = _prefilled_copy(
        plog_path, tmp_path, s_edits={row: '=HYPERLINK("https://evil.invalid")'}
    )
    out = tmp_path / "source-s-safe.xlsx"
    write_annotated_xlsx(
        source, str(out), verdicts, plog.header_row, plog.sheet
    )

    wb = load_workbook(out, data_only=False)
    cell = wb[plog.sheet].cell(row=row, column=19)
    assert cell.data_type == "f"
    assert cell.value == '=HYPERLINK("https://evil.invalid")'
    provenance = [
        json.loads(audit_row[1].value)
        for audit_row in wb["_DMR_AUDIT_META"].iter_rows(min_row=2)
        if audit_row[0].value == "source_column_s"
    ]
    assert provenance == [{
        "cell_data_type": "f",
        "disposition": "preserved",
        "excel_row": row,
        "pipeline_column_s": verdicts[0].column_s(),
        "value": '=HYPERLINK("https://evil.invalid")',
    }]


def test_sparse_content_outside_verdict_rows_moves_evidence_block(
        plog_path, dmr_path, fake_resolver, tmp_path):
    plog = parse_plog(plog_path)
    verdicts = run_pipeline(plog, parse_dmr(dmr_path))
    source = _prefilled_copy(
        plog_path, tmp_path, col_edits={(10_000, 20): "keep sparse note"}
    )
    out = tmp_path / "sparse-evidence.xlsx"
    write_annotated_xlsx(
        source, str(out), verdicts, plog.header_row, plog.sheet
    )
    ws = load_workbook(out)[plog.sheet]
    assert ws.cell(10_000, 20).value == "keep sparse note"
    assert ws.cell(plog.header_row, 20).value in (None, "")
    assert ws.cell(plog.header_row, 21).value == "STATUS"


def test_blank_merged_cells_move_evidence_block(
        plog_path, dmr_path, fake_resolver, tmp_path):
    plog = parse_plog(plog_path)
    verdicts = run_pipeline(plog, parse_dmr(dmr_path))
    source = _prefilled_copy(plog_path, tmp_path)
    source_wb = load_workbook(source)
    verdict_row = verdicts[0].excel_row
    source_wb[plog.sheet].merge_cells(
        start_row=verdict_row, start_column=20,
        end_row=verdict_row, end_column=34,
    )
    source_wb.save(source)

    out = tmp_path / "merged-evidence.xlsx"
    write_annotated_xlsx(
        source, str(out), verdicts, plog.header_row, plog.sheet
    )

    ws = load_workbook(out)[plog.sheet]
    assert f"T{verdict_row}:AH{verdict_row}" in {
        str(cell_range) for cell_range in ws.merged_cells.ranges
    }
    assert ws.cell(plog.header_row, 20).value in (None, "")
    assert ws.cell(plog.header_row, 35).value == "STATUS"


def test_evidence_uses_hidden_sheet_when_no_bounded_column_block_fits(
        plog_path, dmr_path, fake_resolver, tmp_path, monkeypatch):
    import app.reconciler.export as export_mod

    plog = parse_plog(plog_path)
    verdicts = run_pipeline(plog, parse_dmr(dmr_path))
    source = _prefilled_copy(
        plog_path, tmp_path, col_edits={(verdicts[0].excel_row, 20): "keep"}
    )
    source_wb = load_workbook(source)
    source_wb.create_sheet("_DMR_EVIDENCE")["A1"] = "USER EVIDENCE"
    source_wb.save(source)
    # With a 15-column evidence block, T:AH is the only possible location.
    monkeypatch.setattr(export_mod, "EXCEL_MAX_COLUMN", 34)
    out = tmp_path / "fallback-evidence.xlsx"
    write_annotated_xlsx(
        source, str(out), verdicts, plog.header_row, plog.sheet
    )

    wb = load_workbook(out)
    assert wb[plog.sheet].cell(verdicts[0].excel_row, 20).value == "keep"
    assert wb["_DMR_EVIDENCE"]["A1"].value == "USER EVIDENCE"
    evidence = wb["_DMR_EVIDENCE_2"]
    assert evidence.sheet_state == "hidden"
    assert evidence.cell(1, 6).value == "STATUS"
    assert evidence.cell(2, 1).value == plog.sheet
    meta = {
        row[0].value: row[1].value
        for row in wb["_DMR_AUDIT_META"].iter_rows(min_row=2)
    }
    assert meta["evidence_sheet"] == "_DMR_EVIDENCE_2"
