"""Regression tests for adversarially-verified review findings."""
from __future__ import annotations

import json
import threading
from concurrent.futures import ThreadPoolExecutor
from datetime import date, datetime

import httpx
import pytest
from openpyxl import Workbook

from app.reconciler.name_match import name_contains, name_ladder
from app.core.xlsx import to_date as _to_date, to_int as _to_int
from app.reconciler.domain import REVIEW, Verdict
from app.reconciler.parsers import (DmrParse, DmrRow, PlogParse, PlogRow,
                                    parse_dmr, parse_plog)
from app.reconciler.links import (_extract_note_fields, _normalize_url,
                                  _note_id_from_url, _retry_after_seconds,
                                  Resolution)
from tests import fixtures


# ---------------------------------------------------------- parser findings

def _build_dmr_header_row1(path: str) -> None:
    """DMR export with the header on row 1 — no metadata rows at all."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Streaming"
    ws.append(fixtures.DMR_HEADERS)
    for r in fixtures.dmr_rows():
        ws.append([
            "CN", "Beauty", r["blogger"], r["user"], "RED", r["pid"],
            r["likes"], r["favs"], r["date"], 150000, "Fashion", "Dior",
            # a data cell that LOOKS like a window string must not be parsed
            "Makeup", "#tag", "Show Post", r["likes"],
            1.0, "Promo From 2020/01/01 To 2020/02/01", 0, "", r["comments"],
        ])
    wb.save(path)


def test_dmr_header_on_row1_does_not_scan_data_for_window(tmp_path):
    p = tmp_path / "dmr_row1.xlsx"
    _build_dmr_header_row1(str(p))
    d = parse_dmr(str(p))
    assert d.header_row == 1
    # openpyxl treats max_row=0 as unset — the guard must prevent the whole
    # sheet being treated as metadata and a bogus 2020 window being extracted
    assert d.window_from is None and d.window_to is None
    assert d.metadata_text == ""
    assert any("window" in w for w in d.warnings)
    assert len(d.rows) == len(fixtures.dmr_rows())


def test_blank_campaign_cells_forward_fill(tmp_path):
    """Section-style sheets put the CAMPAIGN value only on the first row of
    each section; later rows inherit it."""
    wb = Workbook()
    ws = wb.active
    ws.title = "MASTER KOL LIST"
    ws.append(fixtures.PLOG_HEADERS)
    rows = [
        (1, "PLOG #009", "甲", datetime(2026, 3, 1)),
        (2, "", "乙", datetime(2026, 3, 2)),          # inherits #009
        (1, "PLOG #010", "丙", datetime(2026, 3, 3)),
        (2, "", "丁", datetime(2026, 3, 4)),          # inherits #010
    ]
    for no, camp, name, dt in rows:
        ws.append([no, "", camp, "", "", name, 1, dt, "", f"http://xhslink.com/o/{name}",
                   1, 1, 1, 1, 3, 1, 1, 1])
    p = tmp_path / "plog_sections.xlsx"
    wb.save(str(p))
    parsed = parse_plog(str(p))
    assert [r.campaign for r in parsed.rows] == \
        ["PLOG #009", "PLOG #009", "PLOG #010", "PLOG #010"]
    assert [r.key for r in parsed.rows] == \
        [("PLOG #009", "1"), ("PLOG #009", "2"),
         ("PLOG #010", "1"), ("PLOG #010", "2")]


def test_plog_hyperlink_post_link_cell(tmp_path):
    """POST LINK stored as a display-text hyperlink cell must yield the
    hyperlink target, not the display text."""
    wb = Workbook()
    ws = wb.active
    ws.title = "MASTER KOL LIST"
    ws.append(fixtures.PLOG_HEADERS)
    ws.append([1, "", "C1", "", "", "某博主", 1, datetime(2026, 4, 1), "",
               "链接", 1, 1, 1, 1, 3, 1, 1, 1])
    ws.cell(row=2, column=10).hyperlink = "http://xhslink.com/o/realtarget"
    p = tmp_path / "plog_link.xlsx"
    wb.save(str(p))
    parsed = parse_plog(str(p))
    assert parsed.rows[0].post_link == "http://xhslink.com/o/realtarget"


def test_excel_serial_date_and_comma_numbers():
    assert _to_date(45838.0) is not None          # serial float
    assert _to_date("2026年5月13日") is None       # unparseable → warning path
    assert _to_int("1,234") == 1234
    assert _to_int("１，２３４") == 1234           # full-width digits + comma
    assert _to_int("n/a") is None


def test_non_hex_postid_warns(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Streaming"
    ws.append(["User: x, Top Bloggers - From 2026/01/01 To 2026/07/20"])
    ws.append([""])
    ws.append(fixtures.DMR_HEADERS)
    ws.append(["CN", "B", "某人", "5f00000000000000000000ff", "RED", "N/A",
               1, 1, datetime(2026, 6, 1), 1, "", "", "", "", "", 1, 1, "", 0, "", 1])
    p = tmp_path / "dmr_badpid.xlsx"
    wb.save(str(p))
    d = parse_dmr(str(p))
    assert len(d.rows) == 1  # row kept — the deviation is itself a finding
    assert any("not a 24-char hex" in w for w in d.warnings)


def test_missing_username_column_warns(tmp_path):
    headers = [h for h in fixtures.DMR_HEADERS if h != "Username"]
    wb = Workbook()
    ws = wb.active
    ws.title = "Streaming"
    ws.append(["Top Bloggers - From 2026/01/01 To 2026/07/20"])
    ws.append([""])
    ws.append(headers)
    ws.append(["CN", "B", "某人", "RED", fixtures.N_EXTRA,
               1, 1, datetime(2026, 6, 1), 1, "", "", "", "", "", 1, 1, "", 0, ""])
    p = tmp_path / "dmr_nouser.xlsx"
    wb.save(str(p))
    d = parse_dmr(str(p))
    assert any("Username" in w for w in d.warnings)


def test_unparseable_post_date_warns(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "MASTER KOL LIST"
    ws.append(fixtures.PLOG_HEADERS)
    ws.append([1, "", "C1", "", "", "某博主", 1, "2026年5月13日", "",
               "http://xhslink.com/o/x", 1, 1, 1, 1, 3, 1, 1, 1])
    p = tmp_path / "plog_baddate.xlsx"
    wb.save(str(p))
    parsed = parse_plog(str(p))
    assert parsed.rows[0].post_date is None
    assert any("POST DATE" in w for w in parsed.warnings)


def test_parsers_do_not_drop_data_after_200_blank_rows(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "MASTER KOL LIST"
    ws.append(fixtures.PLOG_HEADERS)
    ws.append([1, "", "C", "", "", "before", 1, datetime(2026, 1, 1), "",
               "https://xhslink.com/o/a", 1, 1, 1, 1, 3, 1, 1, 1])
    ws.cell(row=250, column=1, value=2)
    ws.cell(row=250, column=3, value="C")
    ws.cell(row=250, column=6, value="after")
    ws.cell(row=250, column=10, value="https://xhslink.com/o/b")
    path = tmp_path / "spaced.xlsx"
    wb.save(path)
    assert [row.name for row in parse_plog(str(path)).rows] == ["before", "after"]


def test_sparse_final_row_does_not_force_million_row_scan(tmp_path):
    from openpyxl.styles import PatternFill

    wb = Workbook()
    ws = wb.active
    ws.append(fixtures.PLOG_HEADERS)
    ws.append([1, "", "C", "", "", "real", 1, None, "",
               "https://xhslink.com/o/a", 1, 1, 1, 1, 3, 1, 1, 1])
    ws.cell(row=1_048_576, column=18).fill = PatternFill(
        fill_type="solid", fgColor="FFFFFF"
    )
    path = tmp_path / "sparse-tail.xlsx"
    wb.save(path)
    parsed = parse_plog(str(path))
    assert len(parsed.rows) == 1 and parsed.rows[0].name == "real"


def test_plog_logical_row_limit_is_enforced(tmp_path, monkeypatch):
    from app import config

    monkeypatch.setattr(config, "MAX_PLOG_ROWS", 1)
    wb = Workbook()
    ws = wb.active
    ws.append(fixtures.PLOG_HEADERS)
    for number in (1, 2):
        ws.append([number, "", "C", "", "", f"n{number}", 1,
                   datetime(2026, 1, number), "",
                   f"https://xhslink.com/o/{number}", 1, 1, 1, 1, 3, 1, 1, 1])
    path = tmp_path / "too_many.xlsx"
    wb.save(path)
    import pytest
    with pytest.raises(ValueError, match="more than 1 data rows"):
        parse_plog(str(path))


# --------------------------------------------------------- matcher findings

def test_ladder_rejects_short_dmr_ascii():
    # a 1-char DMR ASCII remainder must not fuzzy-match anything
    assert name_ladder("Poppy-chan", "C酱") == ""
    assert name_ladder("饼饼", "Bin") == ""  # pinyin bridge needs >=4 on DMR side


def test_all_emoji_name_never_flags_mislabel():
    assert name_contains("🥚🥚", "anything at all")
    assert name_contains("", "gungunnnnn")


def _dmr_row(*, post_id: str, username: str, blogger: str = "blogger",
             excel_row: int = 2, post_date: datetime | None = None) -> DmrRow:
    return DmrRow(
        blogger=blogger, username=username, post_id=post_id,
        post_id_raw=post_id, post_date=post_date or datetime(2026, 1, 2),
        likes_retweet=1, share_favorites=1, engagement=2, comments=0,
        link_target="", link_embedded_post_id="", excel_row=excel_row,
    )


def _plog_row(*, name: str = "blogger", link: str = "https://xhslink.com/o/x",
              campaign: str = "C", excel_row: int = 2) -> PlogRow:
    return PlogRow(
        campaign=campaign, no=str(excel_row), name=name,
        post_date=date(2026, 1, 2), post_link=link, like=1, collection=1,
        comment=0, impression=1, ttl_engagement=2, excel_row=excel_row,
    )


def test_author_ids_are_case_insensitive_and_canonical():
    from app.reconciler.pipeline import build_indexes, match_row

    author = "5F00000000000000000000A1"
    dmr = DmrParse("S", 1, {}, rows=[
        _dmr_row(post_id="6a0000000000000000000001", username=author.lower())
    ])
    verdict = match_row(
        _plog_row(), build_indexes(dmr),
        Resolution(status="ok", note_id="6A0000000000000000000002",
                   author_id=author),
        (date(2026, 1, 1), date(2026, 1, 31)),
    )
    assert verdict.status == "NO_POST"
    assert verdict.resolved_author_id == author.lower()


def test_partial_username_index_never_proves_no_blogger():
    from app.reconciler.pipeline import build_indexes, match_row

    dmr = DmrParse("S", 1, {}, rows=[
        _dmr_row(post_id="6a0000000000000000000001",
                 username="5f0000000000000000000001"),
        _dmr_row(post_id="6a0000000000000000000002", username="",
                 excel_row=3),
    ])
    verdict = match_row(
        _plog_row(name="unknown"), build_indexes(dmr),
        Resolution(status="ok", note_id="6a0000000000000000000003",
                   author_id="5f0000000000000000000099"),
        (date(2026, 1, 1), date(2026, 1, 31)),
    )
    assert verdict.status == REVIEW
    assert verdict.tier == "2:partial-username-column"


def test_duplicate_post_id_is_manual_review():
    from app.reconciler.pipeline import build_indexes, match_row

    note = "6a0000000000000000000001"
    dmr = DmrParse("S", 1, {}, rows=[
        _dmr_row(post_id=note, username="5f0000000000000000000001"),
        _dmr_row(post_id=note, username="5f0000000000000000000002",
                 blogger="other", excel_row=3),
    ])
    verdict = match_row(
        _plog_row(), build_indexes(dmr),
        Resolution(status="ok", note_id=note),
        (date(2026, 1, 1), date(2026, 1, 31)),
    )
    assert verdict.status == REVIEW
    assert verdict.tier == "1:duplicate-post-id"
    assert len(verdict.candidates) == 2


def test_run_pipeline_resolves_each_normalized_url_once(monkeypatch):
    import app.reconciler.pipeline as pipeline

    note = "6a0000000000000000000001"
    plog = PlogParse("P", 1, {}, rows=[
        _plog_row(link="xhslink.com/o/shared", excel_row=2),
        _plog_row(link="http://xhslink.com/o/shared", excel_row=3),
    ])
    dmr = DmrParse("D", 1, {}, rows=[
        _dmr_row(post_id=note, username="5f0000000000000000000001")
    ])
    calls = 0

    def fake_resolve(*_args, **_kwargs):
        nonlocal calls
        calls += 1
        return Resolution(status="ok", note_id=note)

    monkeypatch.setattr(pipeline, "resolve_link", fake_resolve)
    verdicts = pipeline.run_pipeline(plog, dmr)
    assert calls == 1
    assert [v.status for v in verdicts] == ["MATCH", "MATCH"]


def test_name_scan_budget_prevents_cartesian_work(monkeypatch):
    import app.reconciler.pipeline as pipeline

    plog = PlogParse("P", 1, {}, rows=[_plog_row(name="unique")])
    dmr = DmrParse("D", 1, {}, rows=[
        _dmr_row(post_id=f"{i + 1:024x}", username=f"{i + 10:024x}",
                 blogger=f"other-{i}", excel_row=i + 2)
        for i in range(3)
    ])
    monkeypatch.setattr(pipeline, "MAX_NAME_SCAN_COMPARISONS", 2)
    monkeypatch.setattr(
        pipeline, "resolve_link",
        lambda *_args, **_kwargs: Resolution(status="failed", error="dead"),
    )
    monkeypatch.setattr(
        pipeline, "scan_by_name",
        lambda *_args, **_kwargs: (_ for _ in ()).throw(
            AssertionError("full DMR scan should be skipped")
        ),
    )
    (verdict,) = pipeline.run_pipeline(plog, dmr)
    assert verdict.status == "LINK_ERROR"
    assert any("safety budget" in note for note in verdict.notes)


def test_candidate_lists_are_bounded(monkeypatch):
    from app import config
    from app.reconciler.pipeline import rank_candidates

    monkeypatch.setattr(config, "MAX_CANDIDATES_PER_VERDICT", 2)
    rows = [
        (_dmr_row(post_id=f"{i:024x}", username=f"{i + 100:024x}",
                  excel_row=i + 2), "same-name")
        for i in range(10)
    ]
    assert len(rank_candidates(_plog_row(), rows, keep_out_of_window=True)) == 2


def test_reverse_audit_uses_campaign_windows_and_requires_date_and_post_id():
    from app.reconciler.reverse_audit import reverse_audit

    author = "5f0000000000000000000001"
    january = _plog_row(campaign="JAN", excel_row=2)
    december = _plog_row(campaign="DEC", excel_row=3)
    december.post_date = date(2026, 12, 1)
    plog = PlogParse("P", 1, {}, rows=[january, december],
                     campaigns=["JAN", "DEC"])
    verdicts = [
        Verdict("JAN", "1", "n", "2026-01-02", "", 2,
                resolved_author_id=author),
        Verdict("DEC", "1", "n", "2026-12-01", "", 3,
                resolved_author_id=author),
    ]
    dmr = DmrParse("D", 1, {}, rows=[
        _dmr_row(post_id="6a0000000000000000000001", username=author,
                 post_date=datetime(2026, 1, 2)),
        _dmr_row(post_id="6a0000000000000000000002", username=author,
                 post_date=datetime(2026, 6, 1), excel_row=3),
        _dmr_row(post_id="", username=author,
                 post_date=datetime(2026, 1, 2), excel_row=4),
        _dmr_row(post_id="6a0000000000000000000003", username=author,
                 post_date=None, excel_row=5),
    ])
    # Dataclass helper supplies a default date; explicitly clear the final row.
    dmr.rows[-1].post_date = None
    rows = reverse_audit(plog, dmr, verdicts)
    assert [(row["post_id"], row["campaigns"]) for row in rows] == [
        ("6a0000000000000000000001", ["JAN"])
    ]

    # A resolved PLOG note still represents the post when duplicate DMR rows
    # force REVIEW rather than MATCH; it must not reappear as an "extra".
    verdicts[0].resolved_note_id = "6a0000000000000000000001"
    assert reverse_audit(plog, dmr, verdicts) == []


# -------------------------------------------------------- resolver findings

def test_note_id_only_from_note_urls():
    uid = "5f00000000000000000000b1"
    nid = "6a1a0000000000000000a001"
    assert _note_id_from_url(f"https://www.xiaohongshu.com/user/profile/{uid}") is None
    assert _note_id_from_url(f"https://www.xiaohongshu.com/discovery/item/{nid}") == nid
    assert _note_id_from_url(f"https://www.xiaohongshu.com/explore/{nid}?xsec=1") == nid
    assert _note_id_from_url(
        f"https://evil.example/?next=xiaohongshu.com/explore/{nid}"
    ) is None
    assert _note_id_from_url(
        f"https://xiaohongshu.com.evil.example/explore/{nid}"
    ) is None


def test_normalize_url_schemeless():
    assert _normalize_url("xhslink.com/o/abc") == "https://xhslink.com/o/abc"
    assert _normalize_url("http://xhslink.com/o/abc") == "https://xhslink.com/o/abc"
    assert _normalize_url("随便写的字") == "随便写的字"  # stays; rejected later


def test_extract_note_fields_no_blind_hex_fallback():
    """An author id in the payload must not be mistaken for the note id."""
    payload = {
        "code": 200,
        "data": {
            "user": {
                "id": "5f00000000000000000000b1",
                "user_id": "5f00000000000000000000b1",
                "nickname": "某人",
            },
            "desc": "no note id here",
        },
    }
    fields = _extract_note_fields(payload)
    assert fields["note_id"] == ""


def test_extract_note_fields_normal_payload():
    payload = {"code": 200, "data": {"note_id": "6A1A0000000000000000A001",
                                     "title": "标题",
                                     "user": {"user_id": "5f00000000000000000000b1",
                                              "nickname": "某人"},
                                     "interact_info": {"liked_count": "607",
                                                       "collected_count": "44",
                                                       "comment_count": "19"}}}
    fields = _extract_note_fields(payload)
    assert fields["note_id"] == "6a1a0000000000000000a001"
    assert fields["author_id"] == "5f00000000000000000000b1"
    assert fields["likes"] == 607 and fields["collects"] == 44


def test_extract_note_fields_keeps_author_in_the_note_subtree():
    note = "6a1a0000000000000000a001"
    correct_author = "5f00000000000000000000b1"
    payload = {
        "user": {
            "user_id": "5f00000000000000000000ff",
            "nickname": "top-level metadata user",
        },
        "data": {
            "recommendation": {
                "user": {
                    "user_id": "5f00000000000000000000ee",
                    "nickname": "unrelated recommendation",
                },
            },
            "note": {
                "note_id": note,
                "user": {
                    "user_id": correct_author,
                    "nickname": "actual note author",
                },
                "interact_info": {"liked_count": "42"},
            },
        },
    }

    fields = _extract_note_fields(payload)

    assert fields["note_id"] == note
    assert fields["author_id"] == correct_author
    assert fields["author_name"] == "actual note author"
    assert fields["likes"] == 42


def test_extract_note_fields_rejects_invalid_author_identity():
    payload = {
        "data": {
            "note_id": "6a1a0000000000000000a001",
            "user": {"user_id": "not-a-platform-id", "nickname": "unknown"},
        },
    }

    assert _extract_note_fields(payload)["author_id"] == ""


def test_retry_after_http_date_does_not_crash():
    resp = httpx.Response(429, headers={"retry-after": "Wed, 22 Jul 2026 07:28:00 GMT"})
    assert _retry_after_seconds(resp, 0) >= 0
    resp = httpx.Response(429, headers={"retry-after": "7"})
    assert _retry_after_seconds(resp, 0) == 7.0
    resp = httpx.Response(429)
    assert _retry_after_seconds(resp, 1) == 4.0


@pytest.mark.parametrize("raw", ["-1", "NaN", "Infinity", "not-a-date"])
def test_invalid_retry_after_falls_back_to_finite_nonnegative_backoff(raw):
    resp = httpx.Response(429, headers={"retry-after": raw})
    assert _retry_after_seconds(resp, 1) == 4.0


def test_direct_resolve_survives_malformed_location(monkeypatch):
    """A garbage Location header must degrade, never raise (a raise here would
    abort the whole run through pool.map)."""
    import app.reconciler.links as resolver_mod

    def handler(request: httpx.Request) -> httpx.Response:
        return httpx.Response(302, headers={"location": "http://[::1/broken"})

    real_client = httpx.Client

    def patched_client(**kwargs):
        kwargs["transport"] = httpx.MockTransport(handler)
        return real_client(**kwargs)

    monkeypatch.setattr(resolver_mod.httpx, "Client", patched_client)
    assert resolver_mod.direct_resolve("http://xhslink.com/o/x") is None


def test_direct_resolve_rejects_redirect_to_private_host(monkeypatch):
    """Every redirect hop is revalidated before any network request."""
    import app.reconciler.links as resolver_mod

    seen = []

    def handler(request: httpx.Request) -> httpx.Response:
        seen.append(str(request.url))
        return httpx.Response(302, headers={"location": "http://127.0.0.1/admin"})

    real_client = httpx.Client

    def patched_client(**kwargs):
        kwargs["transport"] = httpx.MockTransport(handler)
        return real_client(**kwargs)

    monkeypatch.setattr(resolver_mod.httpx, "Client", patched_client)
    assert resolver_mod.direct_resolve("http://xhslink.com/o/x") is None
    assert seen == ["https://xhslink.com/o/x"]


def test_direct_fetch_rejects_redirect_to_private_host(monkeypatch):
    import app.reconciler.links as resolver_mod

    seen = []

    def handler(request: httpx.Request) -> httpx.Response:
        seen.append(str(request.url))
        return httpx.Response(302, headers={"location": "http://[::1]/secret"})

    real_client = httpx.Client

    def patched_client(**kwargs):
        kwargs["transport"] = httpx.MockTransport(handler)
        return real_client(**kwargs)

    monkeypatch.setattr(resolver_mod.httpx, "Client", patched_client)
    assert resolver_mod.direct_fetch_note_detail("https://xhslink.com/o/x") == {}
    assert seen == ["https://xhslink.com/o/x"]


def test_direct_fetch_rejects_embedded_different_note(monkeypatch):
    import app.reconciler.links as resolver_mod

    expected = "6a0000000000000000000001"
    wrong = "6a0000000000000000000002"
    state = {
        "noteData": {"data": {"noteData": {
            "noteId": wrong,
            "user": {"userId": "5f0000000000000000000001"},
            "interactInfo": {"likedCount": 1},
        }}}
    }

    def handler(request: httpx.Request) -> httpx.Response:
        html = f"<script>window.__INITIAL_STATE__ = {json.dumps(state)}</script>"
        return httpx.Response(200, text=html, request=request)

    real_client = httpx.Client

    def patched_client(**kwargs):
        kwargs["transport"] = httpx.MockTransport(handler)
        return real_client(**kwargs)

    monkeypatch.setattr(resolver_mod.httpx, "Client", patched_client)
    url = f"https://www.xiaohongshu.com/explore/{expected}"
    assert resolver_mod.direct_fetch_note_detail(url, expected) == {}


def test_resolve_link_singleflight_across_concurrent_runs(tmp_path, monkeypatch):
    from app import config
    from app.reconciler import links as resolver_mod

    monkeypatch.setattr(config, "DB_PATH", tmp_path / "singleflight.sqlite3")
    note = "6a0000000000000000000001"
    entered = threading.Event()
    release = threading.Event()
    calls = 0

    def fake_direct(_url):
        nonlocal calls
        calls += 1
        entered.set()
        assert release.wait(2)
        return note

    monkeypatch.setattr(resolver_mod, "direct_resolve", fake_direct)
    with ThreadPoolExecutor(max_workers=2) as executor:
        first = executor.submit(resolver_mod.resolve_link,
                                "http://xhslink.com/o/shared")
        assert entered.wait(1)
        second = executor.submit(resolver_mod.resolve_link,
                                 "xhslink.com/o/shared")
        release.set()
        results = [first.result(2), second.result(2)]
    assert calls == 1
    assert all(result.note_id == note for result in results)
    assert not resolver_mod._url_flights


def test_ensure_author_singleflight_rechecks_enriched_cache(tmp_path,
                                                             monkeypatch):
    from app import config
    from app.core import db
    from app.reconciler import links as resolver_mod

    monkeypatch.setattr(config, "DB_PATH", tmp_path / "author-flight.sqlite3")
    url = "https://xhslink.com/o/shared"
    note = "6a0000000000000000000001"
    author = "5f0000000000000000000001"
    db.cache_put(url, status="ok", note_id=note, source="direct")
    monkeypatch.setattr(resolver_mod, "direct_fetch_note_detail",
                        lambda *_args, **_kwargs: {})
    entered = threading.Event()
    release = threading.Event()
    calls = 0

    def fake_tikhub(*_args, **_kwargs):
        nonlocal calls
        calls += 1
        entered.set()
        assert release.wait(2)
        return {"data": {"note_id": note,
                         "user": {"user_id": author}}}

    monkeypatch.setattr(resolver_mod, "tikhub_fetch_note", fake_tikhub)
    with ThreadPoolExecutor(max_workers=2) as executor:
        first = executor.submit(
            resolver_mod.ensure_author, url,
            Resolution(status="ok", note_id=note, source="direct"),
        )
        assert entered.wait(1)
        second = executor.submit(
            resolver_mod.ensure_author, url,
            Resolution(status="ok", note_id=note, source="direct"),
        )
        release.set()
        results = [first.result(2), second.result(2)]
    assert calls == 1
    assert all(result.author_id == author for result in results)
    assert not resolver_mod._url_flights


# ------------------------------------------------------ adjudicator findings

def test_parse_batch_accepts_bare_array_and_fences():
    from app.reconciler.adjudicator import _parse_batch
    item = ('{"row": "C|1|r2", "verdict": "UNSURE", "confidence": 0.5, '
            '"rationale_en": "e", "rationale_zh": "z"}')
    assert _parse_batch(f'{{"items": [{item}]}}') is not None
    assert _parse_batch(f"[{item}]") is not None          # bare array
    assert _parse_batch(f"```json\n[{item}]\n```") is not None
    assert _parse_batch("nonsense") is None


def test_adjudicator_logs_provider_detail_without_exposing_it(
        monkeypatch, caplog):
    import logging
    from app.reconciler import adjudicator

    verdict = Verdict(
        "C", "1", "name", None, "https://xhslink.com/o/x", 2
    )

    def fail(*_args, **_kwargs):
        raise RuntimeError("secret-provider-detail")

    monkeypatch.setattr(adjudicator.llm, "complete", fail)
    with caplog.at_level(logging.INFO):
        adjudicator._adjudicate_chunk(
            object(), [(verdict, {"row": "C|1|r2"})], None
        )

    assert verdict.notes == [
        "LLM adjudication unavailable: provider request failed — kept for "
        "human review."
    ]
    assert "secret-provider-detail" not in verdict.notes[0]
    assert "secret-provider-detail" in caplog.text


# ------------------------------------------------------------ eval findings

def test_eval_classify_review_with_link_mention():
    from tools import evaluate as eval_mod
    assert eval_mod.classify("人工复核（链接已解析但无法获取作者ID）") == "人工复核"
    assert eval_mod.classify("Check链接错误（同名候选: x 6a…）") == "Check链接错误"
    assert eval_mod.classify("有 但是DMR博主名字标注错误") == "有 但是DMR博主名字标注错误"
    assert eval_mod.classify("") == "MATCH"
    assert eval_mod.classify("无帖子（超出DMR导出窗口，预期缺失）") == "无帖子"
