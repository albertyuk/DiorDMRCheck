"""Perimeter cross-check: parsing, caching, and the NO_BLOGGER split.

Uses its own small fixture workbooks (separate from the shared fixtures so
existing campaign-shape assertions stay untouched)."""
from __future__ import annotations

import io
from datetime import datetime

import pytest
from openpyxl import Workbook

from app.reconciler.pipeline import (LINK_ERROR, NO_BLOGGER, NO_BLOGGER_NOT_IN_PERIMETER,
                         NO_POST_IN_PERIMETER, run_pipeline)
from app.reconciler.parsers import parse_dmr, parse_plog
from app.reconciler.perimeter import (PerimeterIndex, file_hash, load_cached,
                           parse_perimeter, store_parsed)
from app.reconciler.links import Resolution
from tests import fixtures

PERIM_HEADERS = [
    "NAME", "NAMEBIS", "DMRID", "COUNTRY", "TYPE", "INSTAGRAM_ID",
    "INSTAGRAM_FOLLOWERS", "YOUTUBE_ID", "YOUTUBE_FOLLOWERS", "TIKTOK_ID",
    "TIKTOK_FOLLOWERS", "WEIBO_ID", "WEIBO_FOLLOWERS", "REDBOOK_ID",
    "REDBOOK_FOLLOWERS",
]

U_IN = "5f000000000000000000c001"      # in-perimeter author
U_NEAR = "5f000000000000000000c002"    # resolved author for the near-miss row
U_NEAR_PERIM = "5f000000000000000000c003"  # perimeter's different account
U_MULTI = "5f000000000000000000c004"   # esther-style multi-hit author
U_NOREG = "5f000000000000000000c005"   # name in perimeter, no REDBOOK_ID
U_NOWHERE = "5f000000000000000000c006" # not in perimeter at all


def _perimeter_rows():
    return [
        ("Yi Ke Ji Dan", "一颗鸡蛋", "100001", U_IN, 52000),
        ("Xiao Ce Shi", "小小测试", "100002", U_NEAR_PERIM, 9000),
        ("esther", "", "100003", "5f000000000000000000c0aa", 1000),
        ("esther lee", "", "100004", "", None),
        ("Esther W", "艾斯特", "100005", "5f000000000000000000c0bb", 2000),
        ("Wu Dang An", "无档案", "100006", "", None),
        ("Unrelated Person", "路人甲", "100007", "5f000000000000000000c0cc", 500),
    ]


def build_perimeter_bytes(extraction="19/05/2026 10:30:00") -> bytes:
    wb = Workbook()
    macro = wb.active
    macro.title = "List Macro"           # must be ignored
    macro.append(["NAME", "REDBOOK_ID"])
    macro.append(["Macro Person", "5f000000000000000000ffff"])
    ws = wb.create_sheet("List Micro")
    ws.append([])
    ws.append(["2026 MICRO social perimeter"])
    ws.append(["LVMH internal use only"])
    ws.append(["The perimeter is regularly updated"])
    ws.append([f"Date of extraction : {extraction}"])
    ws.append([])
    ws.append(PERIM_HEADERS)
    for name, namebis, dmrid, rid, followers in _perimeter_rows():
        ws.append([name, namebis, dmrid, "China", "Influencer",
                   "", None, "", None, "", None, "", None, rid, followers])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.fixture
def perim_index(tmp_path, monkeypatch) -> PerimeterIndex:
    from app import config
    monkeypatch.setattr(config, "DB_PATH", tmp_path / "perim.sqlite3")
    data = build_perimeter_bytes()
    h = file_hash(data)
    parsed = parse_perimeter(io.BytesIO(data), filename="perim.xlsx",
                             content_hash=h)
    store_parsed(parsed)
    idx = load_cached(h)     # exercise the cache round-trip
    assert idx is not None
    return idx


def test_parse_perimeter_contract(perim_index):
    assert perim_index.extraction_date == "19/05/2026 10:30:00"
    assert len(perim_index.rows) == len(_perimeter_rows())
    # Macro sheet ignored
    assert "5f000000000000000000ffff" not in perim_index.by_redbook
    assert U_IN in perim_index.by_redbook
    row = perim_index.by_redbook[U_IN]
    assert row["namebis"] == "一颗鸡蛋" and row["redbook_followers"] == 52000


def test_scan_name_multi_hit(perim_index):
    hits = perim_index.scan_name("esther")
    assert len(hits) >= 2  # esther / esther lee / Esther W collide


# --------------------------------------------------------- pipeline split

def _mini_plog(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "MASTER KOL LIST"
    ws.append(fixtures.PLOG_HEADERS)
    rows = [
        (1, "一颗鸡蛋🥚", "in-perim"),
        (2, "小小测试", "near-miss"),
        (3, "esther", "multi"),
        (4, "无档案", "no-reg"),
        (5, "查无此人", "nowhere"),
        (6, "断链博主", "dead-link"),
    ]
    for no, name, tag in rows:
        ws.append([no, "", "C#1", "", "", name, 1, datetime(2026, 6, 1), "",
                   f"http://xhslink.com/o/{tag}", 1, 1, 1, 1, 3, 1, 1, 1])
    wb.save(path)


def _mini_dmr(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Streaming"
    ws.append(["Top Bloggers - From 2026/01/01 To 2026/07/20"])
    ws.append([])
    ws.append(fixtures.DMR_HEADERS)
    # one unrelated row so indexes are non-empty (none of the test authors)
    ws.append(["CN", "B", "别人", "5f00000000000000000000d1",
               "RED", "6a1a0000000000000000d001", 1, 1,
               datetime(2026, 6, 2), 1, "", "", "", "", "", 1, 1, "", 0, "", 1])
    wb.save(path)


RESOLUTIONS = {
    "http://xhslink.com/o/in-perim": (U_IN, "6a1a0000000000000000e001"),
    "http://xhslink.com/o/near-miss": (U_NEAR, "6a1a0000000000000000e002"),
    "http://xhslink.com/o/multi": (U_MULTI, "6a1a0000000000000000e003"),
    "http://xhslink.com/o/no-reg": (U_NOREG, "6a1a0000000000000000e004"),
    "http://xhslink.com/o/nowhere": (U_NOWHERE, "6a1a0000000000000000e005"),
}


@pytest.fixture
def split_verdicts(tmp_path, monkeypatch, perim_index):
    _mini_plog(tmp_path / "p.xlsx")
    _mini_dmr(tmp_path / "d.xlsx")

    def fake_resolve(url, run_counter=None, retry_failed=False):
        if url in RESOLUTIONS:
            author, note = RESOLUTIONS[url]
            return Resolution(status="ok", note_id=note, author_id=author,
                              source="fixture")
        return Resolution(status="failed", error="dead link")

    import app.reconciler.pipeline as m
    monkeypatch.setattr(m, "resolve_link", fake_resolve)
    monkeypatch.setattr(m, "ensure_author",
                        lambda url, res, run_counter=None, retry_failed=False: res)
    plog = parse_plog(str(tmp_path / "p.xlsx"))
    dmr = parse_dmr(str(tmp_path / "d.xlsx"))
    vs = run_pipeline(plog, dmr, perimeter=perim_index)
    return {v.name: v for v in vs}


def test_primary_id_join_flips_to_in_perimeter(split_verdicts):
    v = split_verdicts["一颗鸡蛋🥚"]
    assert v.status == NO_POST_IN_PERIMETER
    assert v.perimeter_method == "redbook-id"
    assert v.perimeter_redbook_id == U_IN
    assert v.column_s() == "无博主但在Perimeter内→无帖子"


def test_near_miss_same_name_different_id(split_verdicts):
    v = split_verdicts["小小测试"]
    assert v.status == NO_BLOGGER_NOT_IN_PERIMETER
    assert "REDBOOK_ID不同" in v.perimeter_note
    assert v.column_s() == "无博主（不在Perimeter内）"


def test_multi_hit_never_classifies_by_name(split_verdicts):
    v = split_verdicts["esther"]
    assert v.status == NO_BLOGGER_NOT_IN_PERIMETER
    assert len(v.perimeter_candidates) >= 2
    assert v.perimeter_method == ""  # nothing auto-picked


def test_single_hit_without_redbook_id(split_verdicts):
    v = split_verdicts["无档案"]
    assert v.status == NO_BLOGGER_NOT_IN_PERIMETER
    assert "未登记REDBOOK_ID" in v.perimeter_note


def test_not_in_perimeter_plain(split_verdicts):
    v = split_verdicts["查无此人"]
    assert v.status == NO_BLOGGER_NOT_IN_PERIMETER
    assert not v.perimeter_method and not v.perimeter_candidates


def test_dead_link_verdict_never_changes(split_verdicts):
    v = split_verdicts["断链博主"]
    assert v.status == LINK_ERROR
    assert v.column_s().startswith("Check链接错误")


def test_without_perimeter_behavior_unchanged(tmp_path, monkeypatch):
    _mini_plog(tmp_path / "p.xlsx")
    _mini_dmr(tmp_path / "d.xlsx")

    def fake_resolve(url, run_counter=None, retry_failed=False):
        if url in RESOLUTIONS:
            author, note = RESOLUTIONS[url]
            return Resolution(status="ok", note_id=note, author_id=author,
                              source="fixture")
        return Resolution(status="failed", error="dead link")

    import app.reconciler.pipeline as m
    monkeypatch.setattr(m, "resolve_link", fake_resolve)
    monkeypatch.setattr(m, "ensure_author",
                        lambda url, res, run_counter=None, retry_failed=False: res)
    plog = parse_plog(str(tmp_path / "p.xlsx"))
    dmr = parse_dmr(str(tmp_path / "d.xlsx"))
    vs = run_pipeline(plog, dmr, perimeter=None)
    by = {v.name: v for v in vs}
    assert by["一颗鸡蛋🥚"].status == NO_BLOGGER
    assert by["一颗鸡蛋🥚"].column_s() == "无博主"


def test_eval_classify_maps_new_statuses_to_no_blogger():
    from tools import evaluate as ev
    assert ev.classify("无博主但在Perimeter内→无帖子") == "无博主"
    assert ev.classify("无博主（不在Perimeter内）") == "无博主"


# ------------------------------------------- promotion & cached warnings

def test_parse_and_cache_does_not_promote(tmp_path, monkeypatch):
    """Uploading (preview) must not change the app-wide current perimeter;
    promotion is explicit. Cache hits replay the first parse's warnings."""
    from app import config
    from app.reconciler import perimeter as pm
    monkeypatch.setattr(config, "DATA_DIR", tmp_path)
    monkeypatch.setattr(config, "DB_PATH", tmp_path / "p.sqlite3")

    data = build_perimeter_bytes(extraction="")   # no extraction date → warning
    meta, warnings = pm.parse_and_cache(data, "perim.xlsx")
    assert meta["hash"] == pm.file_hash(data)
    assert any("Date of extraction" in w for w in warnings)
    assert pm.current_meta() is None              # NOT promoted

    meta2, warnings2 = pm.parse_and_cache(data, "perim.xlsx")
    assert meta2["rows"] == meta["rows"]
    assert warnings2 == warnings                  # cache hit keeps warnings

    pm.promote_cached(meta["hash"])
    cur = pm.current_meta()
    assert cur and cur["hash"] == meta["hash"]
    assert cur["rows"] == meta["rows"]


def test_promote_cached_unknown_hash_is_noop(tmp_path, monkeypatch):
    from app import config
    from app.reconciler import perimeter as pm
    monkeypatch.setattr(config, "DATA_DIR", tmp_path)
    monkeypatch.setattr(config, "DB_PATH", tmp_path / "p2.sqlite3")
    pm.promote_cached("deadbeef" * 8)
    assert pm.current_meta() is None


# ------------------------------------------------- China-market filter

def _mini_perimeter(headers, rows, sheet="List Micro") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet
    ws.append(["Date of extraction : 19/05/2026 10:30:00"])
    ws.append(headers)
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_country_filter_keeps_only_china():
    """The tool evaluates the Chinese market: Micro sheets (no IN_CHINA
    column) filter by COUNTRY."""
    data = _mini_perimeter(
        ["NAME", "COUNTRY", "REDBOOK_ID"],
        [["甲", "MAINLAND CHINA", "5f0000000000000000000001"],
         ["乙", "Mainland China", ""],          # case-insensitive
         ["丙", "China", "5f0000000000000000000002"],
         ["丁", "U.S.A.", "5f0000000000000000000003"],
         ["戊", "TAIWAN", ""],
         ["己", "Hong Kong", ""]])
    p = parse_perimeter(io.BytesIO(data))
    assert p.china_filter == "COUNTRY"
    assert p.rows_scanned == 6
    assert [r["name"] for r in p.rows] == ["甲", "乙", "丙"]
    assert p.redbook_count == 2                 # 丁's REDBOOK row dropped


def test_in_china_reports_wins_over_country():
    """Macro-style sheets carry an explicit flag — it overrides COUNTRY
    (an influencer abroad can still be in China-market reports)."""
    data = _mini_perimeter(
        ["NAME", "COUNTRY", "REDBOOK_ID", "IN_CHINA_REPORTS"],
        [["在库海外", "U.S.A.", "5f0000000000000000000001", "YES"],
         ["不在库本土", "MAINLAND CHINA", "5f0000000000000000000002", "NO"],
         ["在库本土", "MAINLAND CHINA", "5f0000000000000000000003", "yes"]])
    p = parse_perimeter(io.BytesIO(data))
    assert p.china_filter == "IN_CHINA_REPORTS"
    assert [r["name"] for r in p.rows] == ["在库海外", "在库本土"]


def test_missing_both_columns_keeps_all_and_warns():
    data = _mini_perimeter(
        ["NAME", "REDBOOK_ID"],
        [["甲", "5f0000000000000000000001"], ["乙", ""]])
    p = parse_perimeter(io.BytesIO(data))
    assert p.china_filter == ""
    assert len(p.rows) == 2
    assert any("cannot restrict" in w for w in p.warnings)


def test_all_rows_filtered_warns_loudly():
    data = _mini_perimeter(
        ["NAME", "COUNTRY", "REDBOOK_ID"],
        [["甲", "U.S.A.", ""], ["乙", "JAPAN", ""]])
    p = parse_perimeter(io.BytesIO(data))
    assert not p.rows and p.rows_scanned == 2
    warning = next(w for w in p.warnings if "filtered out" in w)
    # …and it translates
    from app.i18n import make_td
    zh = make_td("zh")(warning)
    assert "全部被过滤" in zh and "2" in zh


def test_file_hash_is_parser_version_salted():
    """Same bytes must NOT hit caches written under old parse semantics
    (they would serve unfiltered, non-China rows)."""
    import hashlib
    assert file_hash(b"same-bytes") != hashlib.sha256(b"same-bytes").hexdigest()


# --------------------------------------------------------- macro perimeter

U_MACRO = "5f000000000000000000c00a"    # macro-only member
U_FOREIGN = "5f000000000000000000c00b"  # macro row filtered by IN_CHINA=NO

MACRO_HEADERS = ["NAME", "NAMEBIS", "DMRID", "IN_CHINA_REPORTS",
                 "REDBOOK_ID", "REDBOOK_FOLLOWERS"]


def build_macro_bytes(extraction="21/05/2026 09:00:00") -> bytes:
    wb = Workbook()
    micro = wb.active
    micro.title = "List Micro"           # must be ignored for kind=macro
    micro.append(["NAME", "REDBOOK_ID"])
    micro.append(["Micro Person", "5f000000000000000000fffe"])
    ws = wb.create_sheet("List Macro")
    ws.append([f"Date of extraction : {extraction}"])
    ws.append(MACRO_HEADERS)
    ws.append(["Macro Star", "宏观大V", "M0001", "YES", U_MACRO, 800000])
    ws.append(["Yi Ke Ji Dan", "一颗鸡蛋", "M0002", "YES", U_IN, 52000])
    ws.append(["Foreign Macro", "海外大V", "M0003", "NO", U_FOREIGN, 100])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.fixture
def macro_index(tmp_path, monkeypatch) -> PerimeterIndex:
    from app import config
    monkeypatch.setattr(config, "DB_PATH", tmp_path / "perim_macro.sqlite3")
    data = build_macro_bytes()
    h = file_hash(data, "macro")
    parsed = parse_perimeter(io.BytesIO(data), filename="macro.xlsx",
                             content_hash=h, kind="macro")
    store_parsed(parsed)
    idx = load_cached(h)
    assert idx is not None
    return idx


def test_macro_parse_contract(macro_index):
    """kind="macro" reads List Macro, honors IN_CHINA_REPORTS, ignores the
    Micro sheet."""
    assert macro_index.extraction_date == "21/05/2026 09:00:00"
    assert U_MACRO in macro_index.by_redbook
    assert U_IN in macro_index.by_redbook
    assert U_FOREIGN not in macro_index.by_redbook       # IN_CHINA=NO dropped
    assert "5f000000000000000000fffe" not in macro_index.by_redbook


def test_macro_parse_error_names_macro_sheet(tmp_path):
    wb = Workbook()
    wb.active.title = "Only Micro Here"
    wb.active.append(["NAME", "REDBOOK_ID"])
    buf = io.BytesIO()
    wb.save(buf)
    with pytest.raises(ValueError, match="List Macro"):
        parse_perimeter(io.BytesIO(buf.getvalue()), kind="macro")


def test_file_hash_is_kind_salted():
    """The same workbook uploaded as Micro and as Macro parses different
    sheets — the two cache entries must never collide."""
    assert file_hash(b"same-bytes") != file_hash(b"same-bytes", "macro")
    assert file_hash(b"same-bytes") == file_hash(b"same-bytes", "micro")


@pytest.fixture
def dual_split_verdicts(tmp_path, monkeypatch, perim_index, macro_index):
    """Pipeline over BOTH lists. Adds a macro-only blogger to the mini PLOG."""
    _mini_plog(tmp_path / "p.xlsx")
    from openpyxl import load_workbook as _lw
    wb = _lw(str(tmp_path / "p.xlsx"))
    ws = wb["MASTER KOL LIST"]
    ws.append([7, "", "C#1", "", "", "宏观大V", 1, datetime(2026, 6, 1), "",
               "http://xhslink.com/o/macro-only", 1, 1, 1, 1, 3, 1, 1, 1])
    wb.save(str(tmp_path / "p.xlsx"))
    _mini_dmr(tmp_path / "d.xlsx")

    table = dict(RESOLUTIONS)
    table["http://xhslink.com/o/macro-only"] = (U_MACRO,
                                                "6a1a0000000000000000e006")

    def fake_resolve(url, run_counter=None, retry_failed=False):
        if url in table:
            author, note = table[url]
            return Resolution(status="ok", note_id=note, author_id=author,
                              source="fixture")
        return Resolution(status="failed", error="dead link")

    import app.reconciler.pipeline as m
    monkeypatch.setattr(m, "resolve_link", fake_resolve)
    monkeypatch.setattr(m, "ensure_author",
                        lambda url, res, run_counter=None, retry_failed=False: res)
    plog = parse_plog(str(tmp_path / "p.xlsx"))
    dmr = parse_dmr(str(tmp_path / "d.xlsx"))
    vs = run_pipeline(plog, dmr, perimeter=perim_index,
                      perimeter_macro=macro_index)
    return {v.name: v for v in vs}


def test_membership_both(dual_split_verdicts):
    v = dual_split_verdicts["一颗鸡蛋🥚"]      # U_IN is in Micro AND Macro
    assert v.status == NO_POST_IN_PERIMETER
    assert v.perimeter_membership == "both"
    assert v.column_s() == "无博主但在Perimeter内→无帖子"
    assert any("Micro perimeter" in n for n in v.notes)
    assert any("Macro perimeter" in n for n in v.notes)


def test_membership_macro_only(dual_split_verdicts):
    v = dual_split_verdicts["宏观大V"]
    assert v.status == NO_POST_IN_PERIMETER
    assert v.perimeter_membership == "macro"
    assert v.perimeter_redbook_id == U_MACRO
    assert any("Macro perimeter" in n for n in v.notes)
    assert not any("Micro perimeter" in n for n in v.notes)


def test_membership_none_keeps_name_evidence(dual_split_verdicts):
    v = dual_split_verdicts["小小测试"]        # micro near-miss, not a member
    assert v.status == NO_BLOGGER_NOT_IN_PERIMETER
    assert v.perimeter_membership == "none"
    assert "REDBOOK_ID不同" in v.perimeter_note
    v2 = dual_split_verdicts["查无此人"]
    assert v2.perimeter_membership == "none"


def test_membership_micro_only_single_list(split_verdicts):
    """A micro-only run records membership against the one checked list."""
    v = split_verdicts["一颗鸡蛋🥚"]
    assert v.perimeter_membership == "micro"
    assert split_verdicts["查无此人"].perimeter_membership == "none"
    # rows that were never 无博主 stay blank
    assert split_verdicts["断链博主"].perimeter_membership == ""


def test_membership_column_in_export(dual_split_verdicts, tmp_path):
    from app.reconciler.export import (EVIDENCE_HEADERS, write_annotated_xlsx)
    from openpyxl import load_workbook as _lw
    verdicts = list(dual_split_verdicts.values())
    out = tmp_path / "ann.xlsx"
    write_annotated_xlsx(str(tmp_path / "p.xlsx"), str(out), verdicts,
                         header_row=1, sheet_name="MASTER KOL LIST")
    ws = _lw(str(out))["MASTER KOL LIST"]
    assert ws.cell(row=1, column=22).value == "PERIMETER (无博主)"
    col = {v.name: ws.cell(row=v.excel_row, column=22).value
           for v in verdicts}
    assert col["一颗鸡蛋🥚"] == "Micro + Macro"
    assert col["宏观大V"] == "Macro"
    assert col["查无此人"] == "None"
    assert col["断链博主"] is None                 # never 无博主 → blank
