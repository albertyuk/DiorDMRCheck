"""Account system: setup bootstrap, login, team management, session security."""
from __future__ import annotations


import pytest


@pytest.fixture
def client(tmp_path, monkeypatch):
    from app import config
    monkeypatch.setattr(config, "DATA_DIR", tmp_path)
    monkeypatch.setattr(config, "UPLOAD_DIR", tmp_path / "uploads")
    monkeypatch.setattr(config, "DB_PATH", tmp_path / "test.sqlite3")
    monkeypatch.setattr(config, "APP_PASSWORD", "setup-code-9")
    monkeypatch.setattr(config, "APP_SECRET", "dmr-setup-code-9")
    from fastapi.testclient import TestClient
    from app.main import app
    with TestClient(app, follow_redirects=False) as c:
        yield c


def _setup_admin(client, username="boss", password="password123"):
    return client.post("/setup", data={
        "code": "setup-code-9", "username": username, "password": password,
        "display": "The Boss"})


def test_no_users_redirects_to_setup(client):
    r = client.get("/")
    assert r.status_code == 303 and r.headers["location"] == "/setup"


def test_setup_rejects_wrong_code(client):
    r = client.post("/setup", data={"code": "nope", "username": "boss",
                                    "password": "password123"})
    assert r.status_code == 401
    r = client.get("/")
    assert r.headers["location"] == "/setup"  # still no users


def test_setup_creates_admin_and_signs_in(client):
    r = _setup_admin(client)
    assert r.status_code == 303 and r.headers["location"] == "/"
    assert "dmr_session" in r.cookies
    r = client.get("/")
    assert r.status_code == 200  # session accepted


def test_session_cookie_is_secure_in_deployment(client, monkeypatch):
    from app import config
    monkeypatch.setattr(config, "SESSION_COOKIE_SECURE", True)
    r = _setup_admin(client)
    assert "Secure" in r.headers["set-cookie"]


def test_login_flow(client):
    _setup_admin(client)
    client.cookies.clear()
    r = client.get("/")
    assert r.headers["location"] == "/login"  # users exist → login, not setup
    r = client.post("/login", data={"username": "BOSS",  # case-insensitive
                                    "password": "password123"})
    assert r.status_code == 303
    r = client.post("/login", data={"username": "boss", "password": "wrong"})
    assert r.status_code == 401


def test_forged_session_rejected(client):
    _setup_admin(client)
    client.cookies.clear()
    client.cookies.set("dmr_session", "boss|9999999999|deadbeef")
    r = client.get("/")
    assert r.status_code == 303 and r.headers["location"] == "/login"


def test_run_error_page_does_not_disclose_internal_traceback(client):
    from app.core import db

    _setup_admin(client)
    db.run_create("failed-run", plog_path="p", dmr_path="d")
    db.run_update(
        "failed-run",
        status="error",
        phase="error",
        message="Run failed because of an internal error.",
        error="Traceback: SENSITIVE_INTERNAL_DETAIL",
    )

    response = client.get("/runs/failed-run/progress")

    assert response.status_code == 200
    assert "Run failed because of an internal error." in response.text
    assert "Traceback" not in response.text
    assert "SENSITIVE_INTERNAL_DETAIL" not in response.text


def test_admin_adds_coworker_who_can_sign_in(client):
    _setup_admin(client)
    r = client.post("/team/add", data={"username": "mei", "password": "meipassword",
                                       "display": "Mei", "is_admin": "0"})
    assert r.status_code == 303
    client.cookies.clear()
    r = client.post("/login", data={"username": "mei", "password": "meipassword"})
    assert r.status_code == 303
    # non-admin cannot add accounts
    r = client.post("/team/add", data={"username": "x1", "password": "xpassword1"})
    assert "Only+admins" in r.headers["location"] or "Only%20admins" in r.headers["location"]
    from app.core import db
    assert db.user_get("x1") is None


def test_non_admin_cannot_approve_or_revoke_shared_header_mapping(client):
    from app.remap import mapper, service as remap_service

    _setup_admin(client)
    client.post("/team/add", data={"username": "mei",
                                   "password": "meipassword"})
    client.cookies.clear()
    assert client.post("/login", data={"username": "mei",
                                       "password": "meipassword"}).status_code == 303
    token = remap_service.PENDING_MAPS.put({
        "flow": "test", "audits": {}, "names": {},
    })
    try:
        assert client.post(f"/remap/{token}/apply", data={}).status_code == 403
        sig = "b" * 32
        mapper.cache_put("eff", sig, "S", 1, {"name": 1}, "boss")
        assert client.post(
            f"/remap/cache/eff/{sig}/delete", data={}).status_code == 403
        assert mapper.cache_get("eff", sig) is not None
    finally:
        remap_service.PENDING_MAPS.clear()
        mapper.cache_delete("eff", "b" * 32)


def test_non_admin_bad_data_cannot_revoke_valid_cached_mapping(client):
    import io

    from openpyxl import load_workbook

    from app.remap import mapper
    from tests.web.test_schema_map import (CN_EFF_PROPOSAL,
                                           build_cn_plog_bytes)

    _setup_admin(client)
    client.post("/team/add", data={"username": "mei",
                                   "password": "meipassword"})
    client.cookies.clear()
    assert client.post("/login", data={"username": "mei",
                                       "password": "meipassword"}).status_code == 303

    original = build_cn_plog_bytes()
    sig = mapper.header_signature(original, "达人列表", 2)
    mapper.cache_put(
        "eff", sig, "达人列表", 2, CN_EFF_PROPOSAL["columns"], "boss")
    wb = load_workbook(io.BytesIO(original))
    ws = wb["达人列表"]
    for row in range(3, ws.max_row + 1):
        ws.cell(row=row, column=16).value = -1
    invalid = io.BytesIO()
    wb.save(invalid)

    try:
        response = client.post(
            "/efficiency",
            files={"report": (
                "invalid-data.xlsx", invalid.getvalue(),
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet")},
            data={},
        )
        assert response.status_code == 422
        assert mapper.cache_get("eff", sig) is not None
    finally:
        mapper.cache_delete("eff", sig)


def test_only_admin_can_remove_or_promote_shared_perimeter(client, monkeypatch):
    from app import config
    from app.core import db
    from app.reconciler import runs

    _setup_admin(client)
    client.post("/team/add", data={"username": "mei",
                                   "password": "meipassword"})
    db.setting_set("current_perimeter", '{"hash":"keep"}')
    client.cookies.clear()
    assert client.post("/login", data={"username": "mei",
                                       "password": "meipassword"}).status_code == 303
    assert client.post("/perimeter/remove").status_code == 403
    assert db.setting_get("current_perimeter") == '{"hash":"keep"}'
    denied_upload = client.post("/upload", files={
        "plog": ("p.xlsx", b"not-read"),
        "dmr": ("d.xlsx", b"not-read"),
        "perimeter": ("perimeter.xlsx", b"not-read"),
    })
    assert denied_upload.status_code == 403

    run_dir = config.UPLOAD_DIR / "member-perimeter"
    run_dir.mkdir(parents=True, exist_ok=True)
    db.run_create(
        "member-perimeter", plog_path="p", dmr_path="d",
        perimeter_hash="uploaded-hash", perimeter_uploaded=True,
        perimeter_name="uploaded.xlsx",
    )
    started = []
    monkeypatch.setattr(runs, "start_run", lambda run_id: started.append(run_id))
    denied = client.post("/runs/member-perimeter/start", data={})
    assert denied.status_code == 403
    assert db.run_get("member-perimeter")["status"] == "pending"
    assert not started

    db.run_create(
        "legacy-perimeter", plog_path="p", dmr_path="d",
        perimeter_hash="legacy-hash", perimeter_uploaded=None,
        perimeter_name="legacy.xlsx",
    )
    with db.connect() as conn:
        conn.execute(
            "UPDATE runs SET perimeter_uploaded = NULL WHERE id = ?",
            ("legacy-perimeter",),
        )
        conn.commit()
    legacy_denied = client.post("/runs/legacy-perimeter/start", data={})
    assert legacy_denied.status_code == 403
    assert db.run_get("legacy-perimeter")["status"] == "pending"

    client.cookies.clear()
    assert client.post("/login", data={"username": "boss",
                                       "password": "password123"}).status_code == 303
    assert client.post("/perimeter/remove").status_code == 303
    assert db.setting_get("current_perimeter") is None


def test_initial_password_field_is_masked(client):
    _setup_admin(client)
    page = client.get("/team")
    assert page.status_code == 200
    assert 'type="password" name="password"' in page.text
    assert 'autocomplete="new-password"' in page.text


def test_cannot_delete_last_admin_or_self(client):
    _setup_admin(client)
    r = client.post("/team/delete", data={"username": "boss"})
    assert "cannot" in r.headers["location"].lower() or "own" in r.headers["location"].lower()
    from app.core import db
    assert db.user_get("boss") is not None


def test_password_change_and_reset(client):
    _setup_admin(client)
    client.post("/team/add", data={"username": "mei", "password": "meipassword"})
    # admin resets mei's password
    r = client.post("/team/password", data={"username": "mei",
                                            "password": "newpassword9"})
    assert r.status_code == 303
    client.cookies.clear()
    assert client.post("/login", data={"username": "mei",
                                       "password": "newpassword9"}).status_code == 303
    # mei (non-admin) cannot reset boss's password
    r = client.post("/team/password", data={"username": "boss",
                                            "password": "hijacked999"})
    client.cookies.clear()
    assert client.post("/login", data={"username": "boss",
                                       "password": "password123"}).status_code == 303


def test_password_change_revokes_existing_session(client):
    _setup_admin(client)
    stolen = client.cookies.get("dmr_session")
    assert stolen

    client.post("/team/password", data={"username": "boss",
                                        "password": "newpassword9"})
    client.cookies.clear()
    client.cookies.set("dmr_session", stolen)
    r = client.get("/")
    assert r.status_code == 303 and r.headers["location"] == "/login"


def test_setup_recovers_admin_password(client):
    _setup_admin(client)
    client.cookies.clear()
    r = client.post("/setup", data={"code": "setup-code-9", "username": "boss",
                                    "password": "recovered999"})
    assert r.status_code == 303
    client.cookies.clear()
    assert client.post("/login", data={"username": "boss",
                                       "password": "recovered999"}).status_code == 303


def test_open_mode_without_app_password(client, monkeypatch):
    from app import config
    monkeypatch.setattr(config, "APP_PASSWORD", "")
    r = client.get("/")
    assert r.status_code == 200


def test_missing_password_fails_closed(client, monkeypatch):
    from app import config
    monkeypatch.setattr(config, "APP_PASSWORD", "")
    monkeypatch.setattr(config, "ALLOW_OPEN_ACCESS", False)
    r = client.get("/")
    assert r.status_code == 503


def test_runtime_requires_explicit_open_mode(monkeypatch):
    from app import config
    monkeypatch.setattr(config, "APP_PASSWORD", "")
    monkeypatch.setattr(config, "ALLOW_OPEN_ACCESS", False)
    with pytest.raises(RuntimeError, match="APP_PASSWORD is required"):
        config.validate_runtime()
    monkeypatch.setattr(config, "ALLOW_OPEN_ACCESS", True)
    config.validate_runtime()


# ------------------------------------------------- signing-secret derivation

def test_secret_not_derived_from_setup_code(client, monkeypatch):
    """Regression: with APP_SECRET unset, a cookie signed with the old
    'dmr-' + APP_PASSWORD derivation must NOT validate — anyone holding the
    shared setup code could forge any user's session."""
    import hashlib
    import hmac as hmac_mod
    import time as time_mod
    from app import config
    from app.auth import service

    _setup_admin(client)
    monkeypatch.setattr(config, "APP_SECRET", "")
    service._secret_cache.clear()
    client.cookies.clear()
    payload = f"boss|{int(time_mod.time()) + 3600}"
    forged_sig = hmac_mod.new(("dmr-" + config.APP_PASSWORD).encode(),
                              payload.encode(), hashlib.sha256).hexdigest()
    client.cookies.set("dmr_session", f"{payload}|{forged_sig}")
    r = client.get("/")
    assert r.status_code == 303 and r.headers["location"] == "/login"


def test_generated_secret_is_persisted_and_survives_restart(client, monkeypatch):
    """With APP_SECRET unset a random secret is generated, written under
    DATA_DIR, and re-read after a 'restart' (cache cleared) so existing
    sessions keep validating."""
    from app import config
    from app.auth import service

    monkeypatch.setattr(config, "APP_SECRET", "")
    service._secret_cache.clear()
    _setup_admin(client)
    token = service.make_session("boss")
    secret_file = config.DATA_DIR / "session_secret"
    assert secret_file.exists() and len(secret_file.read_text().strip()) >= 64
    service._secret_cache.clear()          # simulate process restart
    assert service.read_session(token) == "boss"
    # a genuine login round-trip works end to end
    client.cookies.clear()
    r = client.post("/login", data={"username": "boss",
                                    "password": "password123"})
    assert r.status_code == 303
    assert client.get("/").status_code == 200


# ------------------------------------------------------------- throttling

@pytest.fixture(autouse=True)
def _fresh_throttle():
    from app.auth import throttle
    throttle.reset()
    yield
    throttle.reset()


def test_login_throttles_after_repeated_failures(client):
    _setup_admin(client)
    client.get("/logout")
    for _ in range(5):
        r = client.post("/login", data={"username": "boss",
                                        "password": "wrong-password"})
        assert r.status_code == 401
    r = client.post("/login", data={"username": "boss",
                                    "password": "wrong-password"})
    assert r.status_code == 429
    assert "wait" in r.text and "seconds" in r.text
    # even the CORRECT password is refused while blocked — guesses stay cheap
    r = client.post("/login", data={"username": "boss",
                                    "password": "password123"})
    assert r.status_code == 429


def test_login_success_resets_the_failure_count(client):
    _setup_admin(client)
    client.get("/logout")
    for _ in range(4):
        client.post("/login", data={"username": "boss", "password": "nope!"})
    r = client.post("/login", data={"username": "boss",
                                    "password": "password123"})
    assert r.status_code == 303                      # signed in
    client.get("/logout")
    for _ in range(4):                               # counter started over
        r = client.post("/login", data={"username": "boss", "password": "no"})
        assert r.status_code == 401


def test_setup_code_guessing_throttled(client):
    for _ in range(5):
        r = client.post("/setup", data={"code": "guess", "username": "x",
                                        "password": "password123"})
        assert r.status_code == 401
    r = client.post("/setup", data={"code": "guess", "username": "x",
                                    "password": "password123"})
    assert r.status_code == 429
    # …and the block applies even when the code is suddenly right
    r = client.post("/setup", data={"code": "setup-code-9", "username": "x",
                                    "password": "password123"})
    assert r.status_code == 429


def test_throttle_window_expires(client, monkeypatch):
    from app.auth import throttle
    _setup_admin(client)
    client.get("/logout")
    for _ in range(5):
        client.post("/login", data={"username": "boss", "password": "bad"})
    assert client.post("/login", data={
        "username": "boss", "password": "password123"}).status_code == 429
    real_time = throttle.time.time
    monkeypatch.setattr(throttle.time, "time",
                        lambda: real_time() + throttle.WINDOW_SECONDS + 1)
    assert client.post("/login", data={
        "username": "boss", "password": "password123"}).status_code == 303
