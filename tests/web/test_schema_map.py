"""LLM header mapping + human audit.

The LLM call is mocked throughout — tests pin the contract around it: the
proposal is validated, NOTHING applies without approval, applying rewrites
only header cells, approved mappings are cached by header signature, and the
efficiency flow never touches disk.
"""
from __future__ import annotations

import asyncio
import io
import json
import threading
import uuid
import zipfile
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime
from pathlib import Path

import pytest
from fastapi.responses import HTMLResponse
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

from app import config
from app import main as main_mod
from app.reconciler import routes as reconciler_routes
from app.remap import mapper as schema_map
from app.remap import routes as remap_routes
from app.remap import service as remap_service
from app.core.textnorm import header_key

# ------------------------------------------------------------------ fixtures

CN_PLOG_HEADERS = ["序号", "机构", "项目", "形式", "级别", "博主昵称",
                   "粉丝(千)", "发布日期", "备注", "笔记链接", "曝光量",
                   "点赞", "收藏", "评论", "互动总量", "价格"]


def build_cn_plog_bytes(metadata="内部使用 KOL 投放追踪表",
                        wave="WAVE #1", n_rows=3) -> bytes:
    """A PLOG-equivalent tracker with Chinese headers — the fingerprint
    (NAME + POST LINK) cannot bind. Parameters vary the metadata line and
    the data rows WITHOUT touching the header layout (cache-key tests)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "达人列表"
    ws.append([metadata])                           # metadata row above header
    ws.append(CN_PLOG_HEADERS)
    for no in range(1, n_rows + 1):
        ws.append([no, "MCN", wave, "报备图文", "KOC", f"博主{no}",
                   88, datetime(2026, 6, no), "", f"http://xhslink.com/cn{no}",
                   10000 * no, 500, 40, 10, 550, 2000])
    link_cell = ws.cell(row=3, column=10)           # hyperlink must survive
    link_cell.hyperlink = "http://xhslink.com/cn1"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


CN_PLOG_PROPOSAL = {
    "sheet": "达人列表", "header_row": 2,
    "columns": {"no": 1, "campaign": 3, "name": 6, "postdate": 8,
                "postlink": 10, "impression": 11, "like": 12,
                "collection": 13, "comment": 14, "ttlengagement": 15},
    "confidence": {"name": 0.97, "postlink": 0.95, "postdate": 0.9,
                   "impression": 0.7},
    "warnings": ["粉丝(千) appears to be follower count in thousands"],
}


def _clear_mapping_cache():
    from app.core import db
    with db.connect() as conn:
        conn.execute("DELETE FROM settings WHERE key LIKE 'schemamap:%'")
        conn.commit()


@pytest.fixture
def client(tmp_path, monkeypatch):
    monkeypatch.setattr(config, "APP_PASSWORD", "")
    monkeypatch.setattr(config, "ANTHROPIC_API_KEY", "test-key")
    monkeypatch.setattr(config, "UPLOAD_DIR", tmp_path / "uploads")
    monkeypatch.setattr(
        schema_map, "_call_llm",
        lambda system, user: json.dumps(CN_PLOG_PROPOSAL, ensure_ascii=False))
    remap_service.PENDING_MAPS.clear()
    _clear_mapping_cache()      # tests share one SQLite settings table
    with TestClient(main_mod.app) as c:
        yield c
    remap_service.PENDING_MAPS.clear()
    _clear_mapping_cache()


def _upload_run(client, plog_bytes, *, allow_header_ai=True):
    from tests import fixtures
    import tempfile
    import os
    fd, dmr_path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    fixtures.build_dmr(dmr_path)
    dmr_bytes = open(dmr_path, "rb").read()
    os.unlink(dmr_path)
    mime = ("application/vnd.openxmlformats-officedocument"
            ".spreadsheetml.sheet")
    return client.post("/upload", files={
        "plog": ("cn_tracker.xlsx", plog_bytes, mime),
        "dmr": ("dmr.xlsx", dmr_bytes, mime),
    }, data={"allow_header_ai": "1" if allow_header_ai else "0"},
        follow_redirects=False)


# ----------------------------------------------------------------- invariants

def test_canonical_headers_roundtrip_through_header_key():
    """apply_mapping writes the canonical text; the parsers re-derive keys
    from it — every canonical text must normalize to exactly its key."""
    for kind, fields in schema_map.FIELDS.items():
        for text, key, _req, _desc in fields:
            assert header_key(text) == key, (kind, text, key)


def test_field_descriptions_have_chinese_translations():
    from app.i18n import ZH
    for fields in schema_map.FIELDS.values():
        for _text, _key, _req, desc in fields:
            assert desc in ZH, f"audit-page description missing zh: {desc!r}"


def test_apply_mapping_touches_only_header_cells():
    data = build_cn_plog_bytes()
    out = schema_map.apply_mapping(
        data, "plog", "达人列表", 2,
        {k: v for k, v in CN_PLOG_PROPOSAL["columns"].items()})
    wb = load_workbook(io.BytesIO(out))
    ws = wb["达人列表"]
    assert ws.cell(row=2, column=6).value == "NAME"
    assert ws.cell(row=2, column=10).value == "POST LINK"
    assert ws.cell(row=2, column=2).value == "机构"        # unmapped: untouched
    assert ws.cell(row=1, column=1).value == "内部使用 KOL 投放追踪表"
    assert ws.cell(row=3, column=6).value == "博主1"        # data untouched
    assert ws.cell(row=3, column=10).hyperlink is not None  # hyperlink survives
    from app.reconciler.parsers import parse_plog
    import tempfile
    import os
    fd, p = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    open(p, "wb").write(out)
    parsed = parse_plog(p)
    os.unlink(p)
    assert len(parsed.rows) == 3 and parsed.rows[0].name == "博主1"


def test_apply_mapping_decollides_duplicate_canonical_headers():
    wb = Workbook()
    ws = wb.active
    ws.title = "S"
    ws.append(["NAME", "真实昵称", "链接"])   # a stray literal NAME column
    ws.append(["wrong", "right", "http://x"])
    buf = io.BytesIO()
    wb.save(buf)
    out = schema_map.apply_mapping(buf.getvalue(), "plog", "S", 1,
                                   {"name": 2, "postlink": 3})
    ws2 = load_workbook(io.BytesIO(out))["S"]
    assert ws2.cell(row=1, column=2).value == "NAME"
    assert ws2.cell(row=1, column=1).value == "(original) NAME"


def test_apply_mapping_preserves_formula_cached_values():
    """A remap must not clear formula results needed by data_only parsers."""
    from lxml import etree
    wb = Workbook()
    ws = wb.active
    ws.title = "S"
    ws.append(["昵称", "链接", "价格"])
    ws.append(["right", "https://x/1", "=1+1"])
    buf = io.BytesIO()
    wb.save(buf)

    # openpyxl cannot author a cached result, so model the Excel-produced XML.
    source = io.BytesIO(buf.getvalue())
    cached = io.BytesIO()
    ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    with zipfile.ZipFile(source) as zin, zipfile.ZipFile(cached, "w") as zout:
        for info in zin.infolist():
            payload = zin.read(info.filename)
            if info.filename == "xl/worksheets/sheet1.xml":
                root = etree.fromstring(payload)
                cell = root.find(f".//{{{ns}}}c[@r='C2']")
                value = cell.find(f"{{{ns}}}v")
                value.text = "2"
                payload = etree.tostring(root, xml_declaration=True,
                                         encoding="UTF-8")
            zout.writestr(info, payload)

    out = schema_map.apply_mapping(
        cached.getvalue(), "plog", "S", 1, {"name": 1, "postlink": 2})
    formula_ws = load_workbook(io.BytesIO(out), data_only=False)["S"]
    values_ws = load_workbook(io.BytesIO(out), data_only=True)["S"]
    assert formula_ws["C2"].value == "=1+1"
    assert values_ws["C2"].value == 2


def test_model_sample_excludes_hidden_sheets():
    wb = Workbook()
    visible = wb.active
    visible.title = "Visible"
    visible.append(["public layout"])
    hidden = wb.create_sheet("Hidden secrets")
    hidden.sheet_state = "hidden"
    hidden.append(["CLIENT-SECRET-SHOULD-NOT-LEAVE"])
    buf = io.BytesIO()
    wb.save(buf)
    sample = schema_map.build_sample(buf.getvalue())
    assert [sheet["name"] for sheet in sample["sheets"]] == ["Visible"]
    assert "CLIENT-SECRET" not in json.dumps(sample)


def test_cache_candidates_use_same_visible_sheet_window_as_model_sample():
    wb = Workbook()
    first = wb.active
    first.title = "Hidden 1"
    hidden = [first]
    for index in range(2, 7):
        hidden.append(wb.create_sheet(f"Hidden {index}"))
    target = wb.create_sheet("Visible target")
    target.append(["昵称", "链接"])
    for sheet in hidden:
        sheet.sheet_state = "hidden"
        sheet.append(["lookup"])
    buf = io.BytesIO()
    wb.save(buf)

    sample = schema_map.build_sample(buf.getvalue())
    candidates = schema_map.candidate_signatures(buf.getvalue())

    assert [sheet["name"] for sheet in sample["sheets"]] == ["Visible target"]
    assert any(sheet == "Visible target" for sheet, _row, _sig in candidates)
    assert not any(sheet.startswith("Hidden")
                   for sheet, _row, _sig in candidates)


def test_mapper_xml_parser_does_not_resolve_external_entities(tmp_path):
    secret = tmp_path / "secret.txt"
    secret.write_text("XXE-SHOULD-NOT-BE-READ")
    xml = (f'<!DOCTYPE root [<!ENTITY xxe SYSTEM "{secret.as_uri()}">]>'
           '<root>&xxe;</root>').encode()
    root = schema_map._safe_xml(xml)
    assert "XXE-SHOULD-NOT-BE-READ" not in "".join(root.itertext())


def test_propose_drops_out_of_range_and_duplicate_columns(monkeypatch):
    monkeypatch.setattr(config, "ANTHROPIC_API_KEY", "k")
    bad = dict(CN_PLOG_PROPOSAL,
               columns={"name": 6, "postlink": 6, "impression": 999})
    monkeypatch.setattr(schema_map, "_call_llm",
                        lambda s, u: json.dumps(bad, ensure_ascii=False))
    sample = schema_map.build_sample(build_cn_plog_bytes())
    prop = schema_map.propose(sample, "plog")
    assert prop.columns["name"] == 6
    assert prop.columns["postlink"] is None      # duplicate dropped
    assert prop.columns["impression"] is None    # out of range dropped


# ----------------------------------------------------------------- run flow

def test_run_flow_audit_then_approve(client):
    r = _upload_run(client, build_cn_plog_bytes())
    assert r.status_code == 303 and r.headers["location"].startswith("/remap/")
    token = r.headers["location"].rsplit("/", 1)[1]
    sig = remap_service.PENDING_MAPS[token]["audits"]["plog"]["sig"]

    page = client.get(f"/remap/{token}")
    assert page.status_code == 200
    body = page.text
    assert "博主昵称" in body                      # original header shown
    assert "博主1" in body                         # sample values shown
    assert "follower count in thousands" in body   # model warning surfaced
    assert "Nothing runs until you approve." in body
    assert "97%" in body                           # confidence shown

    form = {f"plog:{k}": str(v)
            for k, v in CN_PLOG_PROPOSAL["columns"].items()}
    r2 = client.post(f"/remap/{token}/apply", data=form)
    assert r2.status_code == 200
    assert "Parse preview" in r2.text
    assert "Headers remapped" in r2.text           # audit trail on preview
    assert "cn_tracker.xlsx" in r2.text or "WAVE #1" in r2.text
    assert f"/remap/cache/plog/{sig}/delete" in r2.text
    assert "Revoke mapping" in r2.text
    assert schema_map.cache_get("plog", sig) is not None

    revoked = client.post(
        f"/remap/cache/plog/{sig}/delete",
        data={"next_path": "/"},
        follow_redirects=False,
    )
    assert revoked.status_code == 303 and revoked.headers["location"] == "/"
    assert schema_map.cache_get("plog", sig) is None


def test_run_flow_required_field_missing_bounces_back(client):
    r = _upload_run(client, build_cn_plog_bytes())
    token = r.headers["location"].rsplit("/", 1)[1]
    form = {f"plog:{k}": str(v)
            for k, v in CN_PLOG_PROPOSAL["columns"].items() if k != "postlink"}
    r2 = client.post(f"/remap/{token}/apply", data=form)
    assert r2.status_code == 422
    assert "POST LINK" in r2.text                  # names the missing field
    assert token in remap_service.PENDING_MAPS         # still pending, fixable


def test_run_flow_reject_discards(client):
    r = _upload_run(client, build_cn_plog_bytes())
    token = r.headers["location"].rsplit("/", 1)[1]
    run_dir = Path(remap_service.PENDING_MAPS[token]["run_dir"])
    assert run_dir.exists()
    r2 = client.post(f"/remap/{token}/reject", follow_redirects=False)
    assert r2.status_code == 303 and r2.headers["location"] == "/"
    assert token not in remap_service.PENDING_MAPS
    assert not run_dir.exists()


def test_approved_mapping_is_cached_and_auto_applied(client):
    r = _upload_run(client, build_cn_plog_bytes())
    token = r.headers["location"].rsplit("/", 1)[1]
    form = {f"plog:{k}": str(v)
            for k, v in CN_PLOG_PROPOSAL["columns"].items()}
    client.post(f"/remap/{token}/apply", data=form)

    # same format again — no audit page, straight to preview, visibly noted
    r2 = _upload_run(client, build_cn_plog_bytes())
    assert r2.status_code == 200
    assert "Parse preview" in r2.text
    assert "Applied automatically" in r2.text
    assert "open-mode" in r2.text                  # who approved it


def test_no_api_key_keeps_plain_error(client, monkeypatch):
    monkeypatch.setattr(config, "ANTHROPIC_API_KEY", "")
    r = _upload_run(client, build_cn_plog_bytes())
    assert r.status_code == 422
    assert "PLOG parse failed" in r.text


def test_remap_inspection_does_not_mask_cache_outages(monkeypatch):
    def cache_outage(*_args, **_kwargs):
        raise RuntimeError("sqlite unavailable")

    monkeypatch.setattr(schema_map, "cache_get_many", cache_outage)

    with pytest.raises(RuntimeError, match="sqlite unavailable"):
        remap_service.inspect_remap("plog", build_cn_plog_bytes())


def test_efficiency_cache_outage_is_logged_generic_500(client, monkeypatch,
                                                        caplog):
    def cache_outage(*_args, **_kwargs):
        raise RuntimeError("sensitive sqlite path")

    monkeypatch.setattr(schema_map, "cache_get_many", cache_outage)
    response = client.post(
        "/efficiency",
        files={"report": (
            "cn_wave.xlsx", build_cn_plog_bytes(),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={},
    )

    assert response.status_code == 500
    assert "Unexpected server error" in response.text
    assert "sensitive sqlite path" not in response.text
    assert "unexpected efficiency header-mapping failure" in caplog.text


def test_reconciler_requires_explicit_consent_before_header_sample_egress(
        client, monkeypatch):
    called = False

    def unexpected_call(*_args, **_kwargs):
        nonlocal called
        called = True
        raise AssertionError("Claude must not be called without opt-in")

    monkeypatch.setattr(schema_map, "_call_llm", unexpected_call)
    response = _upload_run(
        client, build_cn_plog_bytes(), allow_header_ai=False
    )
    assert response.status_code == 422
    assert "not explicitly allowed" in response.text
    assert not called


# ----------------------------------------------------------- efficiency flow

CN_EFF_PROPOSAL = {
    "sheet": "达人列表", "header_row": 2,
    "columns": {"no": 1, "campaign": 3, "type": 4, "level": 5, "name": 6,
                "fanbase(k)": 7, "postdate": 8, "postlink": 10,
                "impression": 11, "like": 12, "collection": 13, "comment": 14,
                "ttlengagement": 15, "price": 16},
    "confidence": {}, "warnings": [],
}


def test_efficiency_unfamiliar_headers_require_explicit_external_opt_in(
        client, monkeypatch):
    calls = 0

    def forbidden_model_call(*_args):
        nonlocal calls
        calls += 1
        raise AssertionError("model must not be called without consent")

    monkeypatch.setattr(schema_map, "_call_llm", forbidden_model_call)
    response = client.post(
        "/efficiency",
        files={"report": ("cn_wave.xlsx", build_cn_plog_bytes(),
                          "application/vnd.openxmlformats-officedocument."
                          "spreadsheetml.sheet")},
        data={"allow_header_ai": "0"}, follow_redirects=False)
    assert response.status_code == 422
    assert "No workbook sample was sent externally" in response.text
    assert calls == 0


def test_efficiency_model_call_is_not_run_under_workbook_gate(client,
                                                               monkeypatch):
    from app.efficiency import routes as eff_routes
    gated_functions = []
    real_runner = eff_routes.run_upload_task
    real_propose = schema_map.propose

    async def recording_runner(request, func, *args, **kwargs):
        gated_functions.append(func)
        return await real_runner(request, func, *args, **kwargs)

    def checked_propose(*args, **kwargs):
        assert schema_map.propose not in gated_functions
        return real_propose(*args, **kwargs)

    monkeypatch.setattr(eff_routes, "run_upload_task", recording_runner)
    monkeypatch.setattr(schema_map, "propose", checked_propose)
    monkeypatch.setattr(
        schema_map, "_call_llm",
        lambda *_args: json.dumps(CN_EFF_PROPOSAL, ensure_ascii=False))
    response = client.post(
        "/efficiency",
        files={"report": ("cn_wave.xlsx", build_cn_plog_bytes(),
                          "application/vnd.openxmlformats-officedocument."
                          "spreadsheetml.sheet")},
        data={"allow_header_ai": "1"}, follow_redirects=False)
    assert response.status_code == 303
    assert schema_map.propose not in gated_functions


def test_efficiency_flow_audit_then_approve_stays_in_memory(client, tmp_path,
                                                            monkeypatch):
    monkeypatch.setattr(
        schema_map, "_call_llm",
        lambda system, user: json.dumps(CN_EFF_PROPOSAL, ensure_ascii=False))
    mime = ("application/vnd.openxmlformats-officedocument"
            ".spreadsheetml.sheet")
    r = client.post("/efficiency",
                    files={"report": ("cn_wave.xlsx", build_cn_plog_bytes(),
                                      mime)},
                    data={"allow_header_ai": "1"}, follow_redirects=False)
    assert r.status_code == 303 and r.headers["location"].startswith("/remap/")
    token = r.headers["location"].rsplit("/", 1)[1]
    assert client.get(f"/remap/{token}").status_code == 200

    form = {f"eff:{k}": str(v) for k, v in CN_EFF_PROPOSAL["columns"].items()}
    r2 = client.post(f"/remap/{token}/apply", data=form,
                     follow_redirects=True)
    assert r2.status_code == 200
    assert "Headers remapped" in r2.text           # note on the report page
    assert "Download .pptx" in r2.text
    # privacy: nothing about this workbook ever lands on disk
    uploads = config.UPLOAD_DIR
    assert not uploads.exists() or not any(uploads.rglob("*cn_wave*"))


def test_invalid_efficiency_mapping_is_not_cached(client, monkeypatch):
    monkeypatch.setattr(
        schema_map, "_call_llm",
        lambda *_args: json.dumps(CN_EFF_PROPOSAL, ensure_ascii=False))
    data = build_cn_plog_bytes()
    response = client.post(
        "/efficiency",
        files={"report": ("cn_wave.xlsx", data,
                          "application/vnd.openxmlformats-officedocument."
                          "spreadsheetml.sheet")},
        data={"allow_header_ai": "1"}, follow_redirects=False)
    token = response.headers["location"].rsplit("/", 1)[1]
    sig = remap_service.PENDING_MAPS[token]["audits"]["eff"]["sig"]
    form = {f"eff:{key}": str(value)
            for key, value in CN_EFF_PROPOSAL["columns"].items()}
    form["eff:price"] = "2"  # valid offered column, wrong semantic field
    applied = client.post(f"/remap/{token}/apply", data=form)
    assert applied.status_code == 422
    assert schema_map.cache_get("eff", sig) is None


def test_blocked_diagnostic_report_does_not_promote_mapping(client,
                                                            monkeypatch):
    monkeypatch.setattr(
        schema_map, "_call_llm",
        lambda *_args: json.dumps(CN_EFF_PROPOSAL, ensure_ascii=False))
    source = build_cn_plog_bytes()
    wb = load_workbook(io.BytesIO(source))
    wb["达人列表"].cell(row=3, column=16).value = -1
    modified = io.BytesIO()
    wb.save(modified)

    response = client.post(
        "/efficiency",
        files={"report": (
            "mixed-validity.xlsx", modified.getvalue(),
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet")},
        data={"allow_header_ai": "1"}, follow_redirects=False)
    token = response.headers["location"].rsplit("/", 1)[1]
    sig = remap_service.PENDING_MAPS[token]["audits"]["eff"]["sig"]
    form = {f"eff:{key}": str(value)
            for key, value in CN_EFF_PROPOSAL["columns"].items()}

    applied = client.post(f"/remap/{token}/apply", data=form)

    assert applied.status_code == 200
    assert "Deck not generated" in applied.text
    assert "not saved for reuse" in applied.text
    assert "Revoke mapping" not in applied.text
    assert schema_map.cache_get("eff", sig) is None


def test_cached_efficiency_mapping_survives_row_validation_failure(client):
    data = build_cn_plog_bytes()
    sig = schema_map.header_signature(data, "达人列表", 2)
    wrong = dict(CN_EFF_PROPOSAL["columns"])
    wrong["price"] = 2
    schema_map.cache_put("eff", sig, "达人列表", 2, wrong, "admin")
    response = client.post(
        "/efficiency",
        files={"report": ("cn_wave.xlsx", data,
                          "application/vnd.openxmlformats-officedocument."
                          "spreadsheetml.sheet")},
        data={}, follow_redirects=False)
    assert response.status_code == 422
    assert schema_map.cache_get("eff", sig) is not None


def test_cached_efficiency_mapping_is_revoked_on_schema_failure(client):
    data = build_cn_plog_bytes()
    sig = schema_map.header_signature(data, "达人列表", 2)
    # This corrupt legacy cache entry cannot create the required schema.
    incomplete = {"name": 6, "postlink": 10}
    schema_map.cache_put("eff", sig, "达人列表", 2, incomplete, "admin")

    response = client.post(
        "/efficiency",
        files={"report": ("cn_wave.xlsx", data,
                          "application/vnd.openxmlformats-officedocument."
                          "spreadsheetml.sheet")},
        data={}, follow_redirects=False)

    assert response.status_code == 422
    assert "has been revoked" in response.text
    assert schema_map.cache_get("eff", sig) is None


def test_cached_efficiency_mapping_is_revoked_when_apply_fails(
        client, monkeypatch):
    data = build_cn_plog_bytes()
    sig = schema_map.header_signature(data, "达人列表", 2)
    schema_map.cache_put(
        "eff", sig, "达人列表", 2, CN_EFF_PROPOSAL["columns"], "admin")

    def fail_apply(*_args, **_kwargs):
        raise ValueError("stale cached worksheet mapping")

    monkeypatch.setattr(schema_map, "apply_mapping", fail_apply)
    response = client.post(
        "/efficiency",
        files={"report": ("cn_wave.xlsx", data,
                          "application/vnd.openxmlformats-officedocument."
                          "spreadsheetml.sheet")},
        data={}, follow_redirects=False)

    assert response.status_code == 422
    assert "has been revoked" in response.text
    assert schema_map.cache_get("eff", sig) is None


def test_expired_mapping_token_404s(client):
    r = _upload_run(client, build_cn_plog_bytes())
    token = r.headers["location"].rsplit("/", 1)[1]
    run_dir = Path(remap_service.PENDING_MAPS[token]["run_dir"])
    remap_service.PENDING_MAPS[token]["created"] -= remap_service.PENDING_MAPS.ttl_seconds + 1
    assert client.get(f"/remap/{token}").status_code == 404
    assert not run_dir.exists()
    assert client.post(f"/remap/{token}/apply", data={}).status_code == 404


def test_reconciler_rejects_oversize_and_removes_staging(client, monkeypatch):
    monkeypatch.setattr(config, "MAX_UPLOAD_BYTES", 10)
    mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    r = client.post("/upload", files={
        "plog": ("large.xlsx", b"x" * 11, mime),
        "dmr": ("dmr.xlsx", b"x", mime),
    })
    assert r.status_code == 413
    assert not config.UPLOAD_DIR.exists() or not any(config.UPLOAD_DIR.iterdir())


def test_reconciler_parse_failure_removes_staging(client):
    mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    r = client.post("/upload", files={
        "plog": ("bad.xlsx", b"not a zip", mime),
        "dmr": ("bad.xlsx", b"not a zip", mime),
    })
    assert r.status_code == 422
    assert not config.UPLOAD_DIR.exists() or not any(config.UPLOAD_DIR.iterdir())


def test_reconciler_schema_probe_does_not_duplicate_full_parse(
        client, tmp_path, monkeypatch):
    from tests import fixtures

    plog_path = tmp_path / "normal-plog.xlsx"
    dmr_path = tmp_path / "normal-dmr.xlsx"
    fixtures.build_plog(str(plog_path))
    fixtures.build_dmr(str(dmr_path))
    real_plog = reconciler_routes.parse_plog
    real_dmr = reconciler_routes.parse_dmr
    calls = {"plog": 0, "dmr": 0}

    def counted_plog(path):
        calls["plog"] += 1
        return real_plog(path)

    def counted_dmr(path):
        calls["dmr"] += 1
        return real_dmr(path)

    monkeypatch.setattr(reconciler_routes, "parse_plog", counted_plog)
    monkeypatch.setattr(reconciler_routes, "parse_dmr", counted_dmr)
    mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    response = client.post("/upload", files={
        "plog": ("p.xlsx", plog_path.read_bytes(), mime),
        "dmr": ("d.xlsx", dmr_path.read_bytes(), mime),
    })
    assert response.status_code == 200
    assert calls == {"plog": 1, "dmr": 1}


def test_concurrent_xlsx_exports_use_unique_outputs_and_clean_streaming_leases(
        client, monkeypatch):
    from app.core import db
    from app.core.uploads import active_upload_names
    from tests import fixtures

    run_id = "export" + uuid.uuid4().hex[:8]
    run_dir = config.UPLOAD_DIR / run_id
    run_dir.mkdir(parents=True)
    plog_path = run_dir / "plog.xlsx"
    dmr_path = run_dir / "dmr.xlsx"
    fixtures.build_plog(str(plog_path))
    fixtures.build_dmr(str(dmr_path))
    db.run_create(run_id, plog_path=str(plog_path), dmr_path=str(dmr_path))
    db.run_update(
        run_id,
        status="done",
        result_json=json.dumps({
            "verdicts": [],
            "plog_meta": {"header_row": 2, "sheet": "MASTER KOL LIST"},
        }),
    )
    calls = 0
    calls_lock = threading.Lock()
    first_writer_entered = threading.Event()
    both_requests_started = threading.Event()
    release_first_writer = threading.Event()
    output_paths: list[Path] = []
    real_runner = reconciler_routes.run_upload_task
    real_writer = reconciler_routes.write_annotated_xlsx

    async def tracking_runner(*args, **kwargs):
        nonlocal calls
        with calls_lock:
            calls += 1
            if calls == 2:
                both_requests_started.set()
        return await real_runner(*args, **kwargs)

    def tracking_writer(plog_path, output_path, *args, **kwargs):
        with calls_lock:
            output_paths.append(Path(output_path))
            is_first = len(output_paths) == 1
        if is_first:
            first_writer_entered.set()
            assert release_first_writer.wait(2)
        return real_writer(plog_path, output_path, *args, **kwargs)

    monkeypatch.setattr(
        reconciler_routes, "run_upload_task", tracking_runner
    )
    monkeypatch.setattr(
        reconciler_routes, "write_annotated_xlsx", tracking_writer
    )

    try:
        with ThreadPoolExecutor(max_workers=2) as executor:
            first = executor.submit(
                client.get, f"/runs/{run_id}/export.xlsx"
            )
            assert first_writer_entered.wait(2)
            second = executor.submit(
                client.get, f"/runs/{run_id}/export.xlsx"
            )
            assert both_requests_started.wait(2)
            release_first_writer.set()
            responses = [first.result(timeout=5), second.result(timeout=5)]

        download_name = f'filename="PLOG_DMR_CHECK_{run_id}.xlsx"'
        assert all(response.status_code == 200 for response in responses)
        assert all(response.content[:2] == b"PK" for response in responses)
        assert all(
            download_name in response.headers["content-disposition"]
            for response in responses
        )
        assert calls == 2
        assert len(output_paths) == 2
        assert len(set(output_paths)) == 2
        assert all(path.parent == run_dir for path in output_paths)
        assert all(not path.exists() for path in output_paths)
        assert run_id not in active_upload_names()
    finally:
        release_first_writer.set()
        db.run_delete(run_id)


def test_xlsx_export_releases_lease_when_temp_cleanup_fails(client,
                                                              monkeypatch):
    from app.core import db
    from app.core.uploads import (active_upload_names,
                                  unregister_active_upload)
    from tests import fixtures

    run_id = "export" + uuid.uuid4().hex[:8]
    run_dir = config.UPLOAD_DIR / run_id
    run_dir.mkdir(parents=True)
    plog_path = run_dir / "plog.xlsx"
    dmr_path = run_dir / "dmr.xlsx"
    fixtures.build_plog(str(plog_path))
    fixtures.build_dmr(str(dmr_path))
    db.run_create(run_id, plog_path=str(plog_path), dmr_path=str(dmr_path))
    db.run_update(
        run_id,
        status="done",
        result_json=json.dumps({
            "verdicts": [],
            "plog_meta": {"header_row": 2, "sheet": "MASTER KOL LIST"},
        }),
    )

    def failing_writer(_plog_path, output_path, *_args, **_kwargs):
        Path(output_path).write_bytes(b"partial export")
        raise RuntimeError("export failed")

    real_unlink = Path.unlink

    def failing_export_unlink(path, *args, **kwargs):
        if path.parent == run_dir and path.name.startswith(
                f".export-{run_id}-"):
            raise PermissionError("cannot remove partial export")
        return real_unlink(path, *args, **kwargs)

    try:
        with monkeypatch.context() as patch:
            patch.setattr(
                reconciler_routes, "write_annotated_xlsx", failing_writer
            )
            patch.setattr(Path, "unlink", failing_export_unlink)
            with pytest.raises((RuntimeError, PermissionError)):
                client.get(f"/runs/{run_id}/export.xlsx")

        assert run_id not in active_upload_names()
    finally:
        while run_id in active_upload_names():
            unregister_active_upload(run_dir)
        for export_path in run_dir.glob(f".export-{run_id}-*.xlsx"):
            export_path.unlink()
        db.run_delete(run_id)


def test_xlsx_export_stream_admission_is_bounded(client, monkeypatch):
    slots = threading.BoundedSemaphore(1)
    slots.acquire()
    monkeypatch.setattr(reconciler_routes, "_export_stream_slots", slots)

    response = client.get("/runs/not-running/export.xlsx")

    assert response.status_code == 503
    assert response.headers["retry-after"] == "2"
    slots.release()


def test_cached_mapping_failure_is_422_and_removes_staging(client,
                                                            monkeypatch):
    stale_sig = "f" * 64
    cached = {**CN_PLOG_PROPOSAL, "approved_by": "reviewer",
              "sig": stale_sig}
    schema_map.cache_put(
        "plog", stale_sig, cached["sheet"], cached["header_row"],
        cached["columns"], "reviewer",
    )
    monkeypatch.setattr(
        reconciler_routes,
        "inspect_remap",
        lambda _kind, _data: remap_service.RemapOutcome(
            "cached", mapping=cached
        ),
    )
    monkeypatch.setattr(
        schema_map,
        "apply_mapping",
        lambda *_args, **_kwargs: (_ for _ in ()).throw(
            ValueError("stale cached mapping")
        ),
    )

    response = _upload_run(client, build_cn_plog_bytes())

    assert response.status_code == 422
    assert "Header mapping also failed" in response.text
    assert schema_map.cache_get("plog", stale_sig) is None
    assert not config.UPLOAD_DIR.exists() or not any(config.UPLOAD_DIR.iterdir())


def test_cached_mapping_is_evicted_when_remapped_parser_rejects(
        client, monkeypatch):
    stale_sig = "e" * 64
    cached = {**CN_PLOG_PROPOSAL, "approved_by": "reviewer",
              "sig": stale_sig}
    schema_map.cache_put(
        "plog", stale_sig, cached["sheet"], cached["header_row"],
        cached["columns"], "reviewer",
    )
    monkeypatch.setattr(
        reconciler_routes, "inspect_remap",
        lambda _kind, _data: remap_service.RemapOutcome(
            "cached", mapping=cached
        ),
    )
    # Simulate a syntactically successful rewrite that leaves an invalid
    # deterministic schema. The route must validate before trusting/caching.
    monkeypatch.setattr(schema_map, "apply_mapping",
                        lambda data, *_args, **_kwargs: data)
    response = _upload_run(client, build_cn_plog_bytes())
    assert response.status_code == 422
    assert schema_map.cache_get("plog", stale_sig) is None


def _put_pending_plog(flow: str = "test") -> str:
    return remap_service.PENDING_MAPS.put({
        "flow": flow,
        "names": {"plog": "test.xlsx"},
        "audits": {"plog": {
            "proposal": {
                "sheet": "Sheet1", "header_row": 1,
                "columns": {"name": 1, "postlink": 2},
                "confidence": {}, "warnings": [],
            },
            "choices": [
                {"col": 1, "letter": "A", "header": "Name", "samples": []},
                {"col": 2, "letter": "B", "header": "Link", "samples": []},
            ], "sig": "test-signature",
        }},
    })


def _claim_pending_store_to_capacity() -> None:
    for _ in range(remap_service.PENDING_MAPS.max_entries):
        token = _put_pending_plog()
        assert remap_service.PENDING_MAPS.claim(token)[0] == "claimed"


def test_full_mapping_store_returns_503_and_removes_run_staging(client):
    _claim_pending_store_to_capacity()

    response = _upload_run(client, build_cn_plog_bytes())

    assert response.status_code == 503
    assert "Too many mapping audits" in response.text
    assert not config.UPLOAD_DIR.exists() or not any(config.UPLOAD_DIR.iterdir())


def test_full_mapping_store_returns_503_for_efficiency(client, monkeypatch):
    monkeypatch.setattr(
        schema_map, "_call_llm",
        lambda _system, _user: json.dumps(
            CN_EFF_PROPOSAL, ensure_ascii=False
        ),
    )
    _claim_pending_store_to_capacity()
    mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    response = client.post(
        "/efficiency",
        files={"report": ("cn_wave.xlsx", build_cn_plog_bytes(), mime)},
        data={"allow_header_ai": "1"},
        follow_redirects=False,
    )

    assert response.status_code == 503
    assert "Too many mapping audits" in response.text


@pytest.mark.parametrize("bad", ["not-a-number", "3", "0", "-1"])
def test_mapping_apply_rejects_columns_not_offered_by_audit(client, bad):
    token = _put_pending_plog()
    response = client.post(
        f"/remap/{token}/apply",
        data={"plog:name": bad, "plog:postlink": "2"})
    assert response.status_code == 422
    assert "Invalid column selected for NAME" in response.text
    assert token in remap_service.PENDING_MAPS


def test_open_mode_can_revoke_cached_mapping(client):
    sig = "a" * 32
    schema_map.cache_put("eff", sig, "S", 1, {"name": 1}, "operator")
    assert schema_map.cache_get("eff", sig) is not None
    response = client.post(
        f"/remap/cache/eff/{sig}/delete",
        data={"next_path": "/efficiency"}, follow_redirects=False)
    assert response.status_code == 303
    assert response.headers["location"] == "/efficiency"
    assert schema_map.cache_get("eff", sig) is None


def test_apply_claim_allows_only_one_concurrent_handler(client, monkeypatch):
    """Two apply POSTs may validate together, but only one continuation may
    perform the irreversible work for a pending token."""
    token = _put_pending_plog()
    entered = threading.Event()
    finish = threading.Event()
    calls = 0

    async def handler(request, handler_token, entry, approved, username):
        nonlocal calls
        calls += 1
        entered.set()
        assert await asyncio.to_thread(finish.wait, 2)
        return HTMLResponse("applied")

    monkeypatch.setitem(remap_routes.FLOW_HANDLERS, "test", handler)
    form = {"plog:name": "1", "plog:postlink": "2"}
    with ThreadPoolExecutor(max_workers=1) as executor:
        first = executor.submit(
            client.post, f"/remap/{token}/apply", data=form)
        assert entered.wait(2)
        second = client.post(f"/remap/{token}/apply", data=form)
        reject = client.post(f"/remap/{token}/reject")
        finish.set()
        first_response = first.result(timeout=5)

    assert first_response.status_code == 200
    assert second.status_code == 409
    assert reject.status_code == 409
    assert calls == 1
    assert token not in remap_service.PENDING_MAPS


def test_apply_failure_releases_claim_for_retry(client, monkeypatch):
    token = _put_pending_plog()
    attempts = 0

    async def handler(request, handler_token, entry, approved, username):
        nonlocal attempts
        attempts += 1
        if attempts == 1:
            raise RuntimeError("temporary continuation failure")
        return HTMLResponse("applied")

    monkeypatch.setitem(remap_routes.FLOW_HANDLERS, "test", handler)
    form = {"plog:name": "1", "plog:postlink": "2"}
    with pytest.raises(RuntimeError, match="temporary continuation failure"):
        client.post(f"/remap/{token}/apply", data=form)

    assert token in remap_service.PENDING_MAPS
    retry = client.post(f"/remap/{token}/apply", data=form)
    assert retry.status_code == 200
    assert attempts == 2
    assert token not in remap_service.PENDING_MAPS


def test_run_continuation_failure_consumes_token_after_staging_cleanup(
        client, monkeypatch):
    """A fatal failure after the real continuation destroys its run dir must
    not advertise the now-unusable audit token as retryable."""
    from app.core import db

    response = _upload_run(client, build_cn_plog_bytes())
    token = response.headers["location"].rsplit("/", 1)[1]
    run_dir = Path(remap_service.PENDING_MAPS[token]["run_dir"])
    form = {f"plog:{key}": str(value)
            for key, value in CN_PLOG_PROPOSAL["columns"].items()}

    monkeypatch.setattr(
        db, "run_create",
        lambda *_args, **_kwargs: (_ for _ in ()).throw(
            RuntimeError("database unavailable")
        ),
    )
    with pytest.raises(RuntimeError, match="database unavailable"):
        client.post(f"/remap/{token}/apply", data=form)

    assert not run_dir.exists()
    assert token not in remap_service.PENDING_MAPS


def test_preview_render_failure_leaves_no_run_or_retryable_token(
        client, monkeypatch):
    """Rendering is part of the continuation transaction: a response-build
    failure must not commit a run whose audit token can then be retried."""
    from app.core import db

    response = _upload_run(client, build_cn_plog_bytes())
    token = response.headers["location"].rsplit("/", 1)[1]
    entry = remap_service.PENDING_MAPS[token]
    run_id = entry["run_id"]
    run_dir = Path(entry["run_dir"])
    form = {f"plog:{key}": str(value)
            for key, value in CN_PLOG_PROPOSAL["columns"].items()}
    real_template_response = reconciler_routes.templates.TemplateResponse

    def fail_preview(request, name, context, *args, **kwargs):
        if name == "reconciler/preview.html":
            raise RuntimeError("preview render failed")
        return real_template_response(request, name, context, *args, **kwargs)

    monkeypatch.setattr(
        reconciler_routes.templates, "TemplateResponse", fail_preview
    )

    try:
        with pytest.raises(RuntimeError, match="preview render failed"):
            client.post(f"/remap/{token}/apply", data=form)

        assert db.run_get(run_id) is None
        assert not run_dir.exists()
        assert token not in remap_service.PENDING_MAPS
    finally:
        db.run_delete(run_id)


# ------------------------------------------------- layout-keyed cache

def test_header_signature_ignores_data_and_metadata():
    """Regression: the cache key must depend only on the header row's layout.
    The old whole-sample hash covered data rows and the metadata line (which
    carries per-export dates in real DMR files), so the approved cache could
    effectively never hit twice."""
    v1 = build_cn_plog_bytes()
    v2 = build_cn_plog_bytes(metadata="From 2026-07-01 To 2026-07-20 export",
                             wave="WAVE #2", n_rows=7)
    assert (schema_map.header_signature(v1, "达人列表", 2)
            == schema_map.header_signature(v2, "达人列表", 2))
    # a changed header cell IS a different layout
    wb = load_workbook(io.BytesIO(v1))
    wb["达人列表"].cell(row=2, column=6).value = "昵称"
    buf = io.BytesIO()
    wb.save(buf)
    assert (schema_map.header_signature(buf.getvalue(), "达人列表", 2)
            != schema_map.header_signature(v1, "达人列表", 2))
    # candidate enumeration includes the header row's signature
    sigs = {(sheet, row): sig
            for sheet, row, sig in schema_map.candidate_signatures(v2)}
    assert sigs[("达人列表", 2)] == schema_map.header_signature(v1, "达人列表", 2)


def test_header_signature_uses_text_beyond_display_truncation():
    """Headers that look identical in the bounded prompt/UI sample must not
    share an approved-mapping cache identity."""
    prefix = "H" * schema_map.SAMPLE_CELL_CHARS

    def workbook_bytes(header: str) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append([header])
        ws.append(["sample data"])
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    first = workbook_bytes(prefix + "-first-layout")
    second = workbook_bytes(prefix + "-second-layout")

    assert (schema_map.build_sample(first)["sheets"][0]["rows"][0][0]
            == prefix)
    assert schema_map.column_choices(first, "Sheet1", 1)[0]["header"] == prefix
    assert (schema_map.header_signature(first, "Sheet1", 1)
            != schema_map.header_signature(second, "Sheet1", 1))

    first_candidates = {(sheet, row): sig for sheet, row, sig
                        in schema_map.candidate_signatures(first)}
    second_candidates = {(sheet, row): sig for sheet, row, sig
                         in schema_map.candidate_signatures(second)}
    assert (first_candidates[("Sheet1", 1)]
            != second_candidates[("Sheet1", 1)])


def test_cache_hits_for_same_layout_with_different_data(client, monkeypatch):
    """Approve a format once; a later upload of the SAME layout with
    different data and a different metadata line must auto-apply from the
    cache — no audit page and no LLM call."""
    r = _upload_run(client, build_cn_plog_bytes())
    token = r.headers["location"].rsplit("/", 1)[1]
    form = {f"plog:{k}": str(v)
            for k, v in CN_PLOG_PROPOSAL["columns"].items()}
    client.post(f"/remap/{token}/apply", data=form)

    def _no_llm(system, user):
        raise AssertionError("cache miss: LLM was called for a known layout")
    monkeypatch.setattr(schema_map, "_call_llm", _no_llm)

    r2 = _upload_run(client, build_cn_plog_bytes(
        metadata="From 2026-07-01 To 2026-07-20 export",
        wave="WAVE #2", n_rows=9))
    assert r2.status_code == 200
    assert "Parse preview" in r2.text
    assert "Applied automatically" in r2.text
    assert "WAVE #2" in r2.text                    # the new data parsed
