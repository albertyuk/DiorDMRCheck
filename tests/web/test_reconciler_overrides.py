"""Human overrides are authoritative views, with pipeline provenance kept."""
from __future__ import annotations

import io
import json

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

from app import config
from app.core import db
from app.reconciler.domain import NO_POST, Verdict
from tests import fixtures


@pytest.fixture
def completed_run(tmp_path, monkeypatch):
    monkeypatch.setattr(config, "APP_PASSWORD", "")
    monkeypatch.setattr(config, "ALLOW_OPEN_ACCESS", True)
    monkeypatch.setattr(config, "DATA_DIR", tmp_path)
    monkeypatch.setattr(config, "DB_PATH", tmp_path / "overrides.sqlite3")
    monkeypatch.setattr(config, "UPLOAD_DIR", tmp_path / "uploads")
    run_dir = config.UPLOAD_DIR / "override-run"
    run_dir.mkdir(parents=True)
    plog_path = run_dir / "plog.xlsx"
    dmr_path = run_dir / "dmr.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "MASTER KOL LIST"
    ws.append(fixtures.PLOG_HEADERS)
    ws.append([1, "", "REAL-CAMPAIGN", "", "", "real name", 1, None, "",
               "https://xhslink.com/o/x", 1, 1, 1, 1, 3, 1, 1, 1])
    wb.save(plog_path)
    Workbook().save(dmr_path)

    verdict = Verdict(
        campaign="REAL-CAMPAIGN", no="1", name="real name", post_date=None,
        post_link="https://xhslink.com/o/x", excel_row=2,
        status=NO_POST, tier="2:author-id",
        matched_blogger='=HYPERLINK("https://evil.invalid","click")',
        out_of_window=False,
    )
    result = {
        "verdicts": [verdict.to_dict()],
        "counts": {NO_POST: 1},
        "buckets": {"dmr_gaps": 1, "expected_missing": 0,
                    "outside_perimeter": 0},
        "plog_meta": {"header_row": 1, "sheet": ws.title},
        "dmr_meta": {}, "reverse_audit": [],
        "perimeter_meta": {"hash": "perim-hash", "warnings": ["stale"]},
    }

    from app.main import app
    with TestClient(app) as client:
        db.run_create(
            "override-run", plog_path=str(plog_path), dmr_path=str(dmr_path),
            plog_name="plog.xlsx", dmr_name="dmr.xlsx",
        )
        db.run_update(
            "override-run", status="done",
            result_json=json.dumps(result, ensure_ascii=False),
        )
        yield client


def test_override_drives_ui_json_xlsx_and_uses_server_row_identity(completed_run):
    client = completed_run
    invalid = client.post("/runs/override-run/override", data={
        "excel_row": "2", "status": "=1+1",
    })
    assert invalid.status_code == 422

    response = client.post("/runs/override-run/override", data={
        "excel_row": "2", "campaign": "FORGED", "no": "999",
        "status": "无博主", "note": "-1+1",
    })
    assert response.status_code == 200
    assert 'data-status="NO_BLOGGER"' in response.text
    assert "Pipeline status" in response.text
    stored = db.overrides_for_run("override-run")[2]
    assert stored["campaign"] == "REAL-CAMPAIGN" and stored["no"] == "1"

    audit = client.get("/runs/override-run/export.json")
    assert audit.status_code == 200
    doc = audit.json()
    assert doc["counts"] == {"NO_BLOGGER": 1}
    assert doc["pipeline_counts"] == {"NO_POST": 1}
    assert doc["verdicts"][0]["status"] == "NO_BLOGGER"
    assert doc["verdicts"][0]["pipeline_status"] == "NO_POST"

    xlsx = client.get("/runs/override-run/export.xlsx")
    assert xlsx.status_code == 200
    ws = load_workbook(io.BytesIO(xlsx.content), data_only=False)[
        "MASTER KOL LIST"
    ]
    assert ws.cell(2, 19).value == "无博主"
    assert "human override; pipeline NO_POST" in ws.cell(2, 20).value
    assert ws.cell(2, 23).data_type == "s"
    assert ws.cell(2, 23).value.startswith("'")
    assert ws.cell(2, 34).data_type == "s"
    assert ws.cell(2, 34).value.startswith("'")


def test_invalid_legacy_override_does_not_break_results(completed_run):
    db.override_set(
        "override-run", 2, "REAL-CAMPAIGN", "1", "=legacy()", "old", "old"
    )
    response = completed_run.get("/runs/override-run/results")
    assert response.status_code == 200
    assert 'data-status="NO_POST"' in response.text
    assert "Invalid legacy override ignored" in response.text
