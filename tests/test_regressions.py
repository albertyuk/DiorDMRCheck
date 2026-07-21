"""Regression tests for adversarially-verified review findings."""
from __future__ import annotations

from datetime import date, datetime

import httpx
import pytest
from openpyxl import Workbook

from app.reconciler.pipeline import name_contains, name_ladder
from app.reconciler.parsers import _to_date, _to_int, parse_dmr, parse_plog
from app.reconciler.links import (_extract_note_fields, _normalize_url,
                          _note_id_from_url, _retry_after_seconds)
from tests import fixtures


# ---------------------------------------------------------- parser findings

def _build_dmr_header_row1(path: str) -> None:
    """DMR export with the header on row 1 — no metadata rows at all."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Streaming"
    ws.append(fixtures.DMR_HEADERS)
    for r in fixtures.dmr_rows():
        ws.append([
            "CN", "Beauty", r["blogger"], r["user"], "RED", r["pid"],
            r["likes"], r["favs"], r["date"], 150000, "Fashion", "Dior",
            # a data cell that LOOKS like a window string must not be parsed
            "Makeup", "#tag", "Show Post", r["likes"],
            1.0, "Promo From 2020/01/01 To 2020/02/01", 0, "", r["comments"],
        ])
    wb.save(path)


def test_dmr_header_on_row1_does_not_scan_data_for_window(tmp_path):
    p = tmp_path / "dmr_row1.xlsx"
    _build_dmr_header_row1(str(p))
    d = parse_dmr(str(p))
    assert d.header_row == 1
    # openpyxl treats max_row=0 as unset — the guard must prevent the whole
    # sheet being treated as metadata and a bogus 2020 window being extracted
    assert d.window_from is None and d.window_to is None
    assert d.metadata_text == ""
    assert any("window" in w for w in d.warnings)
    assert len(d.rows) == len(fixtures.dmr_rows())


def test_blank_campaign_cells_forward_fill(tmp_path):
    """Section-style sheets put the CAMPAIGN value only on the first row of
    each section; later rows inherit it."""
    wb = Workbook()
    ws = wb.active
    ws.title = "MASTER KOL LIST"
    ws.append(fixtures.PLOG_HEADERS)
    rows = [
        (1, "PLOG #009", "甲", datetime(2026, 3, 1)),
        (2, "", "乙", datetime(2026, 3, 2)),          # inherits #009
        (1, "PLOG #010", "丙", datetime(2026, 3, 3)),
        (2, "", "丁", datetime(2026, 3, 4)),          # inherits #010
    ]
    for no, camp, name, dt in rows:
        ws.append([no, "", camp, "", "", name, 1, dt, "", f"http://xhslink.com/o/{name}",
                   1, 1, 1, 1, 3, 1, 1, 1])
    p = tmp_path / "plog_sections.xlsx"
    wb.save(str(p))
    parsed = parse_plog(str(p))
    assert [r.campaign for r in parsed.rows] == \
        ["PLOG #009", "PLOG #009", "PLOG #010", "PLOG #010"]
    assert [r.key for r in parsed.rows] == \
        [("PLOG #009", "1"), ("PLOG #009", "2"),
         ("PLOG #010", "1"), ("PLOG #010", "2")]


def test_plog_hyperlink_post_link_cell(tmp_path):
    """POST LINK stored as a display-text hyperlink cell must yield the
    hyperlink target, not the display text."""
    wb = Workbook()
    ws = wb.active
    ws.title = "MASTER KOL LIST"
    ws.append(fixtures.PLOG_HEADERS)
    ws.append([1, "", "C1", "", "", "某博主", 1, datetime(2026, 4, 1), "",
               "链接", 1, 1, 1, 1, 3, 1, 1, 1])
    ws.cell(row=2, column=10).hyperlink = "http://xhslink.com/o/realtarget"
    p = tmp_path / "plog_link.xlsx"
    wb.save(str(p))
    parsed = parse_plog(str(p))
    assert parsed.rows[0].post_link == "http://xhslink.com/o/realtarget"


def test_excel_serial_date_and_comma_numbers():
    assert _to_date(45838.0) is not None          # serial float
    assert _to_date("2026年5月13日") is None       # unparseable → warning path
    assert _to_int("1,234") == 1234
    assert _to_int("１，２３４") == 1234           # full-width digits + comma
    assert _to_int("n/a") is None


def test_non_hex_postid_warns(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Streaming"
    ws.append(["User: x, Top Bloggers - From 2026/01/01 To 2026/07/20"])
    ws.append([""])
    ws.append(fixtures.DMR_HEADERS)
    ws.append(["CN", "B", "某人", "5f00000000000000000000ff", "RED", "N/A",
               1, 1, datetime(2026, 6, 1), 1, "", "", "", "", "", 1, 1, "", 0, "", 1])
    p = tmp_path / "dmr_badpid.xlsx"
    wb.save(str(p))
    d = parse_dmr(str(p))
    assert len(d.rows) == 1  # row kept — the deviation is itself a finding
    assert any("not a 24-char hex" in w for w in d.warnings)


def test_missing_username_column_warns(tmp_path):
    headers = [h for h in fixtures.DMR_HEADERS if h != "Username"]
    wb = Workbook()
    ws = wb.active
    ws.title = "Streaming"
    ws.append(["Top Bloggers - From 2026/01/01 To 2026/07/20"])
    ws.append([""])
    ws.append(headers)
    ws.append(["CN", "B", "某人", "RED", fixtures.N_EXTRA,
               1, 1, datetime(2026, 6, 1), 1, "", "", "", "", "", 1, 1, "", 0, ""])
    p = tmp_path / "dmr_nouser.xlsx"
    wb.save(str(p))
    d = parse_dmr(str(p))
    assert any("Username" in w for w in d.warnings)


def test_unparseable_post_date_warns(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "MASTER KOL LIST"
    ws.append(fixtures.PLOG_HEADERS)
    ws.append([1, "", "C1", "", "", "某博主", 1, "2026年5月13日", "",
               "http://xhslink.com/o/x", 1, 1, 1, 1, 3, 1, 1, 1])
    p = tmp_path / "plog_baddate.xlsx"
    wb.save(str(p))
    parsed = parse_plog(str(p))
    assert parsed.rows[0].post_date is None
    assert any("POST DATE" in w for w in parsed.warnings)


# --------------------------------------------------------- matcher findings

def test_ladder_rejects_short_dmr_ascii():
    # a 1-char DMR ASCII remainder must not fuzzy-match anything
    assert name_ladder("Poppy-chan", "C酱") == ""
    assert name_ladder("饼饼", "Bin") == ""  # pinyin bridge needs >=4 on DMR side


def test_all_emoji_name_never_flags_mislabel():
    assert name_contains("🥚🥚", "anything at all")
    assert name_contains("", "gungunnnnn")


# -------------------------------------------------------- resolver findings

def test_note_id_only_from_note_urls():
    uid = "5f00000000000000000000b1"
    nid = "6a1a0000000000000000a001"
    assert _note_id_from_url(f"https://www.xiaohongshu.com/user/profile/{uid}") is None
    assert _note_id_from_url(f"https://www.xiaohongshu.com/discovery/item/{nid}") == nid
    assert _note_id_from_url(f"https://www.xiaohongshu.com/explore/{nid}?xsec=1") == nid


def test_normalize_url_schemeless():
    assert _normalize_url("xhslink.com/o/abc") == "http://xhslink.com/o/abc"
    assert _normalize_url("http://xhslink.com/o/abc") == "http://xhslink.com/o/abc"
    assert _normalize_url("随便写的字") == "随便写的字"  # stays; rejected later


def test_extract_note_fields_no_blind_hex_fallback():
    """An author id in the payload must not be mistaken for the note id."""
    payload = {"code": 200, "data": {"user": {"user_id": "5f00000000000000000000b1",
                                              "nickname": "某人"},
                                     "desc": "no note id here"}}
    fields = _extract_note_fields(payload)
    assert fields["note_id"] == ""


def test_extract_note_fields_normal_payload():
    payload = {"code": 200, "data": {"note_id": "6A1A0000000000000000A001",
                                     "title": "标题",
                                     "user": {"user_id": "5f00000000000000000000b1",
                                              "nickname": "某人"},
                                     "interact_info": {"liked_count": "607",
                                                       "collected_count": "44",
                                                       "comment_count": "19"}}}
    fields = _extract_note_fields(payload)
    assert fields["note_id"] == "6a1a0000000000000000a001"
    assert fields["author_id"] == "5f00000000000000000000b1"
    assert fields["likes"] == 607 and fields["collects"] == 44


def test_retry_after_http_date_does_not_crash():
    resp = httpx.Response(429, headers={"retry-after": "Wed, 22 Jul 2026 07:28:00 GMT"})
    assert _retry_after_seconds(resp, 0) >= 0
    resp = httpx.Response(429, headers={"retry-after": "7"})
    assert _retry_after_seconds(resp, 0) == 7.0
    resp = httpx.Response(429)
    assert _retry_after_seconds(resp, 1) == 4.0


def test_direct_resolve_survives_malformed_location(monkeypatch):
    """A garbage Location header must degrade, never raise (a raise here would
    abort the whole run through pool.map)."""
    import app.reconciler.links as resolver_mod

    def handler(request: httpx.Request) -> httpx.Response:
        return httpx.Response(302, headers={"location": "http://[::1/broken"})

    real_client = httpx.Client

    def patched_client(**kwargs):
        kwargs["transport"] = httpx.MockTransport(handler)
        return real_client(**kwargs)

    monkeypatch.setattr(resolver_mod.httpx, "Client", patched_client)
    assert resolver_mod.direct_resolve("http://xhslink.com/o/x") is None


# ------------------------------------------------------ adjudicator findings

def test_parse_batch_accepts_bare_array_and_fences():
    from app.reconciler.adjudicator import _parse_batch
    item = ('{"row": "C|1|r2", "verdict": "UNSURE", "confidence": 0.5, '
            '"rationale_en": "e", "rationale_zh": "z"}')
    assert _parse_batch(f'{{"items": [{item}]}}') is not None
    assert _parse_batch(f"[{item}]") is not None          # bare array
    assert _parse_batch(f"```json\n[{item}]\n```") is not None
    assert _parse_batch("nonsense") is None


# ------------------------------------------------------------ eval findings

def test_eval_classify_review_with_link_mention():
    import eval as eval_mod
    assert eval_mod.classify("人工复核（链接已解析但无法获取作者ID）") == "人工复核"
    assert eval_mod.classify("Check链接错误（同名候选: x 6a…）") == "Check链接错误"
    assert eval_mod.classify("有 但是DMR博主名字标注错误") == "有 但是DMR博主名字标注错误"
    assert eval_mod.classify("") == "MATCH"
    assert eval_mod.classify("无帖子（超出DMR导出窗口，预期缺失）") == "无帖子"
