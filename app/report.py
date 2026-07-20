"""Exports.

1. Annotated .xlsx — the original PLOG workbook byte-identical in layout
   (columns A–R untouched, values and formats preserved), column S carrying
   the human vocabulary with no header (matching PLOG_DMR_CHECK_1), and
   richer evidence in columns T+ (which the reference leaves free).
2. JSON audit log of the full run.
"""
from __future__ import annotations

import json
from typing import Optional

from openpyxl import load_workbook
from openpyxl.styles import Font

from .matcher import Verdict

S_COL = 19  # column S
# The reference uses "已匹配/blank" for matched rows; this sentinel lets a human
# override force a blank S cell (asserting MATCH) rather than clearing the
# override.
OVERRIDE_MATCH_BLANK = "已匹配（清空S）"
EVIDENCE_HEADERS = [
    ("T", "STATUS"),
    ("U", "TIER"),
    ("V", "MATCHED DMR POSTID"),
    ("W", "MATCHED DMR BLOGGER"),
    ("X", "RESOLVED NOTE ID"),
    ("Y", "RESOLVED AUTHOR ID"),
    ("Z", "NAME METHOD"),
    ("AA", "DATE Δ (days)"),
    ("AB", "PLOG LIKE"),
    ("AC", "DMR LIKES (early snapshot — NOT comparable)"),
    ("AD", "CANDIDATES"),
    ("AE", "CLAUDE VERDICT"),
    ("AF", "CLAUDE RATIONALE"),
    ("AG", "NOTES"),
]
EVIDENCE_START_COL = 20  # column T


def _candidates_text(v: Verdict) -> str:
    parts = []
    for c in v.candidates[:5]:
        delta = f"Δ{c.date_delta_days:+d}d" if c.date_delta_days is not None else "Δ?"
        parts.append(f"{c.blogger} [{c.post_id}] {c.post_date or '?'} {delta} ({c.name_method})")
    return " ; ".join(parts)


def write_annotated_xlsx(plog_path: str, out_path: str, verdicts: list[Verdict],
                         header_row: int, sheet_name: Optional[str] = None,
                         overrides: Optional[dict] = None) -> None:
    """Copy the PLOG workbook and add column S (+ evidence T..).

    The workbook is loaded without data_only so formulas and formats in A–R
    survive untouched; we only ever write to columns >= S. The target sheet is
    the one parse_plog actually read (passed by name) — re-detection on the
    formula view could pick a different sheet.
    """
    wb = load_workbook(plog_path)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else None
    if ws is None:
        from .parsers import PLOG_REQUIRED, _find_header_row
        for candidate in wb.worksheets:
            if _find_header_row(candidate, PLOG_REQUIRED):
                ws = candidate
                break
    if ws is None:
        ws = wb.active

    overrides = overrides or {}
    bold = Font(bold=True)
    for col_idx, (_, title) in enumerate(EVIDENCE_HEADERS, start=EVIDENCE_START_COL):
        cell = ws.cell(row=header_row, column=col_idx, value=title)
        cell.font = bold
    # Column S intentionally has no header — the reference file leaves S1 blank.

    for v in verdicts:
        r = v.excel_row
        ov = overrides.get(r)
        if ov:
            s_text = "" if ov["status"] == OVERRIDE_MATCH_BLANK else ov["status"]
        else:
            s_text = v.column_s()
        status = f"{v.status}{' (override)' if ov else ''}"
        rationale = " / ".join(x for x in (v.llm_rationale_zh, v.llm_rationale_en) if x)
        llm = (f"{v.llm_verdict} ({v.llm_confidence:.0%})"
               if v.llm_verdict and v.llm_confidence is not None else v.llm_verdict)
        ws.cell(row=r, column=S_COL, value=s_text or None)
        values = [
            status,
            v.tier,
            v.matched_post_id or None,
            v.matched_blogger or None,
            v.resolved_note_id or None,
            v.resolved_author_id or None,
            v.name_method or None,
            v.date_delta_days,
            v.plog_like,
            v.dmr_likes_retweet,
            _candidates_text(v) or None,
            llm or None,
            rationale or None,
            " | ".join(v.notes + ([ov["note"]] if ov and ov.get("note") else [])) or None,
        ]
        for col_idx, value in enumerate(values, start=EVIDENCE_START_COL):
            ws.cell(row=r, column=col_idx, value=value)

    wb.save(out_path)


def build_audit_json(run: dict, verdicts: list[Verdict], counts: dict,
                     plog_meta: dict, dmr_meta: dict,
                     reverse_rows: list[dict],
                     overrides: Optional[dict] = None) -> str:
    overrides = overrides or {}
    doc = {
        "run_id": run.get("id"),
        "created_at": run.get("created_at"),
        "files": {"plog": run.get("plog_name"), "dmr": run.get("dmr_name")},
        "plog": plog_meta,
        "dmr": dmr_meta,
        "counts": counts,
        "tikhub_calls": run.get("tikhub_calls"),
        "llm_calls": run.get("llm_calls"),
        "summary": json.loads(run["summary_json"]) if run.get("summary_json") else None,
        "verdicts": [
            {**v.to_dict(),
             "override": overrides.get(v.excel_row)}
            for v in verdicts
        ],
        "reverse_audit": reverse_rows,
    }
    return json.dumps(doc, ensure_ascii=False, indent=2, default=str)
