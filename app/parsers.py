"""Schema-tolerant parsing of the PLOG tracker and the DMR export.

Header rows are located by fingerprint (PLOG: a row containing both NAME and
POST LINK; DMR: a row containing both Blogger and PostID), never by fixed
index. Header names are matched after NFKC + whitespace-collapse + casefold,
so the observed quirks (full-width paren in ``FAN BASE（K)``, double space in
``TTL  ENGAGEMENT``) and future cosmetic drift both map cleanly.
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from typing import Any, Optional
from urllib.parse import parse_qs, unquote, urlparse

from openpyxl import load_workbook

from .normalize import HEX24, header_key, nfkc

HEADER_SCAN_ROWS = 15  # how deep to look for the header fingerprint


# --------------------------------------------------------------------- utils

def _cell_str(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _to_date(v: Any) -> Optional[date]:
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, (int, float)):
        # Excel serial date (1900 system)
        try:
            return (datetime(1899, 12, 30) + timedelta(days=float(v))).date()
        except (OverflowError, ValueError):
            return None
    s = nfkc(str(v)).strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%m/%d/%Y", "%d/%m/%Y",
                "%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S", "%m/%d/%y", "%m-%d-%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    m = re.match(r"(\d{4})[./-](\d{1,2})[./-](\d{1,2})", s)
    if m:
        try:
            return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except ValueError:
            return None
    return None


def _to_datetime(v: Any) -> Optional[datetime]:
    if isinstance(v, datetime):
        return v
    d = _to_date(v)
    return datetime(d.year, d.month, d.day) if d else None


def _to_int(v: Any) -> Optional[int]:
    if v is None or v == "":
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return int(v)
    s = str(v).strip().replace(",", "")
    try:
        return int(float(s))
    except ValueError:
        return None


def _find_header_row(ws, required: set[str]) -> Optional[tuple[int, dict[str, int]]]:
    """Return (row_index, {header_key: column_index}) for the first row whose
    normalized cell values contain every key in *required*."""
    for row in ws.iter_rows(min_row=1, max_row=HEADER_SCAN_ROWS):
        keys: dict[str, int] = {}
        for cell in row:
            k = header_key(_cell_str(cell.value))
            if k and k not in keys:
                keys[k] = cell.column
        if required.issubset(keys.keys()):
            return row[0].row, keys
    return None


# ---------------------------------------------------------------------- PLOG

@dataclass
class PlogRow:
    campaign: str
    no: str
    name: str
    post_date: Optional[date]
    post_link: str
    like: Optional[int]
    collection: Optional[int]
    comment: Optional[int]
    impression: Optional[int]
    ttl_engagement: Optional[int]
    excel_row: int  # 1-based row in the source sheet (for annotated export)

    @property
    def key(self) -> tuple[str, str]:
        return (self.campaign, self.no)


@dataclass
class PlogParse:
    sheet: str
    header_row: int
    columns: dict[str, int]
    rows: list[PlogRow] = field(default_factory=list)
    campaigns: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)

    @property
    def date_range(self) -> tuple[Optional[date], Optional[date]]:
        ds = [r.post_date for r in self.rows if r.post_date]
        return (min(ds), max(ds)) if ds else (None, None)


PLOG_REQUIRED = {"name", "postlink"}


def parse_plog(path: str) -> PlogParse:
    wb = load_workbook(path, data_only=True)
    found = None
    for ws in wb.worksheets:
        hit = _find_header_row(ws, PLOG_REQUIRED)
        if hit:
            found = (ws, *hit)
            break
    if not found:
        raise ValueError(
            "PLOG parse failed: no sheet has a header row containing both "
            "'NAME' and 'POST LINK' within the first "
            f"{HEADER_SCAN_ROWS} rows."
        )
    ws, header_row, cols = found
    result = PlogParse(sheet=ws.title, header_row=header_row, columns=cols)

    def col(key: str) -> Optional[int]:
        return cols.get(key)

    c_no, c_campaign, c_name = col("no"), col("campaign"), col("name")
    c_date, c_link = col("postdate"), col("postlink")
    c_like, c_coll, c_comm = col("like"), col("collection"), col("comment")
    c_impr, c_ttl = col("impression"), col("ttlengagement")

    current_campaign = ""
    seen_keys: set[tuple[str, str]] = set()
    for row in ws.iter_rows(min_row=header_row + 1):
        r = row[0].row
        get = lambda c: ws.cell(row=r, column=c).value if c else None
        name = _cell_str(get(c_name))
        link_cell = ws.cell(row=r, column=c_link) if c_link else None
        link = ""
        if link_cell is not None:
            if link_cell.hyperlink and link_cell.hyperlink.target:
                link = str(link_cell.hyperlink.target).strip()
            else:
                link = _cell_str(link_cell.value)
        if not name and not link:
            continue  # blank separator / spacer rows

        campaign = _cell_str(get(c_campaign))
        if campaign:
            current_campaign = campaign
        no = _cell_str(get(c_no))
        prow = PlogRow(
            campaign=current_campaign,
            no=no,
            name=name,
            post_date=_to_date(get(c_date)),
            post_link=link,
            like=_to_int(get(c_like)),
            collection=_to_int(get(c_coll)),
            comment=_to_int(get(c_comm)),
            impression=_to_int(get(c_impr)),
            ttl_engagement=_to_int(get(c_ttl)),
            excel_row=r,
        )
        if prow.key in seen_keys:
            result.warnings.append(
                f"Duplicate row identity (CAMPAIGN={prow.campaign!r}, NO={prow.no!r}) "
                f"at sheet row {r} — annotations for duplicates land on the first occurrence."
            )
        seen_keys.add(prow.key)
        if prow.campaign and prow.campaign not in result.campaigns:
            result.campaigns.append(prow.campaign)
        result.rows.append(prow)

    if not result.rows:
        result.warnings.append("PLOG sheet parsed but contained no data rows.")
    return result


# ----------------------------------------------------------------------- DMR

@dataclass
class DmrRow:
    blogger: str
    username: str          # XHS author/user id — treated as an opaque string
    post_id: str           # normalized (lowercase) join key
    post_id_raw: str
    post_date: Optional[datetime]
    likes_retweet: Optional[int]
    share_favorites: Optional[int]
    engagement: Optional[int]
    comments: Optional[int]
    link_target: str       # hyperlink target of the Link cell, if any
    link_embedded_post_id: str
    excel_row: int


@dataclass
class DmrParse:
    sheet: str
    header_row: int
    columns: dict[str, int]
    rows: list[DmrRow] = field(default_factory=list)
    window_from: Optional[date] = None
    window_to: Optional[date] = None
    metadata_text: str = ""
    warnings: list[str] = field(default_factory=list)


DMR_REQUIRED = {"blogger", "postid"}

_WINDOW_RE = re.compile(r"From\s+(.+?)\s+To\s+(.+?)(?:\s*$|,)", re.IGNORECASE)


def _extract_link_target(cell) -> str:
    if cell is None:
        return ""
    if cell.hyperlink and cell.hyperlink.target:
        return str(cell.hyperlink.target).strip()
    v = _cell_str(cell.value)
    return v if v.lower().startswith("http") else ""


def _embedded_post_id(link_target: str) -> str:
    """The DMR Link column carries https://www.dmr.st/redi.html?url=<urlencoded
    xiaohongshu.com/discovery/item/{PostID}> — extract that PostID."""
    if not link_target:
        return ""
    try:
        parsed = urlparse(link_target)
        qs = parse_qs(parsed.query)
        inner = unquote(qs.get("url", [""])[0]) or link_target
    except ValueError:
        inner = link_target
    m = HEX24.search(inner)
    return m.group(1).lower() if m else ""


def parse_dmr(path: str) -> DmrParse:
    wb = load_workbook(path, data_only=True)
    found = None
    for ws in wb.worksheets:
        hit = _find_header_row(ws, DMR_REQUIRED)
        if hit:
            found = (ws, *hit)
            break
    if not found:
        raise ValueError(
            "DMR parse failed: no sheet has a header row containing both "
            f"'Blogger' and 'PostID' within the first {HEADER_SCAN_ROWS} rows."
        )
    ws, header_row, cols = found
    result = DmrParse(sheet=ws.title, header_row=header_row, columns=cols)

    # Metadata rows above the header carry the export window
    # ("... Top Bloggers - From <date> To <date>").
    meta_parts = []
    for row in ws.iter_rows(min_row=1, max_row=header_row - 1):
        for cell in row:
            s = _cell_str(cell.value)
            if s:
                meta_parts.append(s)
    result.metadata_text = " | ".join(meta_parts)
    m = _WINDOW_RE.search(result.metadata_text)
    if m:
        result.window_from = _to_date(m.group(1))
        result.window_to = _to_date(m.group(2))
    if not (result.window_from and result.window_to):
        result.warnings.append(
            "Could not parse the DMR export date window from the metadata rows; "
            "out-of-window checks are disabled for this run."
        )

    def col(key: str) -> Optional[int]:
        return cols.get(key)

    c_blogger, c_user, c_pid = col("blogger"), col("username"), col("postid")
    c_pdate = col("postdate")
    c_likes, c_shares = col("likes_retweet"), col("share_favorites")
    c_eng, c_comm, c_link = col("engagement"), col("comments"), col("link")

    for row in ws.iter_rows(min_row=header_row + 1):
        r = row[0].row
        get = lambda c: ws.cell(row=r, column=c).value if c else None
        blogger = _cell_str(get(c_blogger))
        pid_raw = _cell_str(get(c_pid))
        if not blogger and not pid_raw:
            continue
        link_cell = ws.cell(row=r, column=c_link) if c_link else None
        link_target = _extract_link_target(link_cell)
        embedded = _embedded_post_id(link_target)
        pid = pid_raw.lower() if re.fullmatch(r"[0-9a-fA-F]{24}", pid_raw) else pid_raw.lower()
        drow = DmrRow(
            blogger=blogger,
            username=_cell_str(get(c_user)),
            post_id=pid,
            post_id_raw=pid_raw,
            post_date=_to_datetime(get(c_pdate)),
            likes_retweet=_to_int(get(c_likes)),
            share_favorites=_to_int(get(c_shares)),
            engagement=_to_int(get(c_eng)),
            comments=_to_int(get(c_comm)),
            link_target=link_target,
            link_embedded_post_id=embedded,
            excel_row=r,
        )
        if embedded and pid and embedded != pid:
            result.warnings.append(
                f"DMR row {r}: Link hyperlink embeds PostID {embedded} but the "
                f"PostID column says {pid} — using the PostID column for the join."
            )
        result.rows.append(drow)

    if not result.rows:
        result.warnings.append("DMR sheet parsed but contained no data rows.")
    return result
