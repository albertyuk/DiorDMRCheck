"""Schema-tolerant parsing of the PLOG tracker and the DMR export.

Header rows are located by fingerprint (PLOG: a row containing both NAME and
POST LINK; DMR: a row containing both Blogger and PostID), never by fixed
index. Header names are matched after NFKC + whitespace-collapse + casefold,
so the observed quirks (full-width paren in ``FAN BASE（K)``, double space in
``TTL  ENGAGEMENT``) and future cosmetic drift both map cleanly.
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import date, datetime
from typing import Optional
from urllib.parse import parse_qs, unquote, urlparse

from openpyxl import load_workbook

from .. import config
from ..core.xlsx import (HEADER_SCAN_ROWS, cell_str, extract_link_target,
                         find_header_row, to_date, to_datetime, to_int)
from .domain import HEX24, is_hex24


# ---------------------------------------------------------------------- PLOG

@dataclass
class PlogRow:
    campaign: str
    no: str
    name: str
    post_date: Optional[date]
    post_link: str
    like: Optional[int]
    collection: Optional[int]
    comment: Optional[int]
    impression: Optional[int]
    ttl_engagement: Optional[int]
    excel_row: int  # 1-based row in the source sheet (for annotated export)

    @property
    def key(self) -> tuple[str, str]:
        return (self.campaign, self.no)


@dataclass
class PlogParse:
    sheet: str
    header_row: int
    columns: dict[str, int]
    rows: list[PlogRow] = field(default_factory=list)
    campaigns: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)

    @property
    def date_range(self) -> tuple[Optional[date], Optional[date]]:
        ds = [r.post_date for r in self.rows if r.post_date]
        return (min(ds), max(ds)) if ds else (None, None)


PLOG_REQUIRED = {"name", "postlink"}


def _populated_identity_rows(ws, start_row: int,
                             columns: tuple[Optional[int], ...]) -> list[int]:
    """Return only rows with a populated identity cell.

    Iterating ``1..ws.max_row`` is attacker-controlled: a single styled cell
    at Excel's final row creates a million-iteration scan while remaining far
    below the archive's populated-cell limit. Normal-mode openpyxl has already
    materialized real cells in ``_cells``, so enumerate those sparse entries.
    """
    wanted = {column for column in columns if column is not None}
    rows = set()
    for (row, column), cell in ws._cells.items():
        if row < start_row or column not in wanted:
            continue
        if cell.value not in (None, "") or getattr(cell, "hyperlink", None):
            rows.add(row)
    return sorted(rows)


def _probe_schema(path: str, required: set[str], label: str) -> None:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if any(find_header_row(ws, required) for ws in wb.worksheets):
            return
        names = " and ".join(repr(name) for name in sorted(required))
        raise ValueError(
            f"{label} parse failed: no sheet has a header row containing "
            f"{names} within the first {HEADER_SCAN_ROWS} rows."
        )
    finally:
        wb.close()


def probe_plog_schema(path: str) -> None:
    _probe_schema(path, PLOG_REQUIRED, "PLOG")


def parse_plog(path: str) -> PlogParse:
    wb = load_workbook(path, data_only=True)
    try:
        return _parse_plog_workbook(wb)
    finally:
        wb.close()


def _parse_plog_workbook(wb) -> PlogParse:
    found = None
    for ws in wb.worksheets:
        hit = find_header_row(ws, PLOG_REQUIRED)
        if hit:
            found = (ws, *hit)
            break
    if not found:
        raise ValueError(
            "PLOG parse failed: no sheet has a header row containing both "
            "'NAME' and 'POST LINK' within the first "
            f"{HEADER_SCAN_ROWS} rows."
        )
    ws, header_row, cols = found
    result = PlogParse(sheet=ws.title, header_row=header_row, columns=cols)

    def col(key: str) -> Optional[int]:
        return cols.get(key)

    c_no, c_campaign, c_name = col("no"), col("campaign"), col("name")
    c_date, c_link = col("postdate"), col("postlink")
    c_like, c_coll, c_comm = col("like"), col("collection"), col("comment")
    c_impr, c_ttl = col("impression"), col("ttlengagement")

    current_campaign = ""
    seen_keys: set[tuple[str, str]] = set()
    bad_dates = 0
    for r in _populated_identity_rows(
            ws, header_row + 1, (c_name, c_link)):

        def get(column):
            return ws.cell(row=r, column=column).value if column else None

        name = cell_str(get(c_name))
        link_cell = ws.cell(row=r, column=c_link) if c_link else None
        link = ""
        if link_cell is not None:
            if link_cell.hyperlink and link_cell.hyperlink.target:
                link = str(link_cell.hyperlink.target).strip()
            else:
                link = cell_str(link_cell.value)
        if not name and not link:
            # Workbook archive validation bounds the declared cell area, so it
            # is safe to scan through spacer rows. Stopping after an arbitrary
            # blank run silently discarded legitimate later campaign blocks.
            continue

        raw_date = get(c_date)
        parsed_date = to_date(raw_date)
        if parsed_date is None and cell_str(raw_date):
            bad_dates += 1
            if bad_dates <= 5:
                result.warnings.append(
                    f"PLOG row {r}: POST DATE {cell_str(raw_date)!r} could not "
                    "be parsed — date-based checks are skipped for this row."
                )

        campaign = cell_str(get(c_campaign))
        if campaign:
            current_campaign = campaign
        no = cell_str(get(c_no))
        prow = PlogRow(
            campaign=current_campaign,
            no=no,
            name=name,
            post_date=parsed_date,
            post_link=link,
            like=to_int(get(c_like)),
            collection=to_int(get(c_coll)),
            comment=to_int(get(c_comm)),
            impression=to_int(get(c_impr)),
            ttl_engagement=to_int(get(c_ttl)),
            excel_row=r,
        )
        if prow.key in seen_keys:
            result.warnings.append(
                f"Duplicate row identity (CAMPAIGN={prow.campaign!r}, NO={prow.no!r}) "
                f"at sheet row {r} — each row is still annotated individually "
                "(rows are tracked by sheet row), but check the source data."
            )
        seen_keys.add(prow.key)
        if prow.campaign and prow.campaign not in result.campaigns:
            result.campaigns.append(prow.campaign)
        result.rows.append(prow)
        if len(result.rows) > config.MAX_PLOG_ROWS:
            raise ValueError(
                f"PLOG contains more than {config.MAX_PLOG_ROWS:,} data rows; "
                "split the workbook into smaller runs."
            )

    if bad_dates > 5:
        result.warnings.append(
            f"PLOG: {bad_dates} rows in total had unparseable POST DATE values."
        )
    if not result.rows:
        result.warnings.append("PLOG sheet parsed but contained no data rows.")
    return result


# ----------------------------------------------------------------------- DMR

@dataclass
class DmrRow:
    blogger: str
    username: str          # XHS author/user id — treated as an opaque string
    post_id: str           # normalized (lowercase) join key
    post_id_raw: str
    post_date: Optional[datetime]
    likes_retweet: Optional[int]
    share_favorites: Optional[int]
    engagement: Optional[int]
    comments: Optional[int]
    link_target: str       # hyperlink target of the Link cell, if any
    link_embedded_post_id: str
    excel_row: int


@dataclass
class DmrParse:
    sheet: str
    header_row: int
    columns: dict[str, int]
    rows: list[DmrRow] = field(default_factory=list)
    window_from: Optional[date] = None
    window_to: Optional[date] = None
    metadata_text: str = ""
    warnings: list[str] = field(default_factory=list)


DMR_REQUIRED = {"blogger", "postid"}

# The To-date must stop at a comma or the ' | ' joiner we insert between
# metadata cells — otherwise trailing metadata leaks into the capture.
_WINDOW_RE = re.compile(r"From\s+([^,|]+?)\s+To\s+([^,|]+)", re.IGNORECASE)


def _embedded_post_id(link_target: str) -> str:
    """The DMR Link column carries https://www.dmr.st/redi.html?url=<urlencoded
    xiaohongshu.com/discovery/item/{PostID}> — extract that PostID."""
    if not link_target:
        return ""
    try:
        parsed = urlparse(link_target)
        qs = parse_qs(parsed.query)
        inner = unquote(qs.get("url", [""])[0]) or link_target
    except ValueError:
        inner = link_target
    m = HEX24.search(inner)
    return m.group(1).lower() if m else ""


def probe_dmr_schema(path: str) -> None:
    _probe_schema(path, DMR_REQUIRED, "DMR")


def parse_dmr(path: str) -> DmrParse:
    wb = load_workbook(path, data_only=True)
    try:
        return _parse_dmr_workbook(wb)
    finally:
        wb.close()


def _parse_dmr_workbook(wb) -> DmrParse:
    found = None
    for ws in wb.worksheets:
        hit = find_header_row(ws, DMR_REQUIRED)
        if hit:
            found = (ws, *hit)
            break
    if not found:
        raise ValueError(
            "DMR parse failed: no sheet has a header row containing both "
            f"'Blogger' and 'PostID' within the first {HEADER_SCAN_ROWS} rows."
        )
    ws, header_row, cols = found
    result = DmrParse(sheet=ws.title, header_row=header_row, columns=cols)

    # Metadata rows above the header carry the export window
    # ("... Top Bloggers - From <date> To <date>"). Guard header_row == 1:
    # openpyxl treats max_row=0 as "unset" and would iterate the whole sheet.
    meta_parts = []
    if header_row > 1:
        for row in ws.iter_rows(min_row=1, max_row=header_row - 1):
            for cell in row:
                s = cell_str(cell.value)
                if s:
                    meta_parts.append(s)
    result.metadata_text = " | ".join(meta_parts)
    m = _WINDOW_RE.search(result.metadata_text)
    if m:
        # DMR exports use day/month/year in their human-readable metadata.
        # Passing the policy explicitly avoids interpreting 05/06 as May 6.
        result.window_from = to_date(m.group(1), date_order="day-first")
        result.window_to = to_date(m.group(2), date_order="day-first")
    if not (result.window_from and result.window_to):
        result.warnings.append(
            "Could not parse the DMR export date window from the metadata rows; "
            "out-of-window checks are disabled for this run."
        )

    def col(key: str) -> Optional[int]:
        return cols.get(key)

    c_blogger, c_user, c_pid = col("blogger"), col("username"), col("postid")
    c_pdate = col("postdate")
    c_likes, c_shares = col("likes_retweet"), col("share_favorites")
    c_eng, c_comm, c_link = col("engagement"), col("comments"), col("link")

    non_hex_pids = 0
    seen_post_ids: dict[str, tuple[int, tuple]] = {}
    duplicate_post_ids = 0
    missing_usernames = 0
    for r in _populated_identity_rows(
            ws, header_row + 1, (c_blogger, c_pid)):

        def get(column):
            return ws.cell(row=r, column=column).value if column else None

        blogger = cell_str(get(c_blogger))
        pid_raw = cell_str(get(c_pid))
        if not blogger and not pid_raw:
            continue
        link_cell = ws.cell(row=r, column=c_link) if c_link else None
        link_target = extract_link_target(link_cell)
        embedded = _embedded_post_id(link_target)
        pid = pid_raw.lower()
        if pid_raw and not is_hex24(pid_raw):
            # Row is kept (a deviating PostID is itself a finding), but it can
            # never join — tell the operator instead of failing silently.
            non_hex_pids += 1
            if non_hex_pids <= 5:
                result.warnings.append(
                    f"DMR row {r}: PostID {pid_raw!r} is not a 24-char hex note "
                    "id — this row cannot join against resolved links."
                )
        username = cell_str(get(c_user)).strip().lower()
        if c_user is not None and not username:
            missing_usernames += 1
        drow = DmrRow(
            blogger=blogger,
            username=username,
            post_id=pid,
            post_id_raw=pid_raw,
            post_date=to_datetime(get(c_pdate)),
            likes_retweet=to_int(get(c_likes)),
            share_favorites=to_int(get(c_shares)),
            engagement=to_int(get(c_eng)),
            comments=to_int(get(c_comm)),
            link_target=link_target,
            link_embedded_post_id=embedded,
            excel_row=r,
        )
        if pid:
            signature = (username, blogger, drow.post_date, link_target)
            previous = seen_post_ids.get(pid)
            if previous is None:
                seen_post_ids[pid] = (r, signature)
            else:
                duplicate_post_ids += 1
                first_row, first_signature = previous
                qualifier = "conflicting " if signature != first_signature else ""
                if duplicate_post_ids <= 5:
                    result.warnings.append(
                        f"DMR rows {first_row} and {r} contain {qualifier}duplicate "
                        f"PostID {pid}; matching that note requires manual review."
                    )
        if embedded and pid and embedded != pid:
            result.warnings.append(
                f"DMR row {r}: Link hyperlink embeds PostID {embedded} but the "
                f"PostID column says {pid} — using the PostID column for the join."
            )
        result.rows.append(drow)
        if len(result.rows) > config.MAX_DMR_ROWS:
            raise ValueError(
                f"DMR contains more than {config.MAX_DMR_ROWS:,} data rows; "
                "split the export into smaller runs."
            )

    if non_hex_pids > 5:
        result.warnings.append(
            f"DMR: {non_hex_pids} rows in total had non-hex PostID values."
        )
    if duplicate_post_ids > 5:
        result.warnings.append(
            f"DMR: {duplicate_post_ids} duplicate PostID occurrences were found "
            "in total; affected notes require manual review."
        )
    if c_user is None:
        result.warnings.append(
            "DMR has no 'Username' column — blogger-presence checks (无博主 vs "
            "无帖子) cannot be decided deterministically for this file."
        )
    elif result.rows and not any(r.username for r in result.rows):
        result.warnings.append(
            "DMR 'Username' column is entirely empty — blogger-presence checks "
            "(无博主 vs 无帖子) cannot be decided deterministically for this file."
        )
    elif missing_usernames:
        result.warnings.append(
            f"DMR Username is blank on {missing_usernames} data row(s). An "
            "author absent from the partial Username index cannot be classified "
            "as 无博主 deterministically and will require review."
        )
    if not result.rows:
        result.warnings.append("DMR sheet parsed but contained no data rows.")
    return result
