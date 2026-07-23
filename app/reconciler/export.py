"""Exports.

1. Annotated .xlsx — the original PLOG workbook byte-identical in layout
   (columns A–R untouched, values and formats preserved), column S carrying
   the human vocabulary with no header (matching PLOG_DMR_CHECK_1), and
   richer evidence in columns T+ (which the reference leaves free).
2. JSON audit log of the full run.
"""
from __future__ import annotations

import json
import re
from dataclasses import fields as dataclass_fields
from typing import Optional

from openpyxl import load_workbook
from openpyxl.styles import Font

from ..core.xlsx import find_header_row
from .parsers import PLOG_REQUIRED
from .domain import (ENGAGEMENT_CAVEAT, Candidate, Verdict,
                     effective_verdict_dict)
from .domain import OVERRIDE_MATCH_BLANK as OVERRIDE_MATCH_BLANK
from .pipeline import status_counts, summary_buckets

S_COL = 19  # column S
EVIDENCE_HEADERS = [
    ("T", "STATUS"),
    ("U", "TIER"),
    ("V", "MATCHED DMR POSTID"),
    ("W", "MATCHED DMR BLOGGER"),
    ("X", "RESOLVED NOTE ID"),
    ("Y", "RESOLVED AUTHOR ID"),
    ("Z", "NAME METHOD"),
    ("AA", "DATE Δ (days)"),
    ("AB", "PLOG LIKE"),
    ("AC", "DMR LIKES (early snapshot — NOT comparable)"),
    ("AD", "CANDIDATES"),
    ("AE", "CLAUDE VERDICT"),
    ("AF", "CLAUDE RATIONALE"),
    ("AG", "PERIMETER"),
    ("AH", "NOTES"),
]
EVIDENCE_START_COL = 20  # column T
EXCEL_MAX_COLUMN = 16_384
EVIDENCE_SHEET_BASE = "_DMR_EVIDENCE"
OOXML_ILLEGAL_CONTROL_RE = re.compile(r"[\x00-\x08\x0b-\x0c\x0e-\x1f]")


def _perimeter_text(v: Verdict) -> str:
    parts = []
    if v.perimeter_method:
        entry = f"{v.perimeter_method}: {v.perimeter_name or v.perimeter_namebis}"
        if v.perimeter_namebis and v.perimeter_name:
            entry += f" / {v.perimeter_namebis}"
        if v.perimeter_dmrid:
            entry += f" · DMRID {v.perimeter_dmrid}"
        if v.perimeter_redbook_id:
            entry += f" · REDBOOK {v.perimeter_redbook_id}"
        if v.perimeter_followers is not None:
            entry += f" · {v.perimeter_followers:,} followers"
        parts.append(entry)
    if v.perimeter_note:
        parts.append(v.perimeter_note)
    if v.perimeter_candidates:
        parts.append("candidates: " + " ; ".join(v.perimeter_candidates))
    if parts and v.perimeter_extraction_date:
        parts.append(f"extracted {v.perimeter_extraction_date}")
    return " | ".join(parts)


def _candidates_text(v: Verdict) -> str:
    parts = []
    for c in v.candidates[:5]:
        delta = f"Δ{c.date_delta_days:+d}d" if c.date_delta_days is not None else "Δ?"
        parts.append(f"{c.blogger} [{c.post_id}] {c.post_date or '?'} {delta} ({c.name_method})")
    return " ; ".join(parts)


def _safe_excel_value(value):
    """Prevent text copied from uploads, users, or models becoming a formula."""
    if isinstance(value, str):
        # XML 1.0 forbids these C0 controls. openpyxl raises
        # IllegalCharacterError if any reach a cell, so a single malicious or
        # accidentally pasted byte must not poison the whole workbook export.
        value = OOXML_ILLEGAL_CONTROL_RE.sub("", value)
        # Strip controls before checking the prefix; otherwise "\x01=SUM(...)"
        # would evade formula neutralization.
        if value.startswith(("=", "+", "-", "@", "\t", "\r")):
            return "'" + value
    return value


def _unique_sheet_title(wb, base: str) -> str:
    existing = {title.casefold() for title in wb.sheetnames}
    title = base
    suffix = 2
    while title.casefold() in existing:
        title = f"{base}_{suffix}"
        suffix += 1
    return title


def _evidence_start_column(ws, width: int) -> Optional[int]:
    """Find a bounded empty column block without expanding sparse worksheets."""
    loaded_cells = getattr(ws, "_cells", {})
    occupied = {
        cell.column
        for cell in loaded_cells.values()
        if cell.value not in (None, "")
    }
    # Blank merged cells are still unavailable write targets: only the
    # top-left cell of a merged range is writable. Treat every column touched
    # by a merge as occupied so neither the evidence header nor any verdict
    # row can land on a read-only MergedCell.
    for merged in ws.merged_cells.ranges:
        occupied.update(range(merged.min_col, merged.max_col + 1))
    last_start = EXCEL_MAX_COLUMN - width + 1
    for start in range(EVIDENCE_START_COL, last_start + 1):
        if occupied.isdisjoint(range(start, start + width)):
            return start
    return None


def _create_evidence_sheet(wb):
    """Create a compact hidden audit sheet when the source has no safe block."""
    ws = wb.create_sheet(_unique_sheet_title(wb, EVIDENCE_SHEET_BASE))
    ws.sheet_state = "hidden"
    headers = [
        "SOURCE SHEET", "SOURCE EXCEL ROW", "CAMPAIGN", "NO", "NAME",
        *(title for _, title in EVIDENCE_HEADERS),
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    return ws


def _write_perimeter_provenance(wb, metadata: Optional[dict],
                                warning: Optional[str],
                                source_s: Optional[list[dict]] = None,
                                evidence_sheet: Optional[str] = None) -> None:
    if not metadata and not warning and not source_s and not evidence_sheet:
        return
    ws = wb.create_sheet(_unique_sheet_title(wb, "_DMR_AUDIT_META"))
    ws.sheet_state = "hidden"
    ws.append(["FIELD", "VALUE"])
    for key, value in (metadata or {}).items():
        if isinstance(value, (dict, list)):
            value = json.dumps(value, ensure_ascii=False, sort_keys=True)
        ws.append([_safe_excel_value(str(key)), _safe_excel_value(value)])
    if warning:
        ws.append(["warning", _safe_excel_value(warning)])
    if evidence_sheet:
        ws.append(["evidence_sheet", evidence_sheet])
    for item in source_s or []:
        ws.append([
            "source_column_s",
            json.dumps(item, ensure_ascii=False, sort_keys=True, default=str),
        ])


def write_annotated_xlsx(plog_path: str, out_path: str, verdicts: list[Verdict],
                         header_row: int, sheet_name: Optional[str] = None,
                         overrides: Optional[dict] = None,
                         perimeter_meta: Optional[dict] = None,
                         perimeter_warning: Optional[str] = None) -> None:
    """Copy the PLOG workbook and add column S (+ evidence T..).

    The workbook is loaded without data_only so formulas and formats in A–R
    survive untouched; we only ever write to columns >= S. The target sheet is
    the one parse_plog actually read (passed by name) — re-detection on the
    formula view could pick a different sheet.

    Pre-existing content is never overwritten: an S cell that already holds a
    value in the source keeps its value, type, and formatting (an explicit UI
    override still wins). Newly written external text is formula-neutralized
    and stripped of OOXML-illegal controls. Evidence shifts past populated or
    merged source columns; if no bounded block fits before XFD, it moves to a
    dedicated hidden audit sheet.
    """
    wb = load_workbook(plog_path)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else None
    if ws is None:
        for candidate in wb.worksheets:
            if find_header_row(candidate, PLOG_REQUIRED):
                ws = candidate
                break
    if ws is None:
        ws = wb.active

    overrides = overrides or {}
    width = len(EVIDENCE_HEADERS)
    ev_start = _evidence_start_column(ws, width)
    evidence_ws = None
    if ev_start is None:
        evidence_ws = _create_evidence_sheet(wb)

    bold = Font(bold=True)
    if ev_start is not None:
        for col_idx, (_, title) in enumerate(EVIDENCE_HEADERS, start=ev_start):
            cell = ws.cell(row=header_row, column=col_idx, value=title)
            cell.font = bold
    # Column S intentionally has no header — the reference file leaves S1 blank.

    source_s_provenance: list[dict] = []
    for v in verdicts:
        r = v.excel_row
        ov = overrides.get(r)
        effective = effective_verdict_dict(v.to_dict(), ov)
        valid_override = bool(ov) and not effective.get("override_invalid")
        s_text = effective["column_s"]
        status = effective["status"]
        if valid_override:
            status = f"{status} (human override; pipeline {v.status})"
        source_cell = ws.cell(row=r, column=S_COL)
        existing_s = source_cell.value
        preserved = None
        if existing_s not in (None, ""):
            disposition = (
                "replaced_by_human_override" if valid_override else "preserved"
            )
            source_s_provenance.append({
                "excel_row": r,
                "value": str(existing_s),
                "cell_data_type": source_cell.data_type,
                "disposition": disposition,
                "pipeline_column_s": v.column_s(),
            })
            if not valid_override:
                preserved = str(existing_s)
                s_text = preserved
                status += " (S kept from source)"
        rationale = " / ".join(x for x in (v.llm_rationale_zh, v.llm_rationale_en) if x)
        llm = (f"{v.llm_verdict} ({v.llm_confidence:.0%})"
               if v.llm_verdict and v.llm_confidence is not None else v.llm_verdict)
        notes = list(v.notes)
        if preserved is not None:
            pipeline_s = v.column_s()
            if preserved.strip() != (pipeline_s or "").strip():
                notes.append(
                    f"S already contained {preserved!r} — kept; pipeline "
                    f"verdict was {pipeline_s or '(blank=matched)'}")
        # Preserving an existing S value means leaving the source cell alone.
        # Reassigning ``str(existing_s)`` changes numeric, date, boolean, and
        # formula cells into strings even though their displayed text may look
        # similar. Only the pipeline or a valid human override writes S.
        if existing_s in (None, "") or valid_override:
            ws.cell(row=r, column=S_COL,
                    value=_safe_excel_value(s_text) if s_text else None)
        values = [
            status,
            v.tier,
            v.matched_post_id or None,
            v.matched_blogger or None,
            v.resolved_note_id or None,
            v.resolved_author_id or None,
            v.name_method or None,
            v.date_delta_days,
            v.plog_like,
            v.dmr_likes_retweet,
            _candidates_text(v) or None,
            llm or None,
            rationale or None,
            _perimeter_text(v) or None,
            " | ".join(notes + ([ov["note"]] if ov and ov.get("note") else [])) or None,
        ]
        safe_values = [_safe_excel_value(value) for value in values]
        if ev_start is not None:
            for col_idx, value in enumerate(safe_values, start=ev_start):
                ws.cell(row=r, column=col_idx, value=value)
        else:
            evidence_ws.append([
                _safe_excel_value(ws.title),
                r,
                _safe_excel_value(v.campaign),
                _safe_excel_value(v.no),
                _safe_excel_value(v.name),
                *safe_values,
            ])

    _write_perimeter_provenance(
        wb,
        perimeter_meta,
        perimeter_warning,
        source_s=source_s_provenance,
        evidence_sheet=evidence_ws.title if evidence_ws is not None else None,
    )
    wb.save(out_path)


def build_audit_json(run: dict, verdicts: list[Verdict], counts: dict,
                     plog_meta: dict, dmr_meta: dict,
                     reverse_rows: list[dict],
                     overrides: Optional[dict] = None,
                     perimeter_meta: Optional[dict] = None,
                     perimeter_warning: Optional[str] = None) -> str:
    overrides = overrides or {}
    effective_verdicts = [
        effective_verdict_dict(v.to_dict(), overrides.get(v.excel_row))
        for v in verdicts
    ]
    effective_counts = status_counts(effective_verdicts)
    doc = {
        "run_id": run.get("id"),
        "created_at": run.get("created_at"),
        "files": {"plog": run.get("plog_name"), "dmr": run.get("dmr_name")},
        "plog": plog_meta,
        "dmr": dmr_meta,
        "counts": effective_counts,
        "pipeline_counts": counts,
        "buckets": summary_buckets(effective_verdicts),
        "engagement_caveat": ENGAGEMENT_CAVEAT,
        "tikhub_calls": run.get("tikhub_calls"),
        "llm_calls": run.get("llm_calls"),
        "summary": json.loads(run["summary_json"]) if run.get("summary_json") else None,
        "summary_basis": "pipeline_before_human_overrides",
        "verdicts": effective_verdicts,
        "reverse_audit": reverse_rows,
        "perimeter": perimeter_meta,
        "perimeter_warning": perimeter_warning,
    }
    return json.dumps(doc, ensure_ascii=False, indent=2, default=str)


_VERDICT_FIELDS = frozenset(f.name for f in dataclass_fields(Verdict))
_CANDIDATE_FIELDS = frozenset(f.name for f in dataclass_fields(Candidate))


def load_verdicts(run: dict) -> list[Verdict]:
    """Rehydrate Verdict objects from a finished run's result_json.

    Unknown keys are dropped, not fatal: stored documents carry derived
    fields (column_s) and may predate the current schema (older runs stored
    a per-row engagement_caveat) — a rendering-side change must never make
    historical runs unexportable."""
    result = json.loads(run.get("result_json") or "{}")
    out = []
    for d in result.get("verdicts", []):
        d = dict(d)
        cands = [Candidate(**{k: v for k, v in c.items()
                              if k in _CANDIDATE_FIELDS})
                 for c in d.pop("candidates", [])]
        v = Verdict(**{k: v for k, v in d.items() if k in _VERDICT_FIELDS})
        v.candidates = cands
        out.append(v)
    return out
