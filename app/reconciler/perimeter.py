"""LVMH social perimeter (Micro + Macro) — offline cross-check for
NO_BLOGGER rows.

Parses the ``List Micro`` or ``List Macro`` sheet (selected by *kind*),
locating the header by fingerprint (a row containing both NAME and
REDBOOK_ID) and the extraction date from the metadata rows above it.
Parsing a Micro file (~58.8k rows) is slow, so the parsed rows — with all
Tier-0 normalized forms precomputed — are cached in SQLite keyed by the
file's content hash salted with the kind (the same workbook uploaded as
Micro and as Macro parses different sheets, so the two cache entries must
never collide); warm loads take well under 2 s.

REDBOOK_ID is the XHS user id (same key space as the DMR Username column and
the resolver's author_id), treated as an opaque string with case-insensitive
lookup. Most perimeter rows have no REDBOOK_ID (other-platform bloggers).
"""
from __future__ import annotations

import hashlib
import io
import json
import re
import time
from dataclasses import dataclass, field
from typing import IO, Optional, Union

from openpyxl import load_workbook
from pypinyin import lazy_pinyin
from rapidfuzz import fuzz, process

from ..core import db
from ..core.textnorm import ascii_part, cjk, header_key, norm
from .name_match import (FUZZY_CUTOFF, METHOD_ASCII_FUZZY, METHOD_CJK,
                         METHOD_NORM, METHOD_PINYIN, MIN_COMPARE_LEN)

PERIMETER_REQUIRED = {"name", "redbook_id"}
HEADER_SCAN_ROWS = 15

_EXTRACTION_RE = re.compile(
    r"date of extraction\s*[:：]?\s*([0-9]{2}/[0-9]{2}/[0-9]{4}(?:\s+[0-9:]{5,8})?)",
    re.IGNORECASE)


def _s(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _i(v) -> Optional[int]:
    try:
        if v is None or v == "":
            return None
        return int(float(str(v).replace(",", "")))
    except (ValueError, TypeError):
        return None


@dataclass
class PerimeterParse:
    file_hash: str
    filename: str
    sheet: str
    header_row: int
    extraction_date: str
    rows: list[dict] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    rows_scanned: int = 0        # before the China-market filter
    china_filter: str = ""       # IN_CHINA_REPORTS | COUNTRY | "" (no column)

    @property
    def redbook_count(self) -> int:
        return sum(1 for r in self.rows if r["redbook_id"])


# Bumped whenever parsing semantics change: the cache is content-addressed,
# so without a version salt an already-cached file would keep serving rows
# parsed under the OLD semantics (e.g. unfiltered, pre-China-market).
# The micro salt must stay byte-identical while micro parse semantics are
# unchanged — existing cached Micro perimeters keep working after upgrades.
_PARSER_VERSION = b"2:china-market\x00"
_KIND_SALT = {"micro": b"", "macro": b"macro\x00"}
# Which perimeter lists a run checks (run option, "flick" toggle in the UI).
MODES = ("micro", "macro", "both")
SETTING_KEY = {"micro": "current_perimeter", "macro": "current_perimeter_macro"}

# The tool evaluates the Chinese market only. Micro sheets carry no
# IN_CHINA flag, so COUNTRY is the signal there; Macro sheets (future
# feature) have an explicit IN_CHINA_REPORTS YES/NO which wins when present.
# Verified on the real file: every REDBOOK_ID row is MAINLAND CHINA, so this
# filter cannot flip a membership verdict — it only drops the ~52k non-China
# rows that polluted same-name evidence scans.
_CHINA_COUNTRIES = {"MAINLANDCHINA", "CHINA"}


def _is_china_country(value: str) -> bool:
    return value.replace(" ", "").upper() in _CHINA_COUNTRIES


def file_hash(data: bytes, kind: str = "micro") -> str:
    return hashlib.sha256(_PARSER_VERSION + _KIND_SALT[kind] + data).hexdigest()


def _pick_sheet(wb, kind: str):
    want, other = ("listmicro", "macro") if kind == "micro" else ("listmacro", "micro")
    for name in wb.sheetnames:
        if header_key(name) == want:
            return wb[name]
    for name in wb.sheetnames:
        k = header_key(name)
        if kind in k and other not in k:
            return wb[name]
    return None


def parse_perimeter(source: Union[str, IO[bytes]], filename: str = "",
                    content_hash: str = "",
                    kind: str = "micro") -> PerimeterParse:
    """Parse a perimeter workbook (read-only mode — the file is big),
    reading the ``List Micro`` or ``List Macro`` sheet per *kind*."""
    wb = load_workbook(source, read_only=True, data_only=True)
    try:
        ws = _pick_sheet(wb, kind)
        if ws is None:
            raise ValueError(
                f"Perimeter parse failed: no 'List {kind.title()}' sheet "
                f"found (sheets: {wb.sheetnames})")

        header_row = None
        cols: dict[str, int] = {}
        meta_cells: list[str] = []
        for r_idx, row in enumerate(
                ws.iter_rows(min_row=1, max_row=HEADER_SCAN_ROWS,
                             values_only=True), 1):
            keys = {header_key(_s(v)): c for c, v in enumerate(row) if _s(v)}
            if PERIMETER_REQUIRED.issubset(keys.keys()):
                header_row = r_idx
                cols = keys
                break
            meta_cells.extend(_s(v) for v in row if _s(v))
        if header_row is None:
            raise ValueError(
                "Perimeter parse failed: no header row containing both 'NAME' "
                f"and 'REDBOOK_ID' within the first {HEADER_SCAN_ROWS} rows of "
                f"sheet {ws.title!r}.")

        result = PerimeterParse(
            file_hash=content_hash, filename=filename, sheet=ws.title,
            header_row=header_row, extraction_date="")
        m = _EXTRACTION_RE.search(" | ".join(meta_cells))
        if m:
            result.extraction_date = m.group(1).strip()
        else:
            result.warnings.append(
                "Could not find 'Date of extraction' in the metadata rows — "
                "perimeter staleness cannot be shown.")

        c_name, c_namebis = cols.get("name"), cols.get("namebis")
        c_dmrid, c_rid = cols.get("dmrid"), cols.get("redbook_id")
        c_rfol = cols.get("redbook_followers")
        c_inchina = cols.get("in_china_reports")
        c_country = cols.get("country")
        if c_inchina is not None:
            result.china_filter = "IN_CHINA_REPORTS"
        elif c_country is not None:
            result.china_filter = "COUNTRY"
        else:
            result.warnings.append(
                "No IN_CHINA_REPORTS or COUNTRY column found — cannot "
                "restrict the perimeter to the China market; keeping all rows.")

        def col(row, c):
            return _s(row[c]) if c is not None and c < len(row) else ""

        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            name = col(row, c_name)
            namebis = col(row, c_namebis)
            rid = col(row, c_rid).lower()
            if not name and not namebis:
                continue
            result.rows_scanned += 1
            # China market only — this tool evaluates Chinese-market KOLs
            if c_inchina is not None:
                if col(row, c_inchina).strip().upper() != "YES":
                    continue
            elif c_country is not None:
                if not _is_china_country(col(row, c_country)):
                    continue
            # Tier-0 normalized forms are precomputed here so the cached JSON
            # can be scanned without re-normalizing 58.8k names per run.
            result.rows.append({
                "name": name, "namebis": namebis,
                "dmrid": col(row, c_dmrid),
                "redbook_id": rid,
                "redbook_followers": _i(row[c_rfol]) if c_rfol is not None and c_rfol < len(row) else None,
                "nn": norm(name), "an": ascii_part(name),
                "nb": norm(namebis), "cb": cjk(namebis),
                "ab": ascii_part(namebis),
            })
        if not result.rows and result.rows_scanned:
            result.warnings.append(
                f"All {result.rows_scanned} perimeter rows were filtered out "
                "— none are China-market (COUNTRY=MAINLAND CHINA / "
                "IN_CHINA_REPORTS=YES). Check that this is the right file.")
        elif not result.rows:
            result.warnings.append("Perimeter sheet parsed but had no data rows.")
        return result
    finally:
        wb.close()


# ------------------------------------------------------------------ caching

def store_parsed(parsed: PerimeterParse) -> None:
    db.perimeter_cache_put(
        parsed.file_hash, filename=parsed.filename, sheet=parsed.sheet,
        extraction_date=parsed.extraction_date, row_count=len(parsed.rows),
        redbook_count=parsed.redbook_count,
        parsed_json=json.dumps(
            {"rows": parsed.rows, "rows_scanned": parsed.rows_scanned,
             "china_filter": parsed.china_filter}, ensure_ascii=False),
        warnings_json=json.dumps(parsed.warnings, ensure_ascii=False),
    )


def _payload(cached: dict) -> dict:
    """parsed_json is versioned by shape: v1 was a bare rows list (only ever
    stored under pre-salt hashes, but tolerate it), v2 a dict with counters."""
    loaded = json.loads(cached["parsed_json"])
    if isinstance(loaded, list):
        return {"rows": loaded, "rows_scanned": len(loaded), "china_filter": ""}
    return loaded


def load_cached(file_hash: str, filename: str = "") -> Optional["PerimeterIndex"]:
    """Load a cached index, retaining a run-specific upload name when given.

    The cache is content-addressed, so identical bytes uploaded under a new
    filename still share parsed rows without losing the name shown for that
    particular run.
    """
    row = db.perimeter_cache_get(file_hash)
    if not row:
        return None
    rows = _payload(row)["rows"]
    return PerimeterIndex(rows, extraction_date=row["extraction_date"] or "",
                          filename=filename or row["filename"] or "",
                          file_hash=file_hash)


# ------------------------------------------------------------------ indexes

class PerimeterIndex:
    """Lookup structures over the parsed Micro perimeter.

    ``by_redbook`` is the ground-truth join (same key space as Tier 2's
    author ids). ``scan_name`` reuses the Tier-0/Tier-3 name-ladder semantics
    against NAME + NAMEBIS and returns *all* hits — name collisions are real
    (e.g. 'esther' matches several rows), so callers must never auto-pick."""

    def __init__(self, rows: list[dict], extraction_date: str = "",
                 filename: str = "", file_hash: str = ""):
        self.rows = rows
        self.extraction_date = extraction_date
        self.filename = filename
        self.file_hash = file_hash
        self.by_redbook: dict[str, dict] = {}
        for r in rows:
            rid = (r.get("redbook_id") or "").lower()
            if rid and rid not in self.by_redbook:
                self.by_redbook[rid] = r
        # fuzzy choice lists (index-aligned with self.rows; short forms
        # excluded — partial_ratio saturates on 1-3 char strings)
        self._an = [(i, r["an"]) for i, r in enumerate(rows)
                    if len(r["an"]) >= MIN_COMPARE_LEN]
        self._ab = [(i, r["ab"]) for i, r in enumerate(rows)
                    if len(r["ab"]) >= MIN_COMPARE_LEN]

    def lookup_author(self, author_id: str) -> Optional[dict]:
        return self.by_redbook.get((author_id or "").strip().lower())

    def _fuzzy(self, query: str, choices: list[tuple[int, str]],
               hits: dict[int, str], method: str) -> None:
        if len(query) < MIN_COMPARE_LEN or not choices:
            return
        found = process.extract(
            query, [c[1] for c in choices], scorer=fuzz.partial_ratio,
            score_cutoff=FUZZY_CUTOFF, limit=None)
        for _choice, _score, pos in found:
            idx = choices[pos][0]
            hits.setdefault(idx, method)

    def scan_name(self, plog_name: str) -> list[tuple[dict, str]]:
        pc, pn = cjk(plog_name), norm(plog_name)
        pa = ascii_part(plog_name)
        hits: dict[int, str] = {}
        # ladder steps a/b — exact containment over precomputed forms
        for i, r in enumerate(self.rows):
            if pc and r["cb"] and pc in r["cb"]:
                hits.setdefault(i, METHOD_CJK)
            elif pn and ((r["nn"] and pn in r["nn"]) or (r["nb"] and pn in r["nb"])):
                hits.setdefault(i, METHOD_NORM)
        # ladder steps c/d — fuzzy over the ASCII forms (rapidfuzz batch)
        self._fuzzy(pa, self._an, hits, METHOD_ASCII_FUZZY)
        self._fuzzy(pa, self._ab, hits, METHOD_ASCII_FUZZY)
        if pc:
            pinyin = "".join(lazy_pinyin(pc)).casefold()
            self._fuzzy(pinyin, self._an, hits, METHOD_PINYIN)
            self._fuzzy(pinyin, self._ab, hits, METHOD_PINYIN)
        return [(self.rows[i], m) for i, m in sorted(hits.items())]


def parse_and_cache(data: bytes, filename: str,
                    kind: str = "micro") -> tuple[dict, list[str]]:
    """Parse-or-load an uploaded perimeter and cache it — WITHOUT making it
    the app-wide current perimeter. Promotion is a separate explicit step
    (``promote_cached``) tied to actually starting a run, so an abandoned
    preview can no longer swap the global perimeter under other users.

    Returns (meta, warnings). On a cache hit the warnings recorded at first
    parse are replayed instead of silently dropped."""
    h = file_hash(data, kind)
    cached = db.perimeter_cache_get(h)
    if cached:
        warnings = json.loads(cached.get("warnings_json") or "[]")
        payload = _payload(cached)
        meta = {
            "hash": h, "filename": filename, "kind": kind,
            "extraction_date": cached["extraction_date"] or "",
            "rows": cached["row_count"],
            "redbook_count": cached["redbook_count"],
            "rows_scanned": payload["rows_scanned"],
            "china_filter": payload["china_filter"],
        }
        return meta, warnings
    parsed = parse_perimeter(io.BytesIO(data), filename=filename,
                             content_hash=h, kind=kind)
    store_parsed(parsed)
    meta = {
        "hash": h, "filename": filename, "kind": kind,
        "extraction_date": parsed.extraction_date,
        "rows": len(parsed.rows), "redbook_count": parsed.redbook_count,
        "rows_scanned": parsed.rows_scanned,
        "china_filter": parsed.china_filter,
    }
    return meta, parsed.warnings


def promote_cached(file_hash_: str, filename: str = "",
                   kind: str = "micro") -> None:
    """Make a cached perimeter the app-wide default ('current') for its
    kind. Called when a run that uses it is actually started — never during
    preview. No-op if the hash is no longer cached."""
    row = db.perimeter_cache_get(file_hash_)
    if not row:
        return
    payload = _payload(row)
    db.setting_set(SETTING_KEY[kind], json.dumps({
        "hash": file_hash_,
        "filename": filename or row["filename"] or "",
        "kind": kind,
        "extraction_date": row["extraction_date"] or "",
        "rows": row["row_count"], "redbook_count": row["redbook_count"],
        "rows_scanned": payload["rows_scanned"],
        "china_filter": payload["china_filter"],
        "uploaded_at": time.time(),
    }, ensure_ascii=False))


def current_meta(kind: str = "micro") -> Optional[dict]:
    raw = db.setting_get(SETTING_KEY[kind])
    if not raw:
        return None
    try:
        return json.loads(raw)
    except ValueError:
        return None
