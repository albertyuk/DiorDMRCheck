"""KOL Efficiency Report — analysis engine.

xlsx upload → schema check → validation (warnings surfaced, never swallowed)
→ tier × coop classification → pooled + per-post metrics → verification.

Everything here mirrors the handoff spec worked out against the real
PLOG_DMR_CHECK.xlsx; the edge cases are real, not hypothetical:
- (CAMPAIGN, NO) is the row identity — NO restarts per campaign,
- TYPE is matched by substring (报备图文 / 软植图文 / bare 报备),
- 尾部 and 底部 are deliberately merged into BOT (documented judgment call,
  warned about whenever both coexist; tier_mode="fanbase" re-tiers instead),
- empty groups are null (no bar), never zero,
- the source file's CPM column is PRICE÷IMPRESSION (per SINGLE impression) —
  never reused for standard CPM,
- metrics are computed twice (raw loop + pandas) and reconciled before any
  rendering; cross-wave deltas are never generated (no prior-wave input).
"""
from __future__ import annotations

from dataclasses import dataclass, field, asdict
from datetime import date
import math
from typing import Optional

from openpyxl import load_workbook

from .. import config
from ..core.textnorm import header_key, nfkc
from ..core.uploads import UploadLimitError
from ..core.xlsx import cell_str, find_header_row, to_date, to_float
from ..i18n import make_td

TIERS = ["TOP", "MID", "BOT", "KOC"]
COOPS = ["SOFT", "PAID"]

REQUIRED_COLS = {
    "no": "NO", "campaign": "CAMPAIGN", "type": "TYPE", "level": "LEVEL",
    "name": "NAME", "fanbase(k)": "FAN BASE（K)", "postdate": "POST DATE",
    "postlink": "POST LINK", "impression": "IMPRESSION", "like": "LIKE",
    "collection": "COLLECTION", "comment": "COMMENT",
    "ttlengagement": "TTL ENGAGEMENT", "price": "PRICE",
}
METRIC_COLS = ("impression", "like", "collection", "comment",
               "ttl_engagement", "price")


@dataclass
class ReportConfig:
    basis: str = "pooled"                # pooled | per_post
    tier_mode: str = "label"             # label | fanbase
    fanbase_top_k: float = 1000
    fanbase_mid_k: float = 400
    fanbase_bot_k: float = 200           # below → KOC
    concentration_flag_pct: float = 50
    min_group_n: int = 3
    share_decimals: int = 1
    language: str = "en"                 # en | zh
    missing_row_policy: str = "exclude_warn"  # exclude_warn | block


@dataclass
class Row:
    idx: int                 # 0-based data row index (stable unique key)
    excel_row: int
    campaign: str
    no: str
    name: str
    type_raw: str
    level_raw: str
    fanbase_k: Optional[float]
    post_date: Optional[date]
    post_link: str
    impression: Optional[int]
    like: Optional[int]
    collection: Optional[int]
    comment: Optional[int]
    ttl_engagement: Optional[int]
    price: Optional[float]
    cpm_src: Optional[float]
    cpe_src: Optional[float]
    # filled by classification
    coop: str = ""           # PAID | SOFT | ""
    tier: str = ""           # TOP | MID | BOT | KOC | ""
    excluded: bool = False   # dropped from metrics (V2)
    invalid_metrics: list[str] = field(default_factory=list, repr=False)
    advisory_invalid_metrics: list[str] = field(default_factory=list,
                                                       repr=False)

    @property
    def key(self) -> tuple[str, str]:
        return (self.campaign, self.no)

    @property
    def group(self) -> str:
        return f"{self.tier} {self.coop}" if self.tier and self.coop else ""


@dataclass
class Finding:
    code: str
    severity: str            # ERROR | WARN
    message: str
    rows: list[int] = field(default_factory=list)   # excel rows

    def to_dict(self) -> dict:
        return asdict(self)


# ---------------------------------------------------------------- parsing

def _select_report_sheet(
        wb, findings: Optional[list[Finding]] = None):
    """Choose the preferred sheet that actually carries the full schema.

    A cover/instructions tab is common.  Falling back blindly to worksheet 1
    made valid reports fail whenever their data lived on a later tab.
    """
    named = next((ws for ws in wb.worksheets
                  if header_key(ws.title) == "masterkollist"), None)
    ordered = ([named] if named is not None else []) + [
        ws for ws in wb.worksheets if ws is not named
    ]
    for candidate in ordered:
        hit = find_header_row(candidate, set(REQUIRED_COLS))
        if hit:
            if findings is not None and candidate is not named:
                if named is None:
                    message = (
                        "Sheet 'MASTER KOL LIST' not found — using "
                        f"schema-compatible sheet {candidate.title!r}.")
                else:
                    message = (
                        "Sheet 'MASTER KOL LIST' lacks the required schema — "
                        f"using schema-compatible sheet {candidate.title!r}.")
                findings.append(Finding("V1", "WARN", message))
            return candidate, hit

    # Preserve a useful sheet name in the V1 error when no sheet matches.
    fallback = named or wb.worksheets[0]
    return fallback, None


def probe_report_schema(path_or_file) -> tuple[str, int]:
    """Cheap schema-only probe used to decide whether remapping is needed."""
    wb = load_workbook(path_or_file, read_only=True, data_only=True)
    try:
        ws, hit = _select_report_sheet(wb)
        if not hit:
            raise ValueError(
                "V1: no complete efficiency-report header row found in "
                f"sheet {ws.title!r} within the first 15 rows.")
        return ws.title, hit[0]
    finally:
        wb.close()


def parse_report(path_or_file) -> tuple[list[Row], list[Finding], dict]:
    """Parse the KOL sheet. Returns (rows, findings, meta). V1 failures are
    the only thing that raises — everything else is a Finding."""
    wb = load_workbook(path_or_file, data_only=True)
    try:
        return _parse_report_workbook(wb)
    finally:
        wb.close()


def _parse_report_workbook(wb) -> tuple[list[Row], list[Finding], dict]:
    """Parse an already-open workbook; :func:`parse_report` owns closing it."""
    findings: list[Finding] = []
    ws, hit = _select_report_sheet(wb, findings)

    # Bind only to a complete schema.  A metadata row containing the two old
    # fingerprint labels must not shadow a later, valid header row.
    if not hit:
        raise ValueError(
            "V1: no complete efficiency-report header row found in "
            f"sheet {ws.title!r} within the first 15 rows.")
    header_row, cols = hit

    missing = [label for key, label in REQUIRED_COLS.items() if key not in cols]
    if missing:
        raise ValueError(
            "V1: required columns missing after header normalization: "
            + ", ".join(missing))

    def col(key):
        return cols.get(key)

    c = {k: col(k) for k in (
        "no", "mcn", "campaign", "type", "level", "name", "fanbase(k)",
        "postdate", "micromacro", "postlink", "impression", "like",
        "collection", "comment", "ttlengagement", "price", "cpm", "cpe")}

    rows: list[Row] = []
    current_campaign = ""
    # Iterate physical rows that can actually be report records instead of the
    # worksheet dimension.  This handles legitimate blocks separated by 200+
    # blank rows without walking to a styled cell at Excel row 1,048,576.
    # Normal-mode openpyxl has already materialized every physical cell in
    # `_cells`; limiting to identity columns keeps this set bounded by the
    # logical row limit even for style-heavy workbooks.
    identity_cols = {c["name"], c["postlink"]}
    data_rows: set[int] = set()
    for (row_index, column_index), cell in ws._cells.items():
        if (row_index > header_row and column_index in identity_cols
                and (cell.value not in (None, "")
                     or getattr(cell, "hyperlink", None))):
            data_rows.add(row_index)
            if len(data_rows) > config.MAX_EFFICIENCY_ROWS:
                raise UploadLimitError(
                    "V1: workbook contains more than "
                    f"{config.MAX_EFFICIENCY_ROWS:,} efficiency rows.")

    for r in sorted(data_rows):

        def get(key):
            column = c.get(key)
            return ws.cell(row=r, column=column).value if column else None

        name = cell_str(get("name"))
        link_cell = ws.cell(row=r, column=c["postlink"]) if c.get("postlink") else None
        link = ""
        if link_cell is not None:
            if link_cell.hyperlink and link_cell.hyperlink.target:
                link = str(link_cell.hyperlink.target).strip()
            else:
                link = cell_str(link_cell.value)
        if not name and not link:
            continue
        campaign = cell_str(get("campaign"))
        if campaign:
            current_campaign = campaign
        invalid_metrics: list[str] = []
        advisory_invalid_metrics: list[str] = []

        def count_value(key: str) -> Optional[int]:
            raw = get(key)
            if raw is None or raw == "":
                return None
            value = to_float(raw)
            if (value is None or not math.isfinite(value) or value < 0
                    or not value.is_integer()):
                invalid_metrics.append(key.upper())
                return None
            return int(value)

        def amount_value(key: str, *, blocking: bool = True) -> Optional[float]:
            raw = get(key)
            if raw is None or raw == "":
                return None
            value = to_float(raw)
            if value is None or not math.isfinite(value) or value < 0:
                target = (invalid_metrics if blocking
                          else advisory_invalid_metrics)
                target.append(key.upper())
                return None
            return value

        # FAN BASE is only an input under fanbase tiering. Source CPM/CPE are
        # diagnostics and never feed report metrics, so malformed values in
        # those columns must not discard an otherwise valid report row.
        fanbase = amount_value("fanbase(k)", blocking=False)
        row = Row(
            idx=len(rows), excel_row=r,
            campaign=current_campaign, no=cell_str(get("no")), name=name,
            type_raw=cell_str(get("type")), level_raw=cell_str(get("level")),
            fanbase_k=fanbase,
            post_date=to_date(get("postdate")),
            post_link=link.strip(),
            impression=count_value("impression"), like=count_value("like"),
            collection=count_value("collection"),
            comment=count_value("comment"),
            ttl_engagement=count_value("ttlengagement"),
            price=amount_value("price"),
            cpm_src=amount_value("cpm", blocking=False),
            cpe_src=amount_value("cpe", blocking=False),
            invalid_metrics=invalid_metrics,
            advisory_invalid_metrics=advisory_invalid_metrics,
        )
        rows.append(row)
    meta = {"sheet": ws.title, "header_row": header_row, "rows": len(rows),
            "campaigns": sorted({r.campaign for r in rows if r.campaign})}
    return rows, findings, meta


# ----------------------------------------------------------- classification

def classify(rows: list[Row], cfg: ReportConfig,
             findings: list[Finding]) -> None:
    """COOP from TYPE (substring — 报备图文/软植图文/bare 报备 all classify);
    TIER from LEVEL (label mode merges 尾部+底部 into BOT) or from fan base
    thresholds (fanbase mode). Unknown values go to the warning bucket —
    never guessed."""
    bad_types: dict[str, list[int]] = {}
    bad_levels: dict[str, list[int]] = {}
    for r in rows:
        t = nfkc(r.type_raw)
        if "报备" in t:
            r.coop = "PAID"
        elif "软植" in t:
            r.coop = "SOFT"
        else:
            r.coop = ""
            bad_types.setdefault(r.type_raw or "(blank)", []).append(r.excel_row)

        if cfg.tier_mode == "fanbase":
            fb = r.fanbase_k
            if fb is None:
                r.tier = ""
                bad_levels.setdefault(
                    f"(no FAN BASE for fanbase tiering; LEVEL={r.level_raw!r})",
                    []).append(r.excel_row)
            elif fb >= cfg.fanbase_top_k:
                r.tier = "TOP"
            elif fb >= cfg.fanbase_mid_k:
                r.tier = "MID"
            elif fb >= cfg.fanbase_bot_k:
                r.tier = "BOT"
            else:
                r.tier = "KOC"
        else:
            lvl = nfkc(r.level_raw)
            if "头部" in lvl:
                r.tier = "TOP"
            elif "腰部" in lvl:
                r.tier = "MID"
            elif "尾部" in lvl or "底部" in lvl:
                r.tier = "BOT"   # deliberate merge — see V8
            elif "koc" in lvl.casefold():
                r.tier = "KOC"
            else:
                r.tier = ""
                bad_levels.setdefault(r.level_raw or "(blank)", []).append(r.excel_row)

    for value, rws in {**bad_types}.items():
        findings.append(Finding(
            "V7", "WARN",
            f"Unclassified TYPE value {value!r} on {len(rws)} row(s) — "
            "excluded from groups, counted in totals.", rws))
    for value, rws in {**bad_levels}.items():
        findings.append(Finding(
            "V7", "WARN",
            f"Unclassified LEVEL value {value!r} on {len(rws)} row(s) — "
            "excluded from groups, counted in totals.", rws))


# --------------------------------------------------------------- validation

def _eng(r: Row) -> Optional[int]:
    return r.ttl_engagement


def validate(rows: list[Row], cfg: ReportConfig,
             findings: list[Finding]) -> None:
    if cfg.tier_mode == "fanbase":
        for row in rows:
            if "FANBASE(K)" in row.advisory_invalid_metrics:
                row.invalid_metrics.append("FAN BASE（K)")

    # Invalid numeric domains are blocking.  Counts must be whole,
    # non-negative and finite; monetary/rate fields must be non-negative and
    # finite.  Excluding the rows prevents diagnostic metrics from carrying
    # nonsense values even though no deck will be generated.
    invalid_rows = [r for r in rows if r.invalid_metrics]
    if invalid_rows:
        for r in invalid_rows:
            r.excluded = True
        details = sorted({name for r in invalid_rows for name in r.invalid_metrics})
        findings.append(Finding(
            "V11", "ERROR",
            f"{len(invalid_rows)} row(s) contain invalid numeric values "
            "(counts must be whole; all values must be finite and "
            f"non-negative) in: {', '.join(details)}.",
            [r.excel_row for r in invalid_rows]))

    advisory: dict[str, list[int]] = {}
    for row in rows:
        for name in row.advisory_invalid_metrics:
            if name == "FANBASE(K)" and cfg.tier_mode == "fanbase":
                continue
            label = "FAN BASE（K)" if name == "FANBASE(K)" else name
            advisory.setdefault(label, []).append(row.excel_row)
    if advisory:
        fields = ", ".join(sorted(advisory))
        excel_rows = sorted({r for values in advisory.values() for r in values})
        findings.append(Finding(
            "V11", "WARN",
            f"{len(excel_rows)} row(s) contain invalid diagnostics-only "
            f"numeric values in {fields} — those source values were ignored.",
            excel_rows))

    # V2 — missing metric values
    missing_rows = [r for r in rows if not r.invalid_metrics and any(
        getattr(r, col) is None for col in METRIC_COLS)]
    if missing_rows:
        if cfg.missing_row_policy == "block":
            # ERROR blocks generation (caller checks `blocked`), but the rows
            # are still excluded so the diagnostic metrics stay computable.
            for r in missing_rows:
                r.excluded = True
            findings.append(Finding(
                "V2", "ERROR",
                f"{len(missing_rows)} row(s) missing IMPRESSION/LIKE/"
                "COLLECTION/COMMENT/TTL ENGAGEMENT/PRICE values "
                "(missing_row_policy=block).",
                [r.excel_row for r in missing_rows]))
        else:
            for r in missing_rows:
                r.excluded = True
            findings.append(Finding(
                "V2", "WARN",
                f"{len(missing_rows)} row(s) missing metric values — excluded "
                "from all metrics (missing_row_policy=exclude_warn).",
                [r.excel_row for r in missing_rows]))

    active = [r for r in rows if not r.excluded]

    # V3 — zero denominators
    zero_impr = [r.excel_row for r in active if (r.impression or 0) == 0]
    zero_eng = [r.excel_row for r in active if (_eng(r) or 0) == 0]
    if zero_impr:
        findings.append(Finding(
            "V3", "WARN",
            f"{len(zero_impr)} row(s) with IMPRESSION=0 — excluded from CPM "
            "ratios only.", zero_impr))
    if zero_eng:
        findings.append(Finding(
            "V3", "WARN",
            f"{len(zero_eng)} row(s) with TTL ENGAGEMENT=0 — excluded from "
            "CPE ratios only.", zero_eng))

    # V4 — engagement identity
    bad = [r.excel_row for r in active
           if None not in (r.ttl_engagement, r.like, r.collection, r.comment)
           and r.ttl_engagement != (r.like + r.collection + r.comment)]
    if bad:
        findings.append(Finding(
            "V4", "WARN",
            f"TTL ENGAGEMENT ≠ LIKE+COLLECTION+COMMENT on {len(bad)} row(s).",
            bad))

    # V5 — source CPM/CPE columns vs recomputed (source CPM is PRICE per
    # SINGLE impression, not per mille — never reuse it as standard CPM)
    drift = []
    for r in active:
        if r.cpm_src is not None and r.price is not None and (r.impression or 0) > 0:
            expected = r.price / r.impression
            if expected and abs(r.cpm_src - expected) / expected > 0.02:
                drift.append(r.excel_row)
        if r.cpe_src is not None and r.price is not None and (_eng(r) or 0) > 0:
            expected = r.price / r.ttl_engagement
            if expected and abs(r.cpe_src - expected) / expected > 0.02:
                drift.append(r.excel_row)
    if drift:
        findings.append(Finding(
            "V5", "WARN",
            "Source CPM/CPE drift from recomputation on "
            f"{len(sorted(set(drift)))} row(s). Note the source CPM column is "
            "PRICE÷IMPRESSION — cost per SINGLE impression, ×1000 off the "
            "industry-standard per-mille CPM; this report never reuses it.",
            sorted(set(drift))))

    # V6 — duplicate POST LINK (stripped)
    seen: dict[str, list[Row]] = {}
    for r in rows:
        if r.post_link:
            seen.setdefault(r.post_link, []).append(r)
    for link, rs in seen.items():
        if len(rs) > 1:
            names = " / ".join(x.name for x in rs)
            findings.append(Finding(
                "V6", "WARN",
                f"Duplicate POST LINK shared by {len(rs)} rows ({names}): "
                f"…{link[-24:]} — likely a copy-paste error in the source; "
                "metrics keep both rows, but verify.",
                [x.excel_row for x in rs]))

    # V8 — 尾部/底部 coexistence (documented merge) + fan-range detail
    if cfg.tier_mode == "label":
        weibu = [r for r in rows if "尾部" in nfkc(r.level_raw)]
        dibu = [r for r in rows if "底部" in nfkc(r.level_raw)]
        if weibu and dibu:
            fb = [r.fanbase_k for r in dibu if r.fanbase_k is not None]
            koc_sized = sum(1 for x in fb if x < cfg.fanbase_bot_k)
            wr = [r.fanbase_k for r in weibu if r.fanbase_k is not None]

            def _range(vals: list[float]) -> str:
                # either label set can lack FAN BASE entirely — min([]) raises
                return f"{min(vals):.0f}–{max(vals):.0f}" if vals else "?"

            findings.append(Finding(
                "V8", "WARN",
                f"Both 尾部 ({len(weibu)} rows, fans "
                f"{_range(wr)}K) and 底部 ({len(dibu)} rows, "
                f"fans {_range(fb)}K) labels coexist — merged "
                f"into BOT (documented judgment call). {koc_sized} of the "
                f"{len(dibu)} 底部 accounts have <{cfg.fanbase_bot_k:.0f}K "
                "fans (KOC-sized). Set tier_mode=fanbase to re-tier by "
                "thresholds instead."))


def validate_groups(groups: dict[str, dict], total_rows: int,
                    cfg: ReportConfig, findings: list[Finding]) -> None:
    """V9 (small samples) and V10 (concentration) — need computed metrics."""
    for gname, g in groups.items():
        if 0 < g["n"] < cfg.min_group_n:
            findings.append(Finding(
                "V9", "WARN",
                f"{gname}: n={g['n']} — below min_group_n="
                f"{cfg.min_group_n}; on-slide caveat added (not a benchmark)."))
        conc = g.get("concentration_pct")
        # concentration is trivially 100% for n ≤ 2 — V9 covers those groups
        if g["n"] > 2 and conc is not None and conc > cfg.concentration_flag_pct:
            findings.append(Finding(
                "V10", "WARN",
                f"{gname}: top {g['concentration_top_k']} post(s) hold "
                f"{conc:.0f}% of group impressions — pooled CPM is dragged by "
                f"outliers; plan on per-post ≈{g['cpm_perpost']:.0f}, not "
                f"pooled {g['cpm_pooled']:.0f}. On-slide caveat added."))


# ------------------------------------------------------------------ metrics

def _mean(xs: list[float]) -> Optional[float]:
    return sum(xs) / len(xs) if xs else None


def compute_metrics(rows: list[Row], cfg: ReportConfig) -> dict:
    """Primary implementation: raw-loop accumulation. Group key 'TIER COOP'.
    Empty groups are absent (rendered as null bars, never zeros)."""
    active = [r for r in rows if not r.excluded]
    total_rows = len(active)
    groups: dict[str, dict] = {}
    for tier in TIERS:
        for coop in COOPS:
            g = [r for r in active if r.tier == tier and r.coop == coop]
            if not g:
                continue
            prices = [r.price for r in g]
            impr_rows = [r for r in g if (r.impression or 0) > 0]
            eng_rows = [r for r in g if (_eng(r) or 0) > 0]
            spend = sum(prices)
            impr = sum(r.impression for r in impr_rows)
            eng = sum(r.ttl_engagement for r in eng_rows)
            top2 = sorted((r.impression or 0 for r in g), reverse=True)[:2]
            conc = (sum(top2) / impr * 100) if impr else None
            groups[f"{tier} {coop}"] = {
                "tier": tier, "coop": coop,
                "n": len(g),
                "share": round(len(g) / total_rows * 100, cfg.share_decimals),
                "avg_price": _mean(prices),
                "cpm_pooled": spend / impr * 1000 if impr else None,
                "cpe_pooled": spend / eng if eng else None,
                "cpm_perpost": _mean([r.price / r.impression * 1000
                                      for r in impr_rows]),
                "cpe_perpost": _mean([r.price / r.ttl_engagement
                                      for r in eng_rows]),
                "spend": spend,
                "impressions": sum(r.impression or 0 for r in g),
                "engagements": sum(_eng(r) or 0 for r in g),
                "concentration_pct": conc,
                "concentration_top_k": min(2, len(g)),
            }
    unclassified = [r for r in active if not r.group]
    totals = {
        "rows": total_rows,
        "classified": total_rows - len(unclassified),
        "unclassified": len(unclassified),
        "unclassified_share": round(
            len(unclassified) / total_rows * 100, cfg.share_decimals)
        if total_rows else 0.0,
        "spend": sum(r.price for r in active),
        "impressions": sum(r.impression or 0 for r in active),
        "engagements": sum(_eng(r) or 0 for r in active),
        "excluded_rows": len(rows) - total_rows,
    }
    return {"groups": groups, "totals": totals}


def compute_metrics_pandas(rows: list[Row], cfg: ReportConfig) -> dict:
    """Independent second path (pandas groupby) for the dual-path check."""
    # pandas is imported lazily on purpose: it is heavy, and only this
    # verification path needs it.
    import pandas as pd
    active = [r for r in rows if not r.excluded]
    df = pd.DataFrame([{
        "group": r.group, "price": r.price, "impression": r.impression or 0,
        "engagement": _eng(r) or 0} for r in active])
    out: dict[str, dict] = {}
    if df.empty:
        return out
    for gname, g in df[df["group"] != ""].groupby("group"):
        gi = g[g["impression"] > 0]
        ge = g[g["engagement"] > 0]
        spend = float(g["price"].sum())
        impressions = float(g["impression"].sum())
        engagements = float(g["engagement"].sum())
        out[gname] = {
            "n": int(len(g)),
            "share": round(len(g) / len(df) * 100, cfg.share_decimals),
            "avg_price": float(g["price"].mean()),
            "cpm_pooled": spend / impressions * 1000 if impressions else None,
            "cpe_pooled": spend / engagements if engagements else None,
            "cpm_perpost": float((gi["price"] / gi["impression"] * 1000).mean())
            if len(gi) else None,
            "cpe_perpost": float((ge["price"] / ge["engagement"]).mean())
            if len(ge) else None,
        }
    return out


class VerificationError(RuntimeError):
    """An internal cross-check failed — never render output past this."""


def verify_dual_path(primary: dict, secondary: dict) -> None:
    keys = ("n", "share", "avg_price", "cpm_pooled", "cpe_pooled",
            "cpm_perpost", "cpe_perpost")
    if set(primary["groups"]) != set(secondary):
        raise VerificationError(
            f"dual-path group mismatch: {sorted(primary['groups'])} vs "
            f"{sorted(secondary)}")
    for gname, g in primary["groups"].items():
        for k in keys:
            a, b = g[k], secondary[gname][k]
            if a is None and b is None:
                continue
            if a is None or b is None or abs(a - b) > 1e-6:
                raise VerificationError(
                    f"dual-path mismatch {gname}.{k}: {a} vs {b}")


def verify_reconciliation(rows: list[Row], metrics: dict,
                          share_decimals: int = 1) -> None:
    groups, totals = metrics["groups"], metrics["totals"]
    n_sum = sum(g["n"] for g in groups.values()) + totals["unclassified"]
    if n_sum != totals["rows"]:
        raise VerificationError(
            f"Σ group n ({n_sum}) != total rows ({totals['rows']})")
    spend_sum = sum(g["spend"] for g in groups.values()) + sum(
        r.price for r in rows if not r.excluded and not r.group)
    if abs(spend_sum - totals["spend"]) > 1e-6:
        raise VerificationError(
            f"Σ group spend ({spend_sum}) != total spend ({totals['spend']})")
    # Empty/all-excluded input is a user validation failure, handled before
    # this internal verifier.  Keeping this guard makes direct callers safe.
    if totals["rows"] == 0:
        raise ValueError("V2: workbook contains no analyzable efficiency rows.")
    share_sum = sum(g["share"] for g in groups.values()) + totals["unclassified_share"]
    # Each slice is rounded to `share_decimals` BEFORE summing, so legitimate
    # drift grows with the slice count: up to half an ulp per slice (e.g. 9
    # slices at 1 decimal → 0.45). A fixed 0.3 rejected valid distributions.
    n_slices = len(groups) + (1 if totals["unclassified"] else 0)
    tolerance = n_slices * 0.5 * 10.0 ** -share_decimals + 1e-9
    if abs(share_sum - 100.0) > tolerance:
        raise VerificationError(
            f"donut shares sum to {share_sum} at display precision "
            f"(tolerance ±{tolerance:.2f} for {n_slices} slices)")


# ----------------------------------------------------------------- insights

def _fmt_k(v: float) -> str:
    return f"{v / 1000:.1f}K"


def build_insights(metrics: dict, cfg: ReportConfig,
                   findings: list[Finding]) -> dict:
    """Data-driven bullets only. Cross-wave comparisons are NEVER generated —
    there is no prior-wave input, and fabricated wave deltas are the most
    damaging failure mode of this feature."""
    groups = metrics["groups"]
    basis = cfg.basis
    cpm_key = "cpm_pooled" if basis == "pooled" else "cpm_perpost"
    cpe_key = "cpe_pooled" if basis == "pooled" else "cpe_perpost"

    price_bullets: list[str] = []
    premiums, inversions, ties = [], [], []
    for tier in TIERS:
        paid, soft = groups.get(f"{tier} PAID"), groups.get(f"{tier} SOFT")
        if not (paid and soft):
            continue
        delta = paid["avg_price"] - soft["avg_price"]
        if round(delta / 1000, 1) == 0:
            ties.append(tier)
        elif delta > 0:
            premiums.append(f"{tier} +{_fmt_k(delta)}")
        else:
            inversions.append(f"{tier} {_fmt_k(delta)}")
    if premiums and not inversions and not ties:
        price_bullets.append(
            "PAID CARRIES A PREMIUM IN EVERY TIER — "
            + " / ".join(premiums) + " VS SOFT")
    elif premiums or inversions:
        if premiums:
            price_bullets.append("PAID PREMIUM: " + " / ".join(premiums) + " VS SOFT")
        if inversions:
            price_bullets.append(
                "PRICE INVERSION — SOFT PRICED ABOVE PAID: " + " / ".join(inversions))
    if ties:
        price_bullets.append("PRICE TIE: " + " / ".join(ties))

    eff_bullets: list[str] = []
    cpm_parts, cpe_parts = [], []
    for tier in TIERS:
        paid, soft = groups.get(f"{tier} PAID"), groups.get(f"{tier} SOFT")
        if not (paid and soft) or paid[cpm_key] is None or soft[cpm_key] is None:
            pass
        else:
            paid_text, soft_text = f"{paid[cpm_key]:.0f}", f"{soft[cpm_key]:.0f}"
            if paid_text == soft_text:
                cpm_parts.append(f"{tier}: TIE ¥{paid_text}")
            else:
                winner = "PAID" if paid[cpm_key] < soft[cpm_key] else "SOFT"
                lower, higher = sorted((paid[cpm_key], soft[cpm_key]))
                cpm_parts.append(
                    f"{tier}: {winner} ¥{lower:.0f} VS ¥{higher:.0f}")
        if not (paid and soft) or paid[cpe_key] is None or soft[cpe_key] is None:
            continue
        paid_text, soft_text = f"{paid[cpe_key]:.1f}", f"{soft[cpe_key]:.1f}"
        if paid_text == soft_text:
            cpe_parts.append(f"{tier}: TIE ¥{paid_text}")
        else:
            winner = "PAID" if paid[cpe_key] < soft[cpe_key] else "SOFT"
            lower, higher = sorted((paid[cpe_key], soft[cpe_key]))
            cpe_parts.append(
                f"{tier}: {winner} ¥{lower:.1f} VS ¥{higher:.1f}")
    if cpm_parts:
        eff_bullets.append("CPM COMPARISON — " + " · ".join(cpm_parts))
    if cpe_parts:
        eff_bullets.append("CPE COMPARISON — " + " · ".join(cpe_parts))

    caveats: list[str] = []
    for f in findings:
        if f.code == "V9":
            gname = f.message.split(":")[0]
            n = groups[gname]["n"] if gname in groups else "?"
            caveats.append(f"{gname} = {n} POST(S) ONLY — NOT A BENCHMARK")
        if f.code == "V10":
            gname = f.message.split(":")[0]
            g = groups.get(gname)
            if g:
                caveats.append(
                    f"CAUTION: {gname} CARRIED BY {g['concentration_top_k']} "
                    f"VIRAL POSTS ({g['concentration_pct']:.0f}% OF GROUP "
                    f"IMPRESSIONS) — PLAN ON ~¥{g['cpm_perpost']:.0f} CPM, "
                    f"NOT ¥{g['cpm_pooled']:.0f}")

    # pooled↔per-post gap diagnostic (feeds the same caution family)
    for gname, g in groups.items():
        if (g["cpm_pooled"] and g["cpm_perpost"]
                and g["cpm_perpost"] / g["cpm_pooled"] > 3
                and not any(gname in c for c in caveats)):
            caveats.append(
                f"{gname}: PER-POST CPM ¥{g['cpm_perpost']:.0f} IS "
                f">3× POOLED ¥{g['cpm_pooled']:.0f} — RESULTS CONCENTRATED "
                "IN OUTLIERS")

    tier_n = {t: sum(g["n"] for g in groups.values() if g["tier"] == t)
              for t in TIERS}
    if basis == "pooled":
        footnote = (
            "Basis: pooled — group total spend ÷ total impressions "
            "(CPM, ¥ per 1,000) / total engagements (CPE). "
            f"n = {metrics['totals']['rows']} posts: "
            + " · ".join(f"{t} {tier_n[t]}" for t in TIERS if tier_n[t]))
    else:
        footnote = (
            "Basis: per-post average — mean of PRICE÷IMPRESSION×1000 (CPM, "
            "¥ per 1,000) / PRICE÷ENGAGEMENT (CPE) across posts. "
            f"n = {metrics['totals']['rows']} posts: "
            + " · ".join(f"{t} {tier_n[t]}" for t in TIERS if tier_n[t]))

    # The HTML layer can translate dynamic English diagnostics at render time,
    # but deck text is embedded directly.  Translate the complete insight
    # payload here when a Chinese slide was requested.
    td = make_td(cfg.language)

    def localized(text: str) -> str:
        rendered = td(text)
        if cfg.language == "zh":
            rendered = (rendered.replace(": TIE ", "：持平 ")
                        .replace(" VS ¥", "，对比 ¥"))
        return rendered

    return {"price": [localized(x) for x in price_bullets],
            "efficiency": [localized(x) for x in eff_bullets],
            "caveats": [localized(x) for x in caveats],
            "footnote": localized(footnote)}


# ---------------------------------------------------------------- pipeline

def analyze(path_or_file, cfg: Optional[ReportConfig] = None) -> dict:
    """Full analysis: parse → classify → validate → metrics → verify →
    insights. Raises ValueError on V1, VerificationError on failed internal
    cross-checks; ERROR findings block downstream generation (caller checks).
    """
    cfg = cfg or ReportConfig()
    rows, findings, meta = parse_report(path_or_file)
    classify(rows, cfg, findings)
    validate(rows, cfg, findings)

    if not any(not r.excluded for r in rows):
        invalid = [r for r in rows if r.invalid_metrics]
        if invalid:
            fields = sorted({name for r in invalid for name in r.invalid_metrics})
            excel_rows = ", ".join(str(r.excel_row) for r in invalid[:20])
            suffix = "…" if len(invalid) > 20 else ""
            raise ValueError(
                "V11: workbook contains no analyzable rows because invalid "
                f"numeric values occur in {', '.join(fields)} on Excel rows "
                f"{excel_rows}{suffix}.")
        raise ValueError("V2: workbook contains no analyzable efficiency rows.")

    blocked = any(f.severity == "ERROR" for f in findings)
    metrics = compute_metrics(rows, cfg)
    if not blocked:
        verify_dual_path(metrics, compute_metrics_pandas(rows, cfg))
        verify_reconciliation(rows, metrics, cfg.share_decimals)
    validate_groups(metrics["groups"], metrics["totals"]["rows"], cfg, findings)
    insights = build_insights(metrics, cfg, findings)
    return {
        "meta": meta, "config": asdict(cfg),
        "findings": [f.to_dict() for f in findings],
        "blocked": blocked,
        "metrics": metrics, "insights": insights,
        "rows_total": len(rows),
    }
