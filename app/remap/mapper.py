"""LLM-assisted header mapping for unfamiliar sheet formats, human-audited.

When a workbook's headers don't match the deterministic fingerprint, Claude
is shown a small STRUCTURAL sample (sheet names + the first rows, cells
truncated) and asked one thing: which sheet / header row / columns correspond
to the canonical fields. It answers with a column mapping and per-field
confidence — it never rewrites data, and nothing it says takes effect until
a human approves the mapping on the audit screen.

Applying a mapping rewrites ONLY the text of the identified header cells to
the canonical names; every data cell (and hyperlink) is byte-identical, so
the deterministic parsers — and all their guarantees — run unchanged.

Approved mappings are cached by a signature of the header row's LAYOUT
(sheet name, row index, header cell texts — never the data rows), so a given
format needs one LLM call and one human approval ever; later uploads of the
same format apply it automatically (the preview still shows what was mapped
and who approved it).
"""
from __future__ import annotations

import hashlib
import io
import json
import time
from typing import Any, Optional

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from pydantic import BaseModel, Field, ValidationError

from .. import config
from ..core import db
from ..core import llm
from ..core.textnorm import header_key
from ..core.xlsx import HEADER_SCAN_ROWS, cell_str
from .registry import FIELDS, KIND_LABELS  # noqa: F401  (re-exported)

# The sample must cover at least the rows the deterministic parser scans for
# a header, or the mapper could approve a header row the parser cannot find.
SAMPLE_ROWS = HEADER_SCAN_ROWS
SAMPLE_COLS = 24
SAMPLE_CELL_CHARS = 60
MAX_SHEETS = 6


class SchemaMapError(Exception):
    """Mapper unavailable or the model's answer was unusable."""





# ------------------------------------------------------------------- sample

def _sample_cell_str(v: Any) -> str:
    """``cell_str`` truncated only for model samples and audit display."""
    return cell_str(v)[:SAMPLE_CELL_CHARS]


def build_sample(data: bytes) -> dict:
    """Structural sample sent to the model: sheet names + top-left grid."""
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    try:
        sheets = []
        for ws in wb.worksheets[:MAX_SHEETS]:
            rows = []
            for row in ws.iter_rows(min_row=1, max_row=SAMPLE_ROWS,
                                    max_col=SAMPLE_COLS):
                rows.append([_sample_cell_str(c.value) for c in row])
            while rows and not any(rows[-1]):
                rows.pop()
            sheets.append({"name": ws.title, "rows": rows})
        return {"sheets": sheets}
    finally:
        wb.close()


def _layout_sig(sheet: str, row: int, cells: list[str]) -> str:
    return hashlib.sha256(json.dumps(
        {"sheet": sheet, "row": row, "cells": cells},
        ensure_ascii=False, sort_keys=True).encode()).hexdigest()[:32]


def _trim(cells: list[str]) -> list[str]:
    while cells and not cells[-1]:
        cells.pop()
    return cells


def header_signature(data: bytes, sheet: str, header_row: int) -> str:
    """Cache key for one candidate header row: sheet name, row index, and the
    row's cell texts (trailing blanks trimmed). Layout only — the data rows
    below and the metadata rows above do NOT participate, so re-uploads of a
    known format hit the approved cache even though their content differs.
    (The previous whole-sample hash covered data rows and the DMR metadata
    dates, which change on every export — the cache could effectively never
    hit twice.)"""
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    try:
        cells: list[str] = []
        for row in wb[sheet].iter_rows(min_row=header_row, max_row=header_row,
                                       max_col=SAMPLE_COLS):
            cells = _trim([cell_str(c.value) for c in row])
    finally:
        wb.close()
    return _layout_sig(sheet, header_row, cells)


def candidate_signatures(data: bytes) -> list[tuple[str, int, str]]:
    """(sheet, row, signature) for every non-empty row in the header-scan
    region of the first MAX_SHEETS sheets — the lookup keys probed against
    the approved-mapping cache before any LLM call."""
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    out = []
    try:
        for ws in wb.worksheets[:MAX_SHEETS]:
            for i, row in enumerate(
                    ws.iter_rows(min_row=1, max_row=SAMPLE_ROWS,
                                 max_col=SAMPLE_COLS), start=1):
                cells = _trim([cell_str(c.value) for c in row])
                if cells:
                    out.append((ws.title, i, _layout_sig(ws.title, i, cells)))
    finally:
        wb.close()
    return out


# ----------------------------------------------------------------- proposal

class Proposal(BaseModel):
    sheet: str
    header_row: int = Field(ge=1, le=SAMPLE_ROWS)
    columns: dict[str, Optional[int]]         # field key → 1-based column
    confidence: dict[str, float] = {}
    warnings: list[str] = []


SYSTEM_PROMPT = (
    "You map spreadsheet columns to a canonical schema for a KOL-campaign "
    "reconciliation tool. You are given the sheet names and the first rows "
    "of each sheet of an uploaded workbook, plus the canonical fields with "
    "descriptions.\n"
    "Rules:\n"
    "- Answer ONLY with JSON matching the given schema. No prose outside it.\n"
    "- Choose ONE sheet and ONE header row; all fields map to columns of that "
    "row. Headers may be in any language (Chinese exports are common).\n"
    "- Map a field only when the header meaning genuinely corresponds; use "
    "null when absent or when you would be guessing. NEVER map two fields to "
    "the same column.\n"
    "- confidence is 0..1 per mapped field.\n"
    "- Put anything a human must double-check into warnings — especially "
    "suspected unit differences (e.g. impressions that look like thousands), "
    "rate-vs-count columns, or two plausible candidate columns. Do NOT "
    "convert or fix anything; mapping is your only output."
)

def _call_llm(system: str, user: str) -> str:
    """Isolated for tests to monkeypatch."""
    return llm.complete(llm.make_client(), system=system, user=user,
                        max_tokens=2000)


def _parse_proposal(text: str) -> Optional[Proposal]:
    for cand in llm.json_candidates(text):
        try:
            return Proposal.model_validate(json.loads(cand))
        except (ValueError, ValidationError):
            continue
    return None


def propose(sample: dict, kind: str) -> Proposal:
    """One LLM call → validated Proposal. Raises SchemaMapError when the
    mapper can't run or the answer is unusable."""
    if not config.ANTHROPIC_API_KEY:
        raise SchemaMapError("ANTHROPIC_API_KEY is not configured")
    fields = FIELDS[kind]
    user = json.dumps({
        "file_kind": KIND_LABELS[kind],
        "canonical_fields": [
            {"key": key, "canonical_header": text, "required": req,
             "meaning": desc} for text, key, req, desc in fields],
        "workbook_sample": sample,
        "answer_schema": {
            "sheet": "<sheet name>", "header_row": "<1-based row number>",
            "columns": {"<field key>": "<1-based column number or null>"},
            "confidence": {"<field key>": 0.0},
            "warnings": ["<string>"],
        },
    }, ensure_ascii=False)
    try:
        raw = _call_llm(SYSTEM_PROMPT, user)
    except Exception as e:  # network/auth/SDK — surface, never crash the app
        raise SchemaMapError(f"LLM call failed: {e}") from e
    prop = _parse_proposal(raw)
    if prop is None:
        raise SchemaMapError("The model's answer was not valid mapping JSON.")

    names = {s["name"] for s in sample["sheets"]}
    if prop.sheet not in names:
        raise SchemaMapError(f"Model chose unknown sheet {prop.sheet!r}.")
    known = {key for _, key, _, _ in fields}
    prop.columns = {k: v for k, v in prop.columns.items() if k in known}
    seen: dict[int, str] = {}
    for k, v in list(prop.columns.items()):
        if v is None:
            continue
        if v < 1 or v > SAMPLE_COLS or v in seen:
            prop.columns[k] = None      # out of range / duplicate → unmapped
        else:
            seen[v] = k
    return prop


# ------------------------------------------------------------------- apply

def column_choices(data: bytes, sheet: str, header_row: int,
                   n_samples: int = 2) -> list[dict]:
    """For the audit UI: every column of the proposed header row with its
    header text and the first data values underneath."""
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    try:
        ws = wb[sheet]
        out = []
        for col in range(1, SAMPLE_COLS + 1):
            header = _sample_cell_str(
                ws.cell(row=header_row, column=col).value)
            samples = [
                _sample_cell_str(
                    ws.cell(row=header_row + 1 + i, column=col).value)
                for i in range(n_samples)]
            if header or any(samples):
                out.append({"col": col, "letter": get_column_letter(col),
                            "header": header,
                            "samples": [s for s in samples if s]})
        return out
    finally:
        wb.close()


def apply_mapping(data: bytes, kind: str, sheet: str, header_row: int,
                  columns: dict[str, int]) -> bytes:
    """Rewrite ONLY the chosen header cells to the canonical names. Data
    cells, formulas, and hyperlinks are untouched. Unchosen header cells that
    would collide with a canonical name are prefixed so the fingerprint
    cannot bind to the wrong column."""
    canonical = {key: text for text, key, _, _ in FIELDS[kind]}
    wb = load_workbook(io.BytesIO(data))       # NOT data_only: keep formulas
    ws = wb[sheet]
    chosen_cols = set(columns.values())
    mapped_keys = {header_key(canonical[k]) for k in columns}
    # de-collide first, then write (order matters when a canonical name
    # already exists on a column the human did NOT choose)
    for col in range(1, ws.max_column + 1):
        if col in chosen_cols:
            continue
        cell = ws.cell(row=header_row, column=col)
        if cell.value is None:
            continue
        if header_key(cell_str(cell.value)) in mapped_keys:
            cell.value = f"(original) {cell.value}"
    for key, col in columns.items():
        ws.cell(row=header_row, column=col).value = canonical[key]
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# -------------------------------------------------------------------- cache

def _cache_key(kind: str, sig: str) -> str:
    return f"schemamap:{kind}:{sig}"


def cache_get(kind: str, sig: str) -> Optional[dict]:
    raw = db.setting_get(_cache_key(kind, sig))
    return json.loads(raw) if raw else None


def cache_get_many(kind: str, sigs: list[str]) -> dict[str, dict]:
    """sig → approved mapping for every candidate signature that has one
    (single settings query)."""
    hits = db.settings_get_many([_cache_key(kind, s) for s in sigs])
    return {s: json.loads(hits[_cache_key(kind, s)])
            for s in sigs if _cache_key(kind, s) in hits}


def cache_put(kind: str, sig: str, sheet: str, header_row: int,
              columns: dict[str, int], approved_by: str) -> None:
    db.setting_set(_cache_key(kind, sig), json.dumps({
        "sheet": sheet, "header_row": header_row, "columns": columns,
        "approved_by": approved_by,
        "approved_at": time.strftime("%Y-%m-%d %H:%M"),
    }, ensure_ascii=False))
