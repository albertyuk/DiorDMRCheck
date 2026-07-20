"""KOL Efficiency Report — analysis engine.

xlsx upload → schema check → validation (warnings surfaced, never swallowed)
→ tier × coop classification → pooled + per-post metrics → verification.

Everything here mirrors the handoff spec worked out against the real
PLOG_DMR_CHECK.xlsx; the edge cases are real, not hypothetical:
- (CAMPAIGN, NO) is the row identity — NO restarts per campaign,
- TYPE is matched by substring (报备图文 / 软植图文 / bare 报备),
- 尾部 and 底部 are deliberately merged into BOT (documented judgment call,
  warned about whenever both coexist; tier_mode="fanbase" re-tiers instead),
- empty groups are null (no bar), never zero,
- the source file's CPM column is PRICE÷IMPRESSION (per SINGLE impression) —
  never reused for standard CPM,
- metrics are computed twice (raw loop + pandas) and reconciled before any
  rendering; cross-wave deltas are never generated (no prior-wave input).
"""
from __future__ import annotations

from dataclasses import dataclass, field, asdict
from datetime import date
from typing import Any, Optional

from openpyxl import load_workbook

from .normalize import header_key, nfkc
from .parsers import (HEADER_SCAN_ROWS, MAX_CONSECUTIVE_BLANK_ROWS, _cell_str,
                      _find_header_row, _to_date, _to_int)

TIERS = ["TOP", "MID", "BOT", "KOC"]
COOPS = ["SOFT", "PAID"]

REQUIRED_COLS = {
    "no": "NO", "campaign": "CAMPAIGN", "type": "TYPE", "level": "LEVEL",
    "name": "NAME", "fanbase(k)": "FAN BASE（K)", "postdate": "POST DATE",
    "postlink": "POST LINK", "impression": "IMPRESSION", "like": "LIKE",
    "collection": "COLLECTION", "comment": "COMMENT",
    "ttlengagement": "TTL ENGAGEMENT", "price": "PRICE",
}
METRIC_COLS = ("impression", "like", "collection", "comment",
               "ttl_engagement", "price")


@dataclass
class ReportConfig:
    basis: str = "pooled"                # pooled | per_post
    tier_mode: str = "label"             # label | fanbase
    fanbase_top_k: float = 1000
    fanbase_mid_k: float = 400
    fanbase_bot_k: float = 200           # below → KOC
    concentration_flag_pct: float = 50
    min_group_n: int = 3
    share_decimals: int = 1
    language: str = "en"                 # en | zh
    missing_row_policy: str = "exclude_warn"  # exclude_warn | block


@dataclass
class Row:
    idx: int                 # 0-based data row index (stable unique key)
    excel_row: int
    campaign: str
    no: str
    name: str
    type_raw: str
    level_raw: str
    fanbase_k: Optional[float]
    post_date: Optional[date]
    post_link: str
    impression: Optional[int]
    like: Optional[int]
    collection: Optional[int]
    comment: Optional[int]
    ttl_engagement: Optional[int]
    price: Optional[float]
    cpm_src: Optional[float]
    cpe_src: Optional[float]
    # filled by classification
    coop: str = ""           # PAID | SOFT | ""
    tier: str = ""           # TOP | MID | BOT | KOC | ""
    excluded: bool = False   # dropped from metrics (V2)

    @property
    def key(self) -> tuple[str, str]:
        return (self.campaign, self.no)

    @property
    def group(self) -> str:
        return f"{self.tier} {self.coop}" if self.tier and self.coop else ""


@dataclass
class Finding:
    code: str
    severity: str            # ERROR | WARN
    message: str
    rows: list[int] = field(default_factory=list)   # excel rows

    def to_dict(self) -> dict:
        return asdict(self)


def _to_float(v: Any) -> Optional[float]:
    if v is None or v == "" or isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = nfkc(str(v)).strip().replace(",", "").replace("，", "")
    try:
        return float(s)
    except ValueError:
        return None


# ---------------------------------------------------------------- parsing

def parse_report(path_or_file) -> tuple[list[Row], list[Finding], dict]:
    """Parse the KOL sheet. Returns (rows, findings, meta). V1 failures are
    the only thing that raises — everything else is a Finding."""
    findings: list[Finding] = []
    wb = load_workbook(path_or_file, data_only=True)
    ws = None
    for candidate in wb.worksheets:
        if header_key(candidate.title) == "masterkollist":
            ws = candidate
            break
    if ws is None:
        ws = wb.worksheets[0]
        findings.append(Finding(
            "V1", "WARN",
            f"Sheet 'MASTER KOL LIST' not found — using first sheet "
            f"{ws.title!r}."))

    hit = _find_header_row(ws, {"name", "postlink"})
    if not hit:
        raise ValueError(
            "V1: no header row containing NAME and POST LINK found in "
            f"sheet {ws.title!r}.")
    header_row, cols = hit

    missing = [label for key, label in REQUIRED_COLS.items() if key not in cols]
    if missing:
        raise ValueError(
            "V1: required columns missing after header normalization: "
            + ", ".join(missing))

    def col(key):
        return cols.get(key)

    c = {k: col(k) for k in (
        "no", "mcn", "campaign", "type", "level", "name", "fanbase(k)",
        "postdate", "micromacro", "postlink", "impression", "like",
        "collection", "comment", "ttlengagement", "price", "cpm", "cpe")}

    rows: list[Row] = []
    current_campaign = ""
    blank = 0
    for excel_row_cells in ws.iter_rows(min_row=header_row + 1):
        r = excel_row_cells[0].row
        get = lambda k: ws.cell(row=r, column=c[k]).value if c.get(k) else None
        name = _cell_str(get("name"))
        link_cell = ws.cell(row=r, column=c["postlink"]) if c.get("postlink") else None
        link = ""
        if link_cell is not None:
            if link_cell.hyperlink and link_cell.hyperlink.target:
                link = str(link_cell.hyperlink.target).strip()
            else:
                link = _cell_str(link_cell.value)
        if not name and not link:
            blank += 1
            if blank >= MAX_CONSECUTIVE_BLANK_ROWS:
                break
            continue
        blank = 0
        campaign = _cell_str(get("campaign"))
        if campaign:
            current_campaign = campaign
        rows.append(Row(
            idx=len(rows), excel_row=r,
            campaign=current_campaign, no=_cell_str(get("no")), name=name,
            type_raw=_cell_str(get("type")), level_raw=_cell_str(get("level")),
            fanbase_k=_to_float(get("fanbase(k)")),
            post_date=_to_date(get("postdate")),
            post_link=link.strip(),
            impression=_to_int(get("impression")), like=_to_int(get("like")),
            collection=_to_int(get("collection")),
            comment=_to_int(get("comment")),
            ttl_engagement=_to_int(get("ttlengagement")),
            price=_to_float(get("price")),
            cpm_src=_to_float(get("cpm")), cpe_src=_to_float(get("cpe")),
        ))
    meta = {"sheet": ws.title, "header_row": header_row, "rows": len(rows),
            "campaigns": sorted({r.campaign for r in rows if r.campaign})}
    return rows, findings, meta


# ----------------------------------------------------------- classification

def classify(rows: list[Row], cfg: ReportConfig,
             findings: list[Finding]) -> None:
    """COOP from TYPE (substring — 报备图文/软植图文/bare 报备 all classify);
    TIER from LEVEL (label mode merges 尾部+底部 into BOT) or from fan base
    thresholds (fanbase mode). Unknown values go to the warning bucket —
    never guessed."""
    bad_types: dict[str, list[int]] = {}
    bad_levels: dict[str, list[int]] = {}
    for r in rows:
        t = nfkc(r.type_raw)
        if "报备" in t:
            r.coop = "PAID"
        elif "软植" in t:
            r.coop = "SOFT"
        else:
            r.coop = ""
            bad_types.setdefault(r.type_raw or "(blank)", []).append(r.excel_row)

        if cfg.tier_mode == "fanbase":
            fb = r.fanbase_k
            if fb is None:
                r.tier = ""
                bad_levels.setdefault(
                    f"(no FAN BASE for fanbase tiering; LEVEL={r.level_raw!r})",
                    []).append(r.excel_row)
            elif fb >= cfg.fanbase_top_k:
                r.tier = "TOP"
            elif fb >= cfg.fanbase_mid_k:
                r.tier = "MID"
            elif fb >= cfg.fanbase_bot_k:
                r.tier = "BOT"
            else:
                r.tier = "KOC"
        else:
            lvl = nfkc(r.level_raw)
            if "头部" in lvl:
                r.tier = "TOP"
            elif "腰部" in lvl:
                r.tier = "MID"
            elif "尾部" in lvl or "底部" in lvl:
                r.tier = "BOT"   # deliberate merge — see V8
            elif "koc" in lvl.casefold():
                r.tier = "KOC"
            else:
                r.tier = ""
                bad_levels.setdefault(r.level_raw or "(blank)", []).append(r.excel_row)

    for value, rws in {**bad_types}.items():
        findings.append(Finding(
            "V7", "WARN",
            f"Unclassified TYPE value {value!r} on {len(rws)} row(s) — "
            "excluded from groups, counted in totals.", rws))
    for value, rws in {**bad_levels}.items():
        findings.append(Finding(
            "V7", "WARN",
            f"Unclassified LEVEL value {value!r} on {len(rws)} row(s) — "
            "excluded from groups, counted in totals.", rws))


# --------------------------------------------------------------- validation

def _eng(r: Row) -> Optional[int]:
    return r.ttl_engagement


def validate(rows: list[Row], cfg: ReportConfig,
             findings: list[Finding]) -> None:
    # V2 — missing metric values
    missing_rows = [r for r in rows if any(
        getattr(r, col) is None for col in METRIC_COLS)]
    if missing_rows:
        if cfg.missing_row_policy == "block":
            # ERROR blocks generation (caller checks `blocked`), but the rows
            # are still excluded so the diagnostic metrics stay computable.
            for r in missing_rows:
                r.excluded = True
            findings.append(Finding(
                "V2", "ERROR",
                f"{len(missing_rows)} row(s) missing IMPRESSION/LIKE/"
                "COLLECTION/COMMENT/TTL ENGAGEMENT/PRICE values "
                "(missing_row_policy=block).",
                [r.excel_row for r in missing_rows]))
        else:
            for r in missing_rows:
                r.excluded = True
            findings.append(Finding(
                "V2", "WARN",
                f"{len(missing_rows)} row(s) missing metric values — excluded "
                "from all metrics (missing_row_policy=exclude_warn).",
                [r.excel_row for r in missing_rows]))

    active = [r for r in rows if not r.excluded]

    # V3 — zero denominators
    zero_impr = [r.excel_row for r in active if (r.impression or 0) == 0]
    zero_eng = [r.excel_row for r in active if (_eng(r) or 0) == 0]
    if zero_impr:
        findings.append(Finding(
            "V3", "WARN",
            f"{len(zero_impr)} row(s) with IMPRESSION=0 — excluded from CPM "
            "ratios only.", zero_impr))
    if zero_eng:
        findings.append(Finding(
            "V3", "WARN",
            f"{len(zero_eng)} row(s) with TTL ENGAGEMENT=0 — excluded from "
            "CPE ratios only.", zero_eng))

    # V4 — engagement identity
    bad = [r.excel_row for r in active
           if None not in (r.ttl_engagement, r.like, r.collection, r.comment)
           and r.ttl_engagement != (r.like + r.collection + r.comment)]
    if bad:
        findings.append(Finding(
            "V4", "WARN",
            f"TTL ENGAGEMENT ≠ LIKE+COLLECTION+COMMENT on {len(bad)} row(s).",
            bad))

    # V5 — source CPM/CPE columns vs recomputed (source CPM is PRICE per
    # SINGLE impression, not per mille — never reuse it as standard CPM)
    drift = []
    for r in active:
        if r.cpm_src is not None and r.price is not None and (r.impression or 0) > 0:
            expected = r.price / r.impression
            if expected and abs(r.cpm_src - expected) / expected > 0.02:
                drift.append(r.excel_row)
        if r.cpe_src is not None and r.price is not None and (_eng(r) or 0) > 0:
            expected = r.price / r.ttl_engagement
            if expected and abs(r.cpe_src - expected) / expected > 0.02:
                drift.append(r.excel_row)
    if drift:
        findings.append(Finding(
            "V5", "WARN",
            "Source CPM/CPE drift from recomputation on "
            f"{len(sorted(set(drift)))} row(s). Note the source CPM column is "
            "PRICE÷IMPRESSION — cost per SINGLE impression, ×1000 off the "
            "industry-standard per-mille CPM; this report never reuses it.",
            sorted(set(drift))))

    # V6 — duplicate POST LINK (stripped)
    seen: dict[str, list[Row]] = {}
    for r in rows:
        if r.post_link:
            seen.setdefault(r.post_link, []).append(r)
    for link, rs in seen.items():
        if len(rs) > 1:
            names = " / ".join(x.name for x in rs)
            findings.append(Finding(
                "V6", "WARN",
                f"Duplicate POST LINK shared by {len(rs)} rows ({names}): "
                f"…{link[-24:]} — likely a copy-paste error in the source; "
                "metrics keep both rows, but verify.",
                [x.excel_row for x in rs]))

    # V8 — 尾部/底部 coexistence (documented merge) + fan-range detail
    if cfg.tier_mode == "label":
        weibu = [r for r in rows if "尾部" in nfkc(r.level_raw)]
        dibu = [r for r in rows if "底部" in nfkc(r.level_raw)]
        if weibu and dibu:
            fb = [r.fanbase_k for r in dibu if r.fanbase_k is not None]
            koc_sized = sum(1 for x in fb if x < cfg.fanbase_bot_k)
            wr = [r.fanbase_k for r in weibu if r.fanbase_k is not None]
            findings.append(Finding(
                "V8", "WARN",
                f"Both 尾部 ({len(weibu)} rows, fans "
                f"{min(wr):.0f}–{max(wr):.0f}K) and 底部 ({len(dibu)} rows, "
                f"fans {min(fb):.0f}–{max(fb):.0f}K) labels coexist — merged "
                f"into BOT (documented judgment call). {koc_sized} of the "
                f"{len(dibu)} 底部 accounts have <{cfg.fanbase_bot_k:.0f}K "
                "fans (KOC-sized). Set tier_mode=fanbase to re-tier by "
                "thresholds instead."))


def validate_groups(groups: dict[str, dict], total_rows: int,
                    cfg: ReportConfig, findings: list[Finding]) -> None:
    """V9 (small samples) and V10 (concentration) — need computed metrics."""
    for gname, g in groups.items():
        if 0 < g["n"] < cfg.min_group_n:
            findings.append(Finding(
                "V9", "WARN",
                f"{gname}: n={g['n']} — below min_group_n="
                f"{cfg.min_group_n}; on-slide caveat added (not a benchmark)."))
        conc = g.get("concentration_pct")
        # concentration is trivially 100% for n ≤ 2 — V9 covers those groups
        if g["n"] > 2 and conc is not None and conc > cfg.concentration_flag_pct:
            findings.append(Finding(
                "V10", "WARN",
                f"{gname}: top {g['concentration_top_k']} post(s) hold "
                f"{conc:.0f}% of group impressions — pooled CPM is dragged by "
                f"outliers; plan on per-post ≈{g['cpm_perpost']:.0f}, not "
                f"pooled {g['cpm_pooled']:.0f}. On-slide caveat added."))


# ------------------------------------------------------------------ metrics

def _mean(xs: list[float]) -> Optional[float]:
    return sum(xs) / len(xs) if xs else None


def compute_metrics(rows: list[Row], cfg: ReportConfig) -> dict:
    """Primary implementation: raw-loop accumulation. Group key 'TIER COOP'.
    Empty groups are absent (rendered as null bars, never zeros)."""
    active = [r for r in rows if not r.excluded]
    total_rows = len(active)
    groups: dict[str, dict] = {}
    for tier in TIERS:
        for coop in COOPS:
            g = [r for r in active if r.tier == tier and r.coop == coop]
            if not g:
                continue
            prices = [r.price for r in g]
            impr_rows = [r for r in g if (r.impression or 0) > 0]
            eng_rows = [r for r in g if (_eng(r) or 0) > 0]
            spend_impr = sum(r.price for r in impr_rows)
            impr = sum(r.impression for r in impr_rows)
            spend_eng = sum(r.price for r in eng_rows)
            eng = sum(r.ttl_engagement for r in eng_rows)
            top2 = sorted((r.impression or 0 for r in g), reverse=True)[:2]
            conc = (sum(top2) / impr * 100) if impr else None
            groups[f"{tier} {coop}"] = {
                "tier": tier, "coop": coop,
                "n": len(g),
                "share": round(len(g) / total_rows * 100, cfg.share_decimals),
                "avg_price": _mean(prices),
                "cpm_pooled": spend_impr / impr * 1000 if impr else None,
                "cpe_pooled": spend_eng / eng if eng else None,
                "cpm_perpost": _mean([r.price / r.impression * 1000
                                      for r in impr_rows]),
                "cpe_perpost": _mean([r.price / r.ttl_engagement
                                      for r in eng_rows]),
                "spend": sum(prices),
                "impressions": sum(r.impression or 0 for r in g),
                "engagements": sum(_eng(r) or 0 for r in g),
                "concentration_pct": conc,
                "concentration_top_k": min(2, len(g)),
            }
    unclassified = [r for r in active if not r.group]
    totals = {
        "rows": total_rows,
        "classified": total_rows - len(unclassified),
        "unclassified": len(unclassified),
        "unclassified_share": round(
            len(unclassified) / total_rows * 100, cfg.share_decimals)
        if total_rows else 0.0,
        "spend": sum(r.price for r in active),
        "impressions": sum(r.impression or 0 for r in active),
        "engagements": sum(_eng(r) or 0 for r in active),
        "excluded_rows": len(rows) - total_rows,
    }
    return {"groups": groups, "totals": totals}


def compute_metrics_pandas(rows: list[Row], cfg: ReportConfig) -> dict:
    """Independent second path (pandas groupby) for the dual-path check."""
    import pandas as pd
    active = [r for r in rows if not r.excluded]
    df = pd.DataFrame([{
        "group": r.group, "price": r.price, "impression": r.impression or 0,
        "engagement": _eng(r) or 0} for r in active])
    out: dict[str, dict] = {}
    if df.empty:
        return out
    for gname, g in df[df["group"] != ""].groupby("group"):
        gi = g[g["impression"] > 0]
        ge = g[g["engagement"] > 0]
        out[gname] = {
            "n": int(len(g)),
            "share": round(len(g) / len(df) * 100, cfg.share_decimals),
            "avg_price": float(g["price"].mean()),
            "cpm_pooled": float(gi["price"].sum() / gi["impression"].sum() * 1000)
            if len(gi) else None,
            "cpe_pooled": float(ge["price"].sum() / ge["engagement"].sum())
            if len(ge) else None,
            "cpm_perpost": float((gi["price"] / gi["impression"] * 1000).mean())
            if len(gi) else None,
            "cpe_perpost": float((ge["price"] / ge["engagement"]).mean())
            if len(ge) else None,
        }
    return out


class VerificationError(RuntimeError):
    """An internal cross-check failed — never render output past this."""


def verify_dual_path(primary: dict, secondary: dict) -> None:
    keys = ("n", "share", "avg_price", "cpm_pooled", "cpe_pooled",
            "cpm_perpost", "cpe_perpost")
    if set(primary["groups"]) != set(secondary):
        raise VerificationError(
            f"dual-path group mismatch: {sorted(primary['groups'])} vs "
            f"{sorted(secondary)}")
    for gname, g in primary["groups"].items():
        for k in keys:
            a, b = g[k], secondary[gname][k]
            if a is None and b is None:
                continue
            if a is None or b is None or abs(a - b) > 1e-6:
                raise VerificationError(
                    f"dual-path mismatch {gname}.{k}: {a} vs {b}")


def verify_reconciliation(rows: list[Row], metrics: dict) -> None:
    groups, totals = metrics["groups"], metrics["totals"]
    n_sum = sum(g["n"] for g in groups.values()) + totals["unclassified"]
    if n_sum != totals["rows"]:
        raise VerificationError(
            f"Σ group n ({n_sum}) != total rows ({totals['rows']})")
    spend_sum = sum(g["spend"] for g in groups.values()) + sum(
        r.price for r in rows if not r.excluded and not r.group)
    if abs(spend_sum - totals["spend"]) > 1e-6:
        raise VerificationError(
            f"Σ group spend ({spend_sum}) != total spend ({totals['spend']})")
    share_sum = sum(g["share"] for g in groups.values()) + totals["unclassified_share"]
    if abs(share_sum - 100.0) > 0.3:
        raise VerificationError(
            f"donut shares sum to {share_sum} at display precision")


# ----------------------------------------------------------------- insights

TEXTS = {
    "en": {
        "donut_title": "DETAIL SHARE OF KOL",
        "price_title": "AVG COLLAB PRICE",
        "eff_title": "KOL SOFT VS PAID EFFICIENCY COMPARISON (CNY)",
        "cpm": "CPM", "cpe": "CPE",
        "deck_title": "KOL EFFICIENCY REPORT",
    },
    "zh": {
        "donut_title": "达人构成占比",
        "price_title": "平均合作价格",
        "eff_title": "软植 VS 报备 投放效率对比 (CNY)",
        "cpm": "CPM", "cpe": "CPE",
        "deck_title": "KOL 投放效率报告",
    },
}


def _fmt_k(v: float) -> str:
    return f"{v / 1000:.1f}K"


def build_insights(metrics: dict, cfg: ReportConfig,
                   findings: list[Finding]) -> dict:
    """Data-driven bullets only. Cross-wave comparisons are NEVER generated —
    there is no prior-wave input, and fabricated wave deltas are the most
    damaging failure mode of this feature."""
    groups = metrics["groups"]
    basis = cfg.basis
    cpm_key = "cpm_pooled" if basis == "pooled" else "cpm_perpost"
    cpe_key = "cpe_pooled" if basis == "pooled" else "cpe_perpost"

    price_bullets: list[str] = []
    premiums, inversions = [], []
    for tier in TIERS:
        paid, soft = groups.get(f"{tier} PAID"), groups.get(f"{tier} SOFT")
        if not (paid and soft):
            continue
        delta = paid["avg_price"] - soft["avg_price"]
        if delta >= 0:
            premiums.append(f"{tier} +{_fmt_k(delta)}")
        else:
            inversions.append(f"{tier} {_fmt_k(delta)}")
    if premiums and not inversions:
        price_bullets.append(
            "PAID CARRIES A PREMIUM IN EVERY TIER — "
            + " / ".join(premiums) + " VS SOFT")
    elif premiums or inversions:
        if premiums:
            price_bullets.append("PAID PREMIUM: " + " / ".join(premiums) + " VS SOFT")
        if inversions:
            price_bullets.append(
                "PRICE INVERSION — SOFT PRICED ABOVE PAID: " + " / ".join(inversions))

    eff_bullets: list[str] = []
    cpm_parts, cpe_parts = [], []
    for tier in TIERS:
        paid, soft = groups.get(f"{tier} PAID"), groups.get(f"{tier} SOFT")
        if not (paid and soft) or paid[cpm_key] is None or soft[cpm_key] is None:
            pass
        else:
            winner = "PAID" if paid[cpm_key] <= soft[cpm_key] else "SOFT"
            w, l = sorted((paid[cpm_key], soft[cpm_key]))
            cpm_parts.append(f"{tier}: {winner} ¥{w:.0f} VS ¥{l:.0f}")
        if not (paid and soft) or paid[cpe_key] is None or soft[cpe_key] is None:
            continue
        winner = "PAID" if paid[cpe_key] <= soft[cpe_key] else "SOFT"
        w, l = sorted((paid[cpe_key], soft[cpe_key]))
        cpe_parts.append(f"{tier}: {winner} ¥{w:.1f} VS ¥{l:.1f}")
    if cpm_parts:
        eff_bullets.append("CPM WINNER — " + " · ".join(cpm_parts))
    if cpe_parts:
        eff_bullets.append("CPE WINNER — " + " · ".join(cpe_parts))

    caveats: list[str] = []
    for f in findings:
        if f.code == "V9":
            gname = f.message.split(":")[0]
            n = groups[gname]["n"] if gname in groups else "?"
            caveats.append(f"{gname} = {n} POST(S) ONLY — NOT A BENCHMARK")
        if f.code == "V10":
            gname = f.message.split(":")[0]
            g = groups.get(gname)
            if g:
                caveats.append(
                    f"CAUTION: {gname} CARRIED BY {g['concentration_top_k']} "
                    f"VIRAL POSTS ({g['concentration_pct']:.0f}% OF GROUP "
                    f"IMPRESSIONS) — PLAN ON ~¥{g['cpm_perpost']:.0f} CPM, "
                    f"NOT ¥{g['cpm_pooled']:.0f}")

    # pooled↔per-post gap diagnostic (feeds the same caution family)
    for gname, g in groups.items():
        if (g["cpm_pooled"] and g["cpm_perpost"]
                and g["cpm_perpost"] / g["cpm_pooled"] > 3
                and not any(gname in c for c in caveats)):
            caveats.append(
                f"{gname}: PER-POST CPM ¥{g['cpm_perpost']:.0f} IS "
                f">3× POOLED ¥{g['cpm_pooled']:.0f} — RESULTS CONCENTRATED "
                "IN OUTLIERS")

    tier_n = {t: sum(g["n"] for g in groups.values() if g["tier"] == t)
              for t in TIERS}
    if basis == "pooled":
        footnote = (
            "Basis: pooled — group total spend ÷ total impressions "
            "(CPM, ¥ per 1,000) / total engagements (CPE). "
            f"n = {metrics['totals']['rows']} posts: "
            + " · ".join(f"{t} {tier_n[t]}" for t in TIERS if tier_n[t]))
    else:
        footnote = (
            "Basis: per-post average — mean of PRICE÷IMPRESSION×1000 (CPM, "
            "¥ per 1,000) / PRICE÷ENGAGEMENT (CPE) across posts. "
            f"n = {metrics['totals']['rows']} posts: "
            + " · ".join(f"{t} {tier_n[t]}" for t in TIERS if tier_n[t]))

    return {"price": price_bullets, "efficiency": eff_bullets,
            "caveats": caveats, "footnote": footnote}


# ---------------------------------------------------------------- pipeline

def analyze(path_or_file, cfg: Optional[ReportConfig] = None) -> dict:
    """Full analysis: parse → classify → validate → metrics → verify →
    insights. Raises ValueError on V1, VerificationError on failed internal
    cross-checks; ERROR findings block downstream generation (caller checks).
    """
    cfg = cfg or ReportConfig()
    rows, findings, meta = parse_report(path_or_file)
    classify(rows, cfg, findings)
    validate(rows, cfg, findings)

    blocked = any(f.severity == "ERROR" for f in findings)
    metrics = compute_metrics(rows, cfg)
    if not blocked:
        verify_dual_path(metrics, compute_metrics_pandas(rows, cfg))
        verify_reconciliation(rows, metrics)
    validate_groups(metrics["groups"], metrics["totals"]["rows"], cfg, findings)
    insights = build_insights(metrics, cfg, findings)
    return {
        "meta": meta, "config": asdict(cfg),
        "findings": [f.to_dict() for f in findings],
        "blocked": blocked,
        "metrics": metrics, "insights": insights,
        "rows_total": len(rows),
    }
