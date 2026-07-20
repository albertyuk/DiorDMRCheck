"""LVMH social perimeter (Micro) — offline cross-check for NO_BLOGGER rows.

Parses the ``List Micro`` sheet only (~58.8k rows), locating the header by
fingerprint (a row containing both NAME and REDBOOK_ID) and the extraction
date from the metadata rows above it. Parsing a file this size is slow, so
the parsed rows — with all Tier-0 normalized forms precomputed — are cached
in SQLite keyed by the file's content hash; warm loads take well under 2 s.

REDBOOK_ID is the XHS user id (same key space as the DMR Username column and
the resolver's author_id), treated as an opaque string with case-insensitive
lookup. Most perimeter rows have no REDBOOK_ID (other-platform bloggers).
"""
from __future__ import annotations

import hashlib
import json
import re
import time
from dataclasses import dataclass, field
from typing import IO, Optional, Union

from openpyxl import load_workbook
from pypinyin import lazy_pinyin
from rapidfuzz import fuzz, process

from . import db
from .normalize import ascii_part, cjk, header_key, norm

PERIMETER_REQUIRED = {"name", "redbook_id"}
HEADER_SCAN_ROWS = 15
FUZZY_CUTOFF = 85

_EXTRACTION_RE = re.compile(
    r"date of extraction\s*[:：]?\s*([0-9]{2}/[0-9]{2}/[0-9]{4}(?:\s+[0-9:]{5,8})?)",
    re.IGNORECASE)


def _s(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _i(v) -> Optional[int]:
    try:
        if v is None or v == "":
            return None
        return int(float(str(v).replace(",", "")))
    except (ValueError, TypeError):
        return None


@dataclass
class PerimeterParse:
    file_hash: str
    filename: str
    sheet: str
    header_row: int
    extraction_date: str
    rows: list[dict] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)

    @property
    def redbook_count(self) -> int:
        return sum(1 for r in self.rows if r["redbook_id"])


def file_hash(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def _pick_micro_sheet(wb):
    for name in wb.sheetnames:
        if header_key(name) == "listmicro":
            return wb[name]
    for name in wb.sheetnames:
        k = header_key(name)
        if "micro" in k and "macro" not in k:
            return wb[name]
    return None


def parse_perimeter(source: Union[str, IO[bytes]], filename: str = "",
                    content_hash: str = "") -> PerimeterParse:
    """Parse the Micro perimeter workbook (read-only mode — the file is big)."""
    wb = load_workbook(source, read_only=True, data_only=True)
    try:
        ws = _pick_micro_sheet(wb)
        if ws is None:
            raise ValueError(
                "Perimeter parse failed: no 'List Micro' sheet found "
                f"(sheets: {wb.sheetnames})")

        header_row = None
        cols: dict[str, int] = {}
        meta_cells: list[str] = []
        for r_idx, row in enumerate(
                ws.iter_rows(min_row=1, max_row=HEADER_SCAN_ROWS,
                             values_only=True), 1):
            keys = {header_key(_s(v)): c for c, v in enumerate(row) if _s(v)}
            if PERIMETER_REQUIRED.issubset(keys.keys()):
                header_row = r_idx
                cols = keys
                break
            meta_cells.extend(_s(v) for v in row if _s(v))
        if header_row is None:
            raise ValueError(
                "Perimeter parse failed: no header row containing both 'NAME' "
                f"and 'REDBOOK_ID' within the first {HEADER_SCAN_ROWS} rows of "
                f"sheet {ws.title!r}.")

        result = PerimeterParse(
            file_hash=content_hash, filename=filename, sheet=ws.title,
            header_row=header_row, extraction_date="")
        m = _EXTRACTION_RE.search(" | ".join(meta_cells))
        if m:
            result.extraction_date = m.group(1).strip()
        else:
            result.warnings.append(
                "Could not find 'Date of extraction' in the metadata rows — "
                "perimeter staleness cannot be shown.")

        c_name, c_namebis = cols.get("name"), cols.get("namebis")
        c_dmrid, c_rid = cols.get("dmrid"), cols.get("redbook_id")
        c_rfol = cols.get("redbook_followers")

        def col(row, c):
            return _s(row[c]) if c is not None and c < len(row) else ""

        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            name = col(row, c_name)
            namebis = col(row, c_namebis)
            rid = col(row, c_rid).lower()
            if not name and not namebis:
                continue
            # Tier-0 normalized forms are precomputed here so the cached JSON
            # can be scanned without re-normalizing 58.8k names per run.
            result.rows.append({
                "name": name, "namebis": namebis,
                "dmrid": col(row, c_dmrid),
                "redbook_id": rid,
                "redbook_followers": _i(row[c_rfol]) if c_rfol is not None and c_rfol < len(row) else None,
                "nn": norm(name), "an": ascii_part(name),
                "nb": norm(namebis), "cb": cjk(namebis),
                "ab": ascii_part(namebis),
            })
        if not result.rows:
            result.warnings.append("Perimeter sheet parsed but had no data rows.")
        return result
    finally:
        wb.close()


# ------------------------------------------------------------------ caching

def store_parsed(parsed: PerimeterParse) -> None:
    db.perimeter_cache_put(
        parsed.file_hash, filename=parsed.filename, sheet=parsed.sheet,
        extraction_date=parsed.extraction_date, row_count=len(parsed.rows),
        redbook_count=parsed.redbook_count,
        parsed_json=json.dumps(parsed.rows, ensure_ascii=False),
    )


def load_cached(file_hash: str) -> Optional["PerimeterIndex"]:
    row = db.perimeter_cache_get(file_hash)
    if not row:
        return None
    rows = json.loads(row["parsed_json"])
    return PerimeterIndex(rows, extraction_date=row["extraction_date"] or "",
                          filename=row["filename"] or "",
                          file_hash=file_hash)


# ------------------------------------------------------------------ indexes

class PerimeterIndex:
    """Lookup structures over the parsed Micro perimeter.

    ``by_redbook`` is the ground-truth join (same key space as Tier 2's
    author ids). ``scan_name`` reuses the Tier-0/Tier-3 name-ladder semantics
    against NAME + NAMEBIS and returns *all* hits — name collisions are real
    (e.g. 'esther' matches several rows), so callers must never auto-pick."""

    def __init__(self, rows: list[dict], extraction_date: str = "",
                 filename: str = "", file_hash: str = ""):
        self.rows = rows
        self.extraction_date = extraction_date
        self.filename = filename
        self.file_hash = file_hash
        self.by_redbook: dict[str, dict] = {}
        for r in rows:
            rid = (r.get("redbook_id") or "").lower()
            if rid and rid not in self.by_redbook:
                self.by_redbook[rid] = r
        # fuzzy choice lists (index-aligned with self.rows; short forms
        # excluded — partial_ratio saturates on 1-3 char strings)
        self._an = [(i, r["an"]) for i, r in enumerate(rows) if len(r["an"]) >= 4]
        self._ab = [(i, r["ab"]) for i, r in enumerate(rows) if len(r["ab"]) >= 4]

    def lookup_author(self, author_id: str) -> Optional[dict]:
        return self.by_redbook.get((author_id or "").strip().lower())

    def _fuzzy(self, query: str, choices: list[tuple[int, str]],
               hits: dict[int, str], method: str) -> None:
        if len(query) < 4 or not choices:
            return
        found = process.extract(
            query, [c[1] for c in choices], scorer=fuzz.partial_ratio,
            score_cutoff=FUZZY_CUTOFF, limit=None)
        for _choice, _score, pos in found:
            idx = choices[pos][0]
            hits.setdefault(idx, method)

    def scan_name(self, plog_name: str) -> list[tuple[dict, str]]:
        pc, pn = cjk(plog_name), norm(plog_name)
        pa = ascii_part(plog_name)
        hits: dict[int, str] = {}
        # ladder steps a/b — exact containment over precomputed forms
        for i, r in enumerate(self.rows):
            if pc and r["cb"] and pc in r["cb"]:
                hits.setdefault(i, "cjk-substring")
            elif pn and ((r["nn"] and pn in r["nn"]) or (r["nb"] and pn in r["nb"])):
                hits.setdefault(i, "norm-substring")
        # ladder steps c/d — fuzzy over the ASCII forms (rapidfuzz batch)
        self._fuzzy(pa, self._an, hits, "ascii-fuzzy")
        self._fuzzy(pa, self._ab, hits, "ascii-fuzzy")
        if pc:
            pinyin = "".join(lazy_pinyin(pc)).casefold()
            self._fuzzy(pinyin, self._an, hits, "pinyin-bridge")
            self._fuzzy(pinyin, self._ab, hits, "pinyin-bridge")
        return [(self.rows[i], m) for i, m in sorted(hits.items())]


def ingest(data: bytes, filename: str) -> PerimeterParse:
    """Parse-or-load a perimeter upload, cache it, and make it current."""
    import io
    h = file_hash(data)
    cached = db.perimeter_cache_get(h)
    if cached:
        parsed = PerimeterParse(
            file_hash=h, filename=filename, sheet=cached["sheet"] or "",
            header_row=0, extraction_date=cached["extraction_date"] or "")
        meta_rows, meta_redbook = cached["row_count"], cached["redbook_count"]
    else:
        parsed = parse_perimeter(io.BytesIO(data), filename=filename,
                                 content_hash=h)
        store_parsed(parsed)
        meta_rows, meta_redbook = len(parsed.rows), parsed.redbook_count
    db.setting_set("current_perimeter", json.dumps({
        "hash": h, "filename": filename,
        "extraction_date": parsed.extraction_date,
        "rows": meta_rows, "redbook_count": meta_redbook,
        "uploaded_at": time.time(),
    }, ensure_ascii=False))
    return parsed


def current_meta() -> Optional[dict]:
    raw = db.setting_get("current_perimeter")
    if not raw:
        return None
    try:
        return json.loads(raw)
    except ValueError:
        return None
